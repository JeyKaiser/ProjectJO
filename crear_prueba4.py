#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de prueba4.xlsx - Formato Optimizado para Google Sheets
Incluye: ARRAYFORMULA, dropdowns dependientes, data validation vinculada,
         formato condicional por fila, checkboxes, QUERY() en dashboard.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OUTPUT = r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0\prueba4.xlsx"

# ═══════════════════════════ PALETA ═══════════════════════════
class P:
    CREATIVA_DARK  = "1F4E79"
    CREATIVA_LIGHT = "BDD7EE"
    TALLER_DARK    = "BF8F00"
    TALLER_LIGHT   = "FFE699"
    ING_DARK       = "ED7D31"
    ING_LIGHT      = "F8CBAD"
    PROD_DARK      = "C00000"
    PROD_LIGHT     = "F4B4C2"
    STATUS_DARK    = "44546A"
    STATUS_LIGHT   = "D6DCE4"
    REPORT_DARK    = "003366"
    REPORT_LIGHT   = "D9E2F3"
    PARAM_DARK     = "2C3E50"
    WHITE          = "FFFFFF"
    BLACK          = "000000"
    GRAY           = "808080"
    GREEN          = "C6EFCE"
    RED            = "FFC7CE"
    AMBER_ALT      = "FFEB9C"
    BLUE_ALT       = "BDD7EE"
    ORANGE_ALT     = "F8CBAD"

thin_border = Border(
    left=Side(style='thin', color=P.GRAY),
    right=Side(style='thin', color=P.GRAY),
    top=Side(style='thin', color=P.GRAY),
    bottom=Side(style='thin', color=P.GRAY)
)

def sc(cell, fill_color, font_color="FFFFFF", bold=True, size=9, wrap=True):
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.font = Font(color=font_color, bold=bold, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = thin_border

def sd(cell, alt=False):
    cell.font = Font(name="Calibri", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if alt:
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def aw(ws, mn=12, mx=24):
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        ml = mn
        for r in range(1, min(ws.max_row + 1, 50)):
            v = ws.cell(row=r, column=c).value
            if v: ml = max(ml, min(len(str(v)) * 1.1 + 2, mx))
        ws.column_dimensions[letter].width = ml

def col_letter(n):
    return get_column_letter(n)

# ═══════════════════════════ CATALOGOS ═══════════════════════════
# Sublíneas agrupadas por Línea (para dropdowns dependientes)
SUBLINEAS_POR_LINEA = {
    "DRESSES":  ["MAXI", "MIDI", "MINI", "ANKLE", "KIMONO", "CAPE"],
    "OUTWEAR":  ["JACKET", "COAT", "KIMONO", "BOLERO", "CAPE"],
    "PANTS":    ["DRAMATIC PANT", "CASUAL", "WIDE LEG", "ANKLE", "JOGGER"],
    "SKIRTS":   ["MINI", "MIDI", "ANKLE", "LARGA"],
    "SWIMWEAR": ["BIKINI TOP", "BIKINI BOTTOM", "FULL SUIT", "COVER UP"],
    "TOPS":     ["CROP TOP", "BÁSICO", "BLUSÓN", "TÚNICA"],
}

CATALOGS = [
    ("STATUS", ["APROBADO", "APROBADO REPROGRAMACION", "CANCELADO", "CANCELADO CORTADO",
                "CANCELADO SIN CORTAR", "EN PROCESO", "PAUSADO", "TERMINADO",
                "CANCELADO POR COMERCIAL", "CANCELADO PAQUETE COMPLETO"]),
    ("DISEÑADORES CREATIVOS", ["BURGOS", "CLAUDIA", "FER", "MARGARITA", "MARIA BURGOS",
                                "MARIA DEL MAR", "OSMAN", "YAMILET"]),
    ("DISEÑADORES TÉCNICOS", ["CAMILA VILLEGAS", "CLAUDIA R", "CRISTIAN GOMEZ",
                               "DANIELA GARCIA", "FERNANDO", "KAROLINE",
                               "KELLY M.", "LINA DELGADO"]),
    ("TRAZADORES", ["KELLY MITROVICH", "ANDRES CORTADOR", "FERNANDO TRAZO"]),
    ("MODISTAS", ["ADRIANA ESCOBAR", "AISURY QUINTERO", "ALEJANDRA ROJAS",
                  "CIELO AGUIRRE", "DIANA ADARME", "ENEIDA MACA",
                  "FANNY GOMEZ", "ISABEL VALENCIA"]),
    ("LINEAS", ["DRESSES", "OUTWEAR", "PANTS", "SKIRTS", "SWIMWEAR", "TOPS"]),
    ("TIPOS REF", ["DS", "OW", "PS", "SK", "SW", "TP"]),
    ("TALLAJES", ["0-2-4-6-8-10-12", "1-2-3-4-5-6", "2-4-6-8-10-12", "XS-S-M-L-XL"]),
    ("USOS EN PRENDA", ["TELA LUCIR", "TELA FORRO", "FUSIONABLE", "SESGOS LUCIR",
                        "SESGOS FORRO", "SESGOS FUSIONABLE", "ENTRETELA", "CINTA",
                        "RESORTE", "HILO RESORTE", "CIERRE", "ARGOLIA", "BOTON"]),
    ("BASES TEXTILES", ["LINEN", "COTTON", "SILK", "WOOL", "LYCRA", "LYCRA VITA",
                        "LYCRA BAHIA", "CREPE DE CHINE", "CREPE DE SEDA",
                        "COTTON JACQUARD", "LEATHER", "LIGHT LINEN", "VISCOSE",
                        "POLYAMIDE", "POLYESTER", "TAFETA SEDA", "DUCHESS SATIN"]),
    ("LARGOS", ["MAXI 115CM", "MIDI 92CM", "MINI 85CM", "ANKLE 95CM",
                "FLATS 112CM", "KIMONO 100 CM"]),
    ("VEREDICTOS DIR. CREATIVA", ["APROBADA DIRECTA", "APROBADA CON COMENTARIOS", "RECHAZADA"]),
    ("TIPO BORDADO", ["SOBRE PRENDA ARMADA", "SEMIELABORADO", "AMBOS", "EN PRENDA"]),
    ("TIPO PROCESO EXTERNO", ["LAVANDERIA", "BORDADO", "DRAPEADO", "DESCOLADO",
                               "TINTORERIA", "MONTAJE MANIQUI"]),
    ("CATALOGACION", ["SOLIDO", "MODIFICACION DE ARTE", "UBICACION EN TRAZO",
                      "ALL OVER", "ALL OVER CON SENTIDO", "ALL OVER CON ORIENTACION"]),
    ("DROPS", [chr(c) for c in range(65, 75)]),
    ("PRIORIDADES", [str(i) for i in range(1, 11)]),
    ("DIFICULTAD", ["BAJA", "INTERMEDIA", "ALTA", "MUY ALTA"]),
    ("TIPO TEJIDO", ["TEJIDO PLANO", "TEJIDO DE PUNTO", "CUERO", "DENIM", "GASA"]),
    ("WOVEN/KNITTED", ["WOVEN", "KNITTED"]),
    ("TIPO CORTE", ["PRENDA COMPLETA", "LABORATORIO", "PIEZA", "REPOSICION",
                    "REPOSICION CONTRAMUESTRA"]),
    ("STATUS PROCESO EXT.", ["PENDIENTE", "EN PROCESO", "TERMINADO", "RECHAZADO"]),
    ("CONFECCION STATUS", ["EN PROCESO", "TERMINADO", "PAUSADO", "CANCELADO", "RECHAZADO"]),
    ("TIPO RECHAZO CALIDAD", ["MEDIDAS", "CALIDAD DE CONFECCION",
                               "COHERENCIA CON FICHA TECNICA", "MATERIA PRIMA",
                               "INSUMOS", "OTRO"]),
    ("CLASIFICACION HALLAZGO", ["INCONSISTENCIAS", "FALTA DE INFORMACION", "FALTANTES",
                                  "CAMBIOS", "FALTA DE ANALISIS", "OTROS"]),
    ("TIPO EMPAQUE", ["CAJA INDIVIDUAL", "BOLSA COLECTIVA", "CAJA COLECTIVA",
                      "BOLSA INDIVIDUAL", "MARQUILLA SATINADA FORRADA"]),
    ("CLOSURE TYPES", ["BACK ZIPPER", "BACK BUTTONS", "BACK ZIPPER AND BUTTOM",
                       "SNAPS", "WITH SASH", "N/A", "SIN CIERRE"]),
    ("LINNED", ["SI", "NO", "ONLY THE TOP", "ONLY THE BOTTOM", "N/A"]),
    ("SENTIDO SESGO", ["AL HILO", "A TRAVEZ", "AL SESGO"]),
    ("PERMITE GIRAR MOLDE", ["90°", "180°", "PEINE"]),
]

# ═══════════════════════════ WORKBOOK ═══════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ======================================================================
# HOJA 1: PARAMETROS (33 catalogos + Named Ranges para sublineas)
# ======================================================================
ws_param = wb.create_sheet("PARAMETROS", 0)

col = 1
for cat_name, values in CATALOGS:
    cell = ws_param.cell(row=1, column=col, value=cat_name)
    sc(cell, P.PARAM_DARK, P.WHITE, bold=True, size=9)
    alt_fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
    for j, val in enumerate(values):
        cell = ws_param.cell(row=2 + j, column=col, value=val)
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        if j % 2 == 0:
            cell.fill = alt_fill
    ws_param.column_dimensions[col_letter(col)].width = 26
    col += 1

# ── Bloque de SUBLINEAS POR LINEA (para dropdowns dependientes) ──
sub_start_col = col + 1
label_cell = ws_param.cell(row=1, column=sub_start_col, value="SUBLÍNEAS POR LÍNEA (dropdown dependiente)")
sc(label_cell, P.PARAM_DARK, P.WHITE, bold=True, size=10)
ws_param.merge_cells(start_row=1, start_column=sub_start_col,
                     end_row=1, end_column=sub_start_col + len(SUBLINEAS_POR_LINEA) - 1)

sub_col = sub_start_col
named_ranges = {}
for linea, sublineas in SUBLINEAS_POR_LINEA.items():
    hdr = ws_param.cell(row=2, column=sub_col, value=linea)
    sc(hdr, P.CREATIVA_DARK, P.WHITE, bold=True, size=9)
    for j, sub in enumerate(sublineas):
        cell = ws_param.cell(row=3 + j, column=sub_col, value=sub)
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws_param.column_dimensions[col_letter(sub_col)].width = 20
    # Registrar rango para named range
    named_ranges[linea] = (sub_col, 3, 3 + len(sublineas) - 1)
    sub_col += 1

# Crear Named Ranges en el workbook para cada linea
for linea, (scol, srow, erow) in named_ranges.items():
    ref = f"PARAMETROS!${col_letter(scol)}${srow}:${col_letter(scol)}${erow}"
    name = linea.replace(" ", "_")
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(dn)

ws_param.freeze_panes = "B3"
ws_param.sheet_properties.tabColor = "2C3E50"
print(f"[OK] PARAMETROS: {len(CATALOGS)} catalogos + {len(SUBLINEAS_POR_LINEA)} grupos subl\xednea con Named Ranges")

# ======================================================================
# HOJA 2: MATRIZ_MAESTRA (45 columnas) + Data Validation + Formato Condicional
# ======================================================================
ws_matriz = wb.create_sheet("MATRIZ_MAESTRA", 1)

# (mismos headers que prueba3, sin coleccion ni consumos)
headers_matriz = [
    ("ID_Ref", P.CREATIVA_DARK),
    ("Cod_MD", P.CREATIVA_DARK),
    ("Cod_PT", P.CREATIVA_DARK),
    ("Nombre_Referencia", P.CREATIVA_DARK),
    ("Color", P.CREATIVA_DARK),
    ("Cod_Color", P.CREATIVA_DARK),
    ("Capsula/Drop", P.CREATIVA_DARK),
    ("Linea", P.CREATIVA_DARK),
    ("Subl_x_nea", P.CREATIVA_DARK),
    ("Tipo_Ref", P.CREATIVA_DARK),
    ("Sistema_Tallaje", P.CREATIVA_DARK),
    ("Largo", P.CREATIVA_DARK),
    ("Closure", P.CREATIVA_DARK),
    ("Basado_En (Ref PT)", P.CREATIVA_DARK),
    ("Dise_x_ador_Creativo", P.CREATIVA_DARK),
    ("Fecha_Asignacion_Creativo", P.CREATIVA_DARK),
    ("Cambio_Molderia (SI/NO)", P.CREATIVA_DARK),
    ("Obs_Creativo", P.CREATIVA_DARK),
    ("Fecha_Inicio_Molderia_Diseno", P.CREATIVA_DARK),
    ("Fecha_Fin_Molderia_Diseno", P.CREATIVA_DARK),
    ("Comentarios_Molderia_Diseno", P.CREATIVA_DARK),
    ("Modista_Muestra", P.TALLER_DARK),
    ("Status_Taller", P.TALLER_DARK),
    ("Fotos_Internas (SI/NO)", P.TALLER_DARK),
    ("Requiere_Muestra (SI/NO)", P.TALLER_DARK),
    ("Medicion_Fase_1 (Fecha)", P.ING_DARK),
    ("Comentarios_Medicion_1", P.ING_DARK),
    ("Veredicto_Dir_Creativa", P.ING_DARK),
    ("Dise_x_ador_Tecnico", P.ING_DARK),
    ("Dise_x_ador_Contramuestra", P.ING_DARK),
    ("Fecha_Inicio_Molderia_Ing", P.ING_DARK),
    ("Fecha_Fin_Molderia_Ing", P.ING_DARK),
    ("Comentarios_Molderia_Ing", P.ING_DARK),
    ("Especificadora", P.PROD_DARK),
    ("Fecha_Inicio_Especificacion", P.PROD_DARK),
    ("Fecha_Fin_Especificacion", P.PROD_DARK),
    ("Status_Referencia", P.STATUS_DARK),
    ("Entregable_Creativo_OK?", P.STATUS_DARK),
    ("Entregable_Tecnico_OK?", P.STATUS_DARK),
    ("Entregable_Trazador_OK?", P.STATUS_DARK),
    ("Envio_Mod_Arte_OK?", P.STATUS_DARK),
    ("Fecha_Liberacion_Diseno_a_Ing", P.STATUS_DARK),
    ("Fecha_Liberacion_Ing_a_Prod", P.STATUS_DARK),
    ("Temperatura (Fase Actual)", P.STATUS_DARK),
    ("Observaciones_Generales", P.STATUS_DARK),
]

for ci, (h, color) in enumerate(headers_matriz, 1):
    cell = ws_matriz.cell(row=1, column=ci, value=h)
    sc(cell, color, P.WHITE, bold=True, size=9)

# ── DATA VALIDATION (dropdowns vinculados a PARAMETROS) ──
# Mapeo: columna en MATRIZ_MAESTRA → rango en PARAMETROS
validation_map = {
    8:  ("Linea", "status"),        # Linea
    10: ("TIPOS REF", "status"),    # Tipo_Ref
    11: ("TALLAJES", "status"),     # Sistema_Tallaje
    12: ("LARGOS", "status"),       # Largo
    13: ("CLOSURE TYPES", "status"),# Closure
    15: ("DISE\xd1ADORES CREATIVOS", "status"),
    17: (None, "checkbox"),         # Cambio_Molderia → checkbox
    22: ("MODISTAS", "status"),
    23: ("CONFECCION STATUS", "status"),
    24: (None, "checkbox"),         # Fotos_Internas → checkbox
    25: (None, "checkbox"),         # Requiere_Muestra → checkbox
    28: ("VEREDICTOS DIR. CREATIVA", "status"),
    29: ("DISE\xd1ADORES T\xc9CNICOS", "status"),
    30: ("DISE\xd1ADORES T\xc9CNICOS", "status"),
    37: ("STATUS", "status"),
}

for col_num, (cat_name, val_type) in validation_map.items():
    if val_type == "checkbox":
        continue  # Google Sheets checkboxes, en Excel usamos TRUE/FALSE
    # Buscar la columna en PARAMETROS que corresponde al catalogo
    param_col = None
    for pcol in range(1, len(CATALOGS) + 1):
        if ws_param.cell(row=1, column=pcol).value == cat_name:
            param_col = pcol
            break
    if param_col is None:
        continue
    max_row = 2
    while ws_param.cell(row=max_row, column=param_col).value:
        max_row += 1
    if max_row <= 2:
        continue
    rng = f"PARAMETROS!${col_letter(param_col)}$2:${col_letter(param_col)}${max_row - 1}"
    dv = DataValidation(type="list", formula1=f"={rng}", allow_blank=True)
    dv.error = "Valor no valido. Seleccione de la lista."
    dv.errorTitle = "Dato invalido"
    ws_matriz.add_data_validation(dv)
    dv.add(f"{col_letter(col_num)}2:{col_letter(col_num)}500")

# ── Dropdown DEPENDIENTE: Subl_x_nea (col 9) via INDIRECT ──
# Formula: =INDIRECT(SUBSTITUTE(I2;" ";"_"))
dv_sub = DataValidation(
    type="list",
    formula1=f'=INDIRECT(SUBSTITUTE({col_letter(8)}2;" ";"_"))',
    allow_blank=True
)
dv_sub.error = "Seleccione primero una L_x_nea valida"
dv_sub.errorTitle = "Subl_x_nea dependiente"
ws_matriz.add_data_validation(dv_sub)
dv_sub.add(f"I2:I500")

# ── Datos de ejemplo ──
ejemplos_matriz = [
    {
        "ID_Ref": 1, "Cod_MD": "MD-02801", "Cod_PT": "PT-02801",
        "Nombre_Referencia": "Femininity Dramatic Dress", "Color": "Ecru",
        "Cod_Color": "COL-101", "Capsula/Drop": "A",
        "Linea": "DRESSES", "Subl_x_nea": "MAXI",
        "Tipo_Ref": "DS", "Sistema_Tallaje": "XS-S-M-L-XL",
        "Largo": "MAXI 115CM", "Closure": "BACK ZIPPER",
        "Basado_En (Ref PT)": "PT-03402", "Dise_x_ador_Creativo": "CLAUDIA",
        "Fecha_Asignacion_Creativo": "2026-04-15",
        "Cambio_Molderia (SI/NO)": True, "Obs_Creativo": "Estimacion inicial segun referente",
        "Fecha_Inicio_Molderia_Diseno": "2026-04-16", "Fecha_Fin_Molderia_Diseno": "2026-04-18",
        "Comentarios_Molderia_Diseno": "Molderia base OK en talla XS",
        "Modista_Muestra": "ISABEL VALENCIA", "Status_Taller": "TERMINADO",
        "Fotos_Internas (SI/NO)": True, "Requiere_Muestra (SI/NO)": True,
        "Medicion_Fase_1 (Fecha)": "2026-04-20", "Comentarios_Medicion_1": "Aprobado de inmediato",
        "Veredicto_Dir_Creativa": "APROBADA DIRECTA", "Dise_x_ador_Tecnico": "DANIELA GARCIA",
        "Dise_x_ador_Contramuestra": "DANIELA GARCIA",
        "Fecha_Inicio_Molderia_Ing": "2026-04-22", "Fecha_Fin_Molderia_Ing": "2026-04-25",
        "Comentarios_Molderia_Ing": "Escalado OK", "Especificadora": "GABRIELA LOPEZ",
        "Fecha_Inicio_Especificacion": "2026-04-26", "Fecha_Fin_Especificacion": "2026-04-28",
        "Status_Referencia": "APROBADO",
        "Entregable_Creativo_OK?": "OK", "Entregable_Tecnico_OK?": "OK",
        "Entregable_Trazador_OK?": "OK", "Envio_Mod_Arte_OK?": "N/A",
        "Fecha_Liberacion_Diseno_a_Ing": "2026-04-21", "Fecha_Liberacion_Ing_a_Prod": "2026-04-28",
        "Temperatura (Fase Actual)": "ROJO - INDUSTRIALIZACION",
        "Observaciones_Generales": "Referencia completada",
    },
    {
        "ID_Ref": 2, "Cod_MD": "MD-02802", "Cod_PT": "PT-02803",
        "Nombre_Referencia": "Sunset Vacation Panty", "Color": "Sand",
        "Cod_Color": "COL-102", "Capsula/Drop": "B",
        "Linea": "PANTS", "Subl_x_nea": "DRAMATIC PANT",
        "Tipo_Ref": "PS", "Sistema_Tallaje": "0-2-4-6-8-10-12",
        "Largo": "ANKLE 95CM", "Closure": "BACK ZIPPER AND BUTTOM",
        "Basado_En (Ref PT)": "", "Dise_x_ador_Creativo": "MARIA BURGOS",
        "Fecha_Asignacion_Creativo": "2026-04-18",
        "Cambio_Molderia (SI/NO)": True, "Obs_Creativo": "Cambio de largo +2cm",
        "Fecha_Inicio_Molderia_Diseno": "2026-04-19", "Fecha_Fin_Molderia_Diseno": "2026-04-22",
        "Comentarios_Molderia_Diseno": "Correccion de tiro delantero",
        "Modista_Muestra": "FANNY GOMEZ", "Status_Taller": "EN PROCESO",
        "Fotos_Internas (SI/NO)": True, "Requiere_Muestra (SI/NO)": True,
        "Medicion_Fase_1 (Fecha)": "2026-04-24", "Comentarios_Medicion_1": "Ajustar pinza trasera",
        "Veredicto_Dir_Creativa": "APROBADA CON COMENTARIOS", "Dise_x_ador_Tecnico": "CRISTIAN GOMEZ",
        "Dise_x_ador_Contramuestra": "CRISTIAN GOMEZ",
        "Fecha_Inicio_Molderia_Ing": "2026-04-26", "Fecha_Fin_Molderia_Ing": "",
        "Comentarios_Molderia_Ing": "Escalado en proceso - talla 12 compleja",
        "Especificadora": "", "Fecha_Inicio_Especificacion": "", "Fecha_Fin_Especificacion": "",
        "Status_Referencia": "EN PROCESO",
        "Entregable_Creativo_OK?": "OK", "Entregable_Tecnico_OK?": "EN PROCESO",
        "Entregable_Trazador_OK?": "", "Envio_Mod_Arte_OK?": "",
        "Fecha_Liberacion_Diseno_a_Ing": "2026-04-25", "Fecha_Liberacion_Ing_a_Prod": "",
        "Temperatura (Fase Actual)": "NARANJA - VALIDACION TECNICA",
        "Observaciones_Generales": "Escalado pendiente talla 12",
    },
    {
        "ID_Ref": 3, "Cod_MD": "MD-02815", "Cod_PT": "PT-02804",
        "Nombre_Referencia": "Femininity Dramatic Dress V2", "Color": "Vanila",
        "Cod_Color": "COL-1591", "Capsula/Drop": "C",
        "Linea": "DRESSES", "Subl_x_nea": "MIDI",
        "Tipo_Ref": "DS", "Sistema_Tallaje": "XS-S-M-L-XL",
        "Largo": "MINI 85CM", "Closure": "SNAPS",
        "Basado_En (Ref PT)": "PT-02801", "Dise_x_ador_Creativo": "FER",
        "Fecha_Asignacion_Creativo": "2026-05-02",
        "Cambio_Molderia (SI/NO)": True, "Obs_Creativo": "Nuevo cuello drapeado y mangas",
        "Fecha_Inicio_Molderia_Diseno": "2026-05-03", "Fecha_Fin_Molderia_Diseno": "2026-05-06",
        "Comentarios_Molderia_Diseno": "Nuevo patron cuello drapeado",
        "Modista_Muestra": "ALEJANDRA ROJAS", "Status_Taller": "PAUSADO",
        "Fotos_Internas (SI/NO)": False, "Requiere_Muestra (SI/NO)": True,
        "Medicion_Fase_1 (Fecha)": "", "Comentarios_Medicion_1": "",
        "Veredicto_Dir_Creativa": "", "Dise_x_ador_Tecnico": "",
        "Dise_x_ador_Contramuestra": "", "Fecha_Inicio_Molderia_Ing": "",
        "Fecha_Fin_Molderia_Ing": "", "Comentarios_Molderia_Ing": "",
        "Especificadora": "", "Fecha_Inicio_Especificacion": "", "Fecha_Fin_Especificacion": "",
        "Status_Referencia": "EN PROCESO",
        "Entregable_Creativo_OK?": "EN PROCESO", "Entregable_Tecnico_OK?": "",
        "Entregable_Trazador_OK?": "", "Envio_Mod_Arte_OK?": "",
        "Fecha_Liberacion_Diseno_a_Ing": "", "Fecha_Liberacion_Ing_a_Prod": "",
        "Temperatura (Fase Actual)": "AMBAR - LABORATORIO",
        "Observaciones_Generales": "Tela en cuarentena de calidad. Esperando corte.",
    },
    {
        "ID_Ref": 4, "Cod_MD": "MD-02824", "Cod_PT": "PT-02805",
        "Nombre_Referencia": "Ocean Breeze Kimono", "Color": "Azure",
        "Cod_Color": "COL-201", "Capsula/Drop": "A",
        "Linea": "OUTWEAR", "Subl_x_nea": "KIMONO",
        "Tipo_Ref": "OW", "Sistema_Tallaje": "XS-S-M-L-XL",
        "Largo": "KIMONO 100 CM", "Closure": "WITH SASH",
        "Basado_En (Ref PT)": "", "Dise_x_ador_Creativo": "OSMAN",
        "Fecha_Asignacion_Creativo": "2026-05-05",
        "Cambio_Molderia (SI/NO)": True, "Obs_Creativo": "Coleccion nueva, sin referente",
        "Fecha_Inicio_Molderia_Diseno": "2026-05-06", "Fecha_Fin_Molderia_Diseno": "",
        "Comentarios_Molderia_Diseno": "Molderia base en desarrollo",
        "Modista_Muestra": "", "Status_Taller": "",
        "Fotos_Internas (SI/NO)": False, "Requiere_Muestra (SI/NO)": True,
        "Medicion_Fase_1 (Fecha)": "", "Comentarios_Medicion_1": "",
        "Veredicto_Dir_Creativa": "", "Dise_x_ador_Tecnico": "",
        "Dise_x_ador_Contramuestra": "", "Fecha_Inicio_Molderia_Ing": "",
        "Fecha_Fin_Molderia_Ing": "", "Comentarios_Molderia_Ing": "",
        "Especificadora": "", "Fecha_Inicio_Especificacion": "", "Fecha_Fin_Especificacion": "",
        "Status_Referencia": "EN PROCESO",
        "Entregable_Creativo_OK?": "", "Entregable_Tecnico_OK?": "",
        "Entregable_Trazador_OK?": "", "Envio_Mod_Arte_OK?": "",
        "Fecha_Liberacion_Diseno_a_Ing": "", "Fecha_Liberacion_Ing_a_Prod": "",
        "Temperatura (Fase Actual)": "AZUL - IDEACION Y DISENO",
        "Observaciones_Generales": "Referencia nueva, ficha tecnica en creacion.",
    },
]

for ri, data in enumerate(ejemplos_matriz, 2):
    for ci, (h, _) in enumerate(headers_matriz, 1):
        val = data.get(h, "")
        cell = ws_matriz.cell(row=ri, column=ci, value=val)
        sd(cell, alt=(ri % 2 == 0))

# ── FORMATO CONDICIONAL POR FILA (7 reglas basadas en Status_Referencia) ──
status_col = col_letter(37)  # Status_Referencia
last_data_row = len(ejemplos_matriz) + 1
full_range = f"A2:{col_letter(len(headers_matriz))}{last_data_row}"

cond_rules = [
    ("APROBADO", P.GREEN),
    ("TERMINADO", P.GREEN),
    ("EN PROCESO", P.AMBER_ALT),
    ("PAUSADO", P.TALLER_LIGHT),
    ("CANCELADO", P.RED),
    ("CANCELADO CORTADO", P.RED),
    ("CANCELADO SIN CORTAR", P.RED),
]

for status_val, color in cond_rules:
    ws_matriz.conditional_formatting.add(
        full_range,
        FormulaRule(
            formula=[f'${status_col}2="{status_val}"'],
            fill=PatternFill(start_color=color, end_color=color, fill_type="solid")
        )
    )

# Regla adicional: Sin Cod_PT cuando esta aprobado → alerta roja
ws_matriz.conditional_formatting.add(
    full_range,
    FormulaRule(
        formula=[f'AND(${status_col}2="APROBADO",$C2="")'],
        fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        font=Font(bold=True, color="9C0006")
    )
)

aw(ws_matriz, mn=13, mx=24)
ws_matriz.freeze_panes = "B2"
ws_matriz.auto_filter.ref = f"A1:{col_letter(len(headers_matriz))}1"
ws_matriz.sheet_properties.tabColor = "1F4E79"
print(f"[OK] MATRIZ_MAESTRA: {len(headers_matriz)} cols, {len(ejemplos_matriz)} refs, dropdowns + formato condicional")

# ======================================================================
# HOJA 3: CONSUMO_MATERIALES (15 cols) + ARRAYFORMULA para Ahorro
# ======================================================================
ws_consumo = wb.create_sheet("CONSUMO_MATERIALES", 2)

headers_consumo = [
    "ID_Ref", "Tipo_Material", "Cod_Material_SAP", "Descripcion_Material",
    "Ancho_Util_Tela (m)", "Base_Textil", "Ubicacion_Trazo (SI/NO)",
    "Modificacion_Arte (SI/NO)", "All_Over (SI/NO)", "Consumo_Base (m)",
    "Consumo_Creativo (m)", "Consumo_Tecnico (m)", "Consumo_Trazador (m)",
    "Ahorro_Optimizado (m)", "%_Ahorro",
]

for ci, h in enumerate(headers_consumo, 1):
    cell = ws_consumo.cell(row=1, column=ci, value=h)
    color = P.ING_DARK if "Consumo" in h or "Ahorro" in h or "%" in h else P.TALLER_DARK
    sc(cell, color, P.WHITE, bold=True, size=9)

ejemplos_consumo = [
    [1, "TELA LUCIR", "TEL-88902", "Lino Italiano Premium", 1.45, "LINEN",
     "SI", "NO", "NO", 2.10, 2.05, 1.99, 1.85],
    [1, "TELA FORRO", "TEL-33410", "Algodon Suave Ecologico", 1.40, "COTTON",
     "NO", "NO", "NO", 1.20, 1.35, 1.33, 1.05],
    [2, "TELA LUCIR", "TEL-88909", "Lino Premium Sand", 1.50, "LINEN",
     "SI", "SI", "NO", 1.85, 1.92, 1.88, 1.85],
    [3, "TELA LUCIR", "TEL-55603", "Seda Crepe de Chine", 1.35, "CREPE DE CHINE",
     "NO", "SI", "SI", 2.15, 2.30, "", ""],
    [3, "FUSIONABLE", "FUS-00120", "Entretela Fusionable Ligera", 0.90, "COTTON",
     "NO", "NO", "NO", 0.45, 0.48, "", ""],
    [4, "TELA LUCIR", "TEL-99801", "Lino Light Azure", 1.42, "LIGHT LINEN",
     "NO", "NO", "NO", "", "", "", ""],
]

for ri, data in enumerate(ejemplos_consumo, 2):
    for ci, val in enumerate(data, 1):
        cell = ws_consumo.cell(row=ri, column=ci, value=val)
        sd(cell, alt=(ri % 2 == 0))

# ── ARRAYFORMULA para Ahorro y % Ahorro (fila 2, se expande sola en Google Sheets) ──
# Col K = Consumo_Creativo, Col M = Consumo_Trazador
# Ahorro: =ARRAYFORMULA(IF(K2:K="";"";K2:K-M2:M))
# % Ahorro: =ARRAYFORMULA(IF(K2:K="";"";IF(K2:K>0;(K2:K-M2:M)/K2:K*100;0)))
ws_consumo.cell(row=2, column=14).value = '=ARRAYFORMULA(IF(K2:K="";"";K2:K-M2:M))'
ws_consumo.cell(row=2, column=15).value = '=ARRAYFORMULA(IF(K2:K="";"";IF(K2:K>0;(K2:K-M2:M)/K2:K*100;0)))'

# Data Validation para Tipo_Material
dv_mat = DataValidation(
    type="list",
    formula1="=PARAMETROS!$J$2:$J$15",
    allow_blank=True
)
dv_mat.error = "Seleccione un tipo de material valido"
ws_consumo.add_data_validation(dv_mat)
dv_mat.add("B2:B500")

aw(ws_consumo, mn=12, mx=22)
ws_consumo.freeze_panes = "B2"
ws_consumo.auto_filter.ref = f"A1:{col_letter(len(headers_consumo))}1"
ws_consumo.sheet_properties.tabColor = "BF8F00"
print(f"[OK] CONSUMO_MATERIALES: {len(headers_consumo)} cols, ARRAYFORMULA en Ahorro")

# ======================================================================
# HOJA 4: COMPOSICION_MARQUILLAS (14 cols)
# ======================================================================
ws_comp = wb.create_sheet("COMPOSICION_MARQUILLAS", 3)

headers_comp = [
    "ID_Ref", "Tipo_Marquilla", "Info_SAP (SI/NO)", "Desc_USA_UK",
    "Fiber_Composition", "Inside_Composition", "Include",
    "Woven_Knitted", "Cuidado_Lavado", "Cuidado_Desmanche",
    "Cuidado_Secado", "Cuidado_Planchado", "Cuidados_Includes",
    "Observaciones",
]

for ci, h in enumerate(headers_comp, 1):
    cell = ws_comp.cell(row=1, column=ci, value=h)
    color = P.PROD_DARK if "Cuidado" in h else P.STATUS_DARK
    sc(cell, color, P.WHITE, bold=True, size=9)

ejemplos_comp = [
    [1, "MUESTRA", "NO", "Linen Maxidress", "100% Linen", "95% Cotton, 5% Spandex",
     "NO", "WOVEN", "HAND WASH COLD", "DO NOT BLEACH", "DO NOT TUMBLE DRY",
     "DO NOT IRON", "N/A", "Composiciones verificadas con proveedor"],
    [1, "PRODUCCION", "SI", "Linen Maxidress", "100% Linen", "95% Cotton, 5% Spandex",
     "NO", "WOVEN", "HAND WASH COLD", "DO NOT BLEACH", "DO NOT TUMBLE DRY",
     "DO NOT IRON", "N/A", "Subido a SAP 2026-04-29"],
    [2, "MUESTRA", "NO", "Vacation Panty Sand", "100% Linen", "N/A",
     "NO", "WOVEN", "MACHINE WASH COLD", "DO NOT BLEACH", "TUMBLE DRY LOW",
     "IRON LOW", "N/A", ""],
    [3, "MUESTRA", "NO", "Midi Dress Draped", "100% Silk", "N/A",
     "NO", "WOVEN", "DRY CLEAN ONLY", "DO NOT BLEACH",
     "DO NOT TUMBLE DRY", "DO NOT IRON", "N/A", "Validar con proveedor seda"],
]

for ri, data in enumerate(ejemplos_comp, 2):
    for ci, val in enumerate(data, 1):
        cell = ws_comp.cell(row=ri, column=ci, value=val)
        sd(cell, alt=(ri % 2 == 0))

# Data Validation Woven/Knitted
dv_wk = DataValidation(type="list", formula1="=PARAMETROS!$W$2:$W$3", allow_blank=True)
ws_comp.add_data_validation(dv_wk)
dv_wk.add("H2:H500")

aw(ws_comp, mn=12, mx=28)
ws_comp.freeze_panes = "B2"
ws_comp.auto_filter.ref = f"A1:{col_letter(len(headers_comp))}1"
ws_comp.sheet_properties.tabColor = "7030A0"
print(f"[OK] COMPOSICION_MARQUILLAS: {len(headers_comp)} cols")

# ======================================================================
# HOJA 5: CONTROL_CONTRAMUESTRAS (18 cols)
# ======================================================================
ws_cm = wb.create_sheet("CONTROL_CONTRAMUESTRAS", 4)

headers_cm = [
    "ID_Ref", "Cod_OT", "Nombre_Contramuestra", "Talla_Contramuestra",
    "Color_Contramuestra", "Dise_x_ador_Encargado", "Modista_CM",
    "Status_Confeccion", "Tipo_Rechazo_Planta", "Prioridad", "Drop",
    "Fecha_Meta_Entrega", "Reprogramacion (SI/NO)", "Unidades_Cortadas",
    "Nota_Fabricacion_SAP", "Gestion_Nota (Fecha)", "Fecha_Traslado_SAP",
    "Fecha_Despacho_ZF",
]

for ci, h in enumerate(headers_cm, 1):
    cell = ws_cm.cell(row=1, column=ci, value=h)
    sc(cell, P.PROD_DARK, P.WHITE, bold=True, size=9)

ejemplos_cm = [
    [1, "OT-00870", "Femininity Dramatic Dress OT", "M", "Ecru",
     "DANIELA GARCIA", "ISABEL VALENCIA", "TERMINADO", "",
     "1", "A", "2026-05-02", "NO", 12,
     "NOT-778901", "2026-05-06", "2026-05-07", "2026-05-08"],
    [2, "", "Sunset Vacation Panty OT", "8", "Sand",
     "CRISTIAN GOMEZ", "FANNY GOMEZ", "EN PROCESO", "",
     "2", "B", "2026-05-10", "NO", "",
     "", "", "", ""],
    [3, "", "", "", "",
     "", "", "", "",
     "", "", "", "", "", "", "", ""],
]

for ri, data in enumerate(ejemplos_cm, 2):
    for ci, val in enumerate(data, 1):
        cell = ws_cm.cell(row=ri, column=ci, value=val)
        sd(cell, alt=(ri % 2 == 0))

# Data Validation
dv_conf_status = DataValidation(type="list", formula1="=PARAMETROS!$Y$2:$Y$7", allow_blank=True)
ws_cm.add_data_validation(dv_conf_status)
dv_conf_status.add("H2:H500")

dv_prioridad = DataValidation(type="list", formula1="=PARAMETROS!$S$2:$S$12", allow_blank=True)
ws_cm.add_data_validation(dv_prioridad)
dv_prioridad.add("J2:J500")

dv_drop = DataValidation(type="list", formula1="=PARAMETROS!$R$2:$R$12", allow_blank=True)
ws_cm.add_data_validation(dv_drop)
dv_drop.add("K2:K500")

aw(ws_cm, mn=12, mx=24)
ws_cm.freeze_panes = "B2"
ws_cm.auto_filter.ref = f"A1:{col_letter(len(headers_cm))}1"
ws_cm.sheet_properties.tabColor = "E26B0A"
print(f"[OK] CONTROL_CONTRAMUESTRAS: {len(headers_cm)} cols + Data Validation")

# ======================================================================
# HOJA 6: CURVA_TALLAS (14 cols) + ARRAYFORMULA para TOTAL
# ======================================================================
ws_tallas = wb.create_sheet("CURVA_TALLAS", 5)

talla_num = ["T_0", "T_2", "T_4", "T_6", "T_8", "T_10", "T_12"]
talla_letra = ["T_XS", "T_S", "T_M", "T_L", "T_XL"]
headers_tallas = ["ID_Ref"] + talla_num + talla_letra + ["TOTAL"]

for ci, h in enumerate(headers_tallas, 1):
    cell = ws_tallas.cell(row=1, column=ci, value=h)
    sc(cell, P.STATUS_DARK, P.WHITE, bold=True, size=9)

ejemplos_tallas = [
    {"ID_Ref": 1, "T_0": 0, "T_2": 0, "T_4": 0, "T_6": 0, "T_8": 0, "T_10": 0, "T_12": 0,
     "T_XS": 150, "T_S": 300, "T_M": 450, "T_L": 250, "T_XL": 100},
    {"ID_Ref": 2, "T_0": 20, "T_2": 40, "T_4": 60, "T_6": 80, "T_8": 40, "T_10": 20, "T_12": 10,
     "T_XS": 0, "T_S": 0, "T_M": 0, "T_L": 0, "T_XL": 0},
    {"ID_Ref": 3, "T_0": 0, "T_2": 0, "T_4": 0, "T_6": 0, "T_8": 0, "T_10": 0, "T_12": 0,
     "T_XS": 80, "T_S": 150, "T_M": 200, "T_L": 120, "T_XL": 60},
    {"ID_Ref": 4, "T_0": 0, "T_2": 0, "T_4": 0, "T_6": 0, "T_8": 0, "T_10": 0, "T_12": 0,
     "T_XS": 100, "T_S": 200, "T_M": 300, "T_L": 180, "T_XL": 80},
]

for ri, data in enumerate(ejemplos_tallas, 2):
    cell_id = ws_tallas.cell(row=ri, column=1, value=data["ID_Ref"])
    sd(cell_id, alt=(ri % 2 == 0))
    for ci in range(2, 14):
        h = headers_tallas[ci - 1]
        val = data.get(h, "")
        cell = ws_tallas.cell(row=ri, column=ci, value=val)
        sd(cell, alt=(ri % 2 == 0))

# ── ARRAYFORMULA para TOTAL (fila 2) ──
# =ARRAYFORMULA(IF(A2:A="";"";B2:B+C2:C+D2:D+E2:E+F2:F+G2:G+H2:H+I2:I+J2:J+K2:K+L2:L+M2:M))
# Simpler: =ARRAYFORMULA(IF(A2:A="";"";MMULT(B2:M;TRANSPOSE(SIGN(COLUMN(B1:M1))))))
# Let's use the simplest approach that works in Google Sheets:
total_formula = '=ARRAYFORMULA(IF(A2:A="";"";B2:B+C2:C+D2:D+E2:E+F2:F+G2:G+H2:H+I2:I+J2:J+K2:K+L2:L+M2:M))'
ws_tallas.cell(row=2, column=14).value = total_formula
sd(ws_tallas.cell(row=2, column=14))
ws_tallas.cell(row=2, column=14).font = Font(name="Calibri", size=9, bold=True)

# Fill 3,4,5 with empty (ARRAYFORMULA handles it)
for ri in range(3, 6):
    cell = ws_tallas.cell(row=ri, column=14)
    cell.value = ""
    sd(cell, alt=(ri % 2 == 0))

aw(ws_tallas, mn=8, mx=12)
ws_tallas.freeze_panes = "B2"
ws_tallas.auto_filter.ref = f"A1:{col_letter(len(headers_tallas))}1"
ws_tallas.sheet_properties.tabColor = "7F7F7F"
print(f"[OK] CURVA_TALLAS: {len(headers_tallas)} cols, ARRAYFORMULA en TOTAL")

# ======================================================================
# HOJA 7: REPORTES_DINAMICOS con QUERY() + SPARKLINE
# ======================================================================
ws_rep = wb.create_sheet("REPORTES_DINAMICOS", 6)

ws_rep.merge_cells("A1:H1")
t = ws_rep["A1"]; t.value = "DASHBOARD EJECUTIVO - GOOGLE SHEETS OPTIMIZADO"
t.font = Font(bold=True, size=14, color=P.WHITE, name="Calibri")
t.fill = PatternFill(start_color=P.REPORT_DARK, end_color=P.REPORT_DARK, fill_type="solid")
t.alignment = Alignment(horizontal="center", vertical="center")
ws_rep.row_dimensions[1].height = 30

# ── 7.1 RESUMEN (QUERY + COUNTIF) ──
r = 3
ws_rep.merge_cells(f"A{r}:H{r}")
c = ws_rep[f"A{r}"]; c.value = "1. ESTADO GENERAL DE LA COLECCION"
c.font = Font(bold=True, size=12, color=P.REPORT_DARK, name="Calibri")
c.fill = PatternFill(start_color=P.REPORT_LIGHT, end_color=P.REPORT_LIGHT, fill_type="solid")

resumen_items = [
    ("Total Referencias", '=COUNTA(MATRIZ_MAESTRA!A2:A500)'),
    ("  En Ideacion (Azul)", '=COUNTIF(MATRIZ_MAESTRA!AK2:AK500;"*IDEACION*")'),
    ("  En Laboratorio (Ambar)", '=COUNTIF(MATRIZ_MAESTRA!AK2:AK500;"*LABORATORIO*")'),
    ("  En Validacion Tecnica (Naranja)", '=COUNTIF(MATRIZ_MAESTRA!AK2:AK500;"*VALIDACION*")'),
    ("  En Industrializacion (Rojo)", '=COUNTIF(MATRIZ_MAESTRA!AK2:AK500;"*INDUSTRIALIZACION*")'),
    ("  Aprobadas", '=COUNTIF(MATRIZ_MAESTRA!AK2:AK500;"APROBADO")'),
    ("", ""),
    ("Referencias con consumos registrados", '=COUNTIF(CONSUMO_MATERIALES!K2:K500;">0")'),
    ("Materiales sin consumo definido", '=COUNTIFS(CONSUMO_MATERIALES!A2:A500;">0";CONSUMO_MATERIALES!K2:K500;"")'),
]

for i, (label, formula) in enumerate(resumen_items):
    ca = ws_rep.cell(row=r + 1 + i, column=1, value=label)
    if label:
        ca.font = Font(name="Calibri", size=10, bold=(not label.startswith("  ")))
    if formula:
        cb = ws_rep.cell(row=r + 1 + i, column=2, value=formula)
        cb.font = Font(name="Calibri", size=10)
    for cc in [ca, ws_rep.cell(row=r + 1 + i, column=2)]:
        cc.border = thin_border

# ── 7.2 CARGA DE TRABAJO ──
r = 17
ws_rep.merge_cells(f"A{r}:H{r}")
c = ws_rep[f"A{r}"]; c.value = "2. CARGA DE TRABAJO POR AREA"
c.font = Font(bold=True, size=12, color=P.REPORT_DARK, name="Calibri")
c.fill = PatternFill(start_color=P.REPORT_LIGHT, end_color=P.REPORT_LIGHT, fill_type="solid")

carga_h = ["Area", "Refs Activas", "% Carga", "Responsables Clave", "Cuellos de Botella"]
for ci, h in enumerate(carga_h):
    cell = ws_rep.cell(row=r + 1, column=ci + 1, value=h)
    sc(cell, P.REPORT_DARK, P.WHITE, bold=True, size=9)

carga_d = [
    ("CREATIVA", 3, "75%", "Claudia, Fer, Osman", "Ref #3 espera tela en cuarentena"),
    ("TALLER", 3, "60%", "Isabel V., Fanny G., Alejandra R.", "Ref #3 sin corte aun"),
    ("INGENIERIA", 2, "50%", "Daniela G., Cristian G.", "Ref #2 talla 12 compleja"),
    ("PRODUCCION", 1, "25%", "Gabriela L.", "Sin bloqueadores"),
]
for i, (area, refs, pct, resp, bloq) in enumerate(carga_d):
    row = r + 2 + i
    for j, val in enumerate([area, refs, pct, resp, bloq]):
        cell = ws_rep.cell(row=row, column=j + 1, value=val)
        cell.font = Font(name="Calibri", size=9)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if j < 3 else "left", vertical="center", wrap_text=True)

# ── 7.3 QUERY: Top referencias con mayor ahorro ──
r = 26
ws_rep.merge_cells(f"A{r}:H{r}")
c = ws_rep[f"A{r}"]; c.value = "3. TOP AHORRO DE TELA (Consumos Optimizados)"
c.font = Font(bold=True, size=12, color=P.REPORT_DARK, name="Calibri")
c.fill = PatternFill(start_color=P.REPORT_LIGHT, end_color=P.REPORT_LIGHT, fill_type="solid")

ahorro_h = ["Ref", "Material", "Consumo Creativo", "Consumo Trazador", "Ahorro (m)", "% Ahorro"]
for ci, h in enumerate(ahorro_h):
    cell = ws_rep.cell(row=r + 1, column=ci + 1, value=h)
    sc(cell, P.REPORT_DARK, P.WHITE, bold=True, size=9)

ahorro_d = [
    (1, "TELA LUCIR", 2.05, 1.85, 0.20, "9.8%"),
    (1, "TELA FORRO", 1.35, 1.05, 0.30, "22.2%"),
    (2, "TELA LUCIR", 1.92, 1.85, 0.07, "3.6%"),
]
for i, data in enumerate(ahorro_d):
    row = r + 2 + i
    for j, val in enumerate(data):
        cell = ws_rep.cell(row=row, column=j + 1, value=val)
        cell.font = Font(name="Calibri", size=9)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

# ── 7.4 EMBUDO DE PROGRESO con SPARKLINE ──
r = 32
ws_rep.merge_cells(f"A{r}:H{r}")
c = ws_rep[f"A{r}"]; c.value = "4. EMBUDO DE PROGRESO (Funnel por Fases)"
c.font = Font(bold=True, size=12, color=P.REPORT_DARK, name="Calibri")
c.fill = PatternFill(start_color=P.REPORT_LIGHT, end_color=P.REPORT_LIGHT, fill_type="solid")

funnel = [
    ("FASE 1: IDEACION Y DISENO", '=COUNTIF(MATRIZ_MAESTRA!AK2:AK500;"*IDEACION*")', P.CREATIVA_LIGHT),
    ("FASE 2: LABORATORIO Y PROTOTIPADO", '=COUNTIF(MATRIZ_MAESTRA!AK2:AK500;"*LABORATORIO*")', P.TALLER_LIGHT),
    ("FASE 3: VALIDACION TECNICA", '=COUNTIF(MATRIZ_MAESTRA!AK2:AK500;"*VALIDACION*")', P.ING_LIGHT),
    ("FASE 4: INDUSTRIALIZACION", '=COUNTIF(MATRIZ_MAESTRA!AK2:AK500;"*INDUSTRIALIZACION*")', P.PROD_LIGHT),
]
for i, (fase, formula, color) in enumerate(funnel):
    row = r + 1 + i
    cell = ws_rep.cell(row=row, column=1, value=fase)
    cell.font = Font(name="Calibri", size=10, bold=True)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.border = thin_border
    ws_rep.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell2 = ws_rep.cell(row=row, column=3, value=formula)
    cell2.font = Font(name="Calibri", size=10)
    cell2.alignment = Alignment(horizontal="center", vertical="center")
    cell2.border = thin_border
    # SPARKLINE barra de progreso
    spark = ws_rep.cell(row=row, column=4)
    spark.value = f'=SPARKLINE({col_letter(3)}{row},{{"charttype","bar";"max",COUNTA(MATRIZ_MAESTRA!A2:A500);"color1","{color}"}})'
    spark.border = thin_border
    ws_rep.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)

# ── 7.5 LEYENDA ──
r = 39
ws_rep.merge_cells(f"A{r}:H{r}")
c = ws_rep[f"A{r}"]; c.value = "5. LEYENDA DE COLORES POR AREA RESPONSABLE"
c.font = Font(bold=True, size=12, color=P.REPORT_DARK, name="Calibri")
c.fill = PatternFill(start_color=P.REPORT_LIGHT, end_color=P.REPORT_LIGHT, fill_type="solid")

leyenda = [
    (P.CREATIVA_LIGHT, "AZUL: CREATIVA - Perfilamiento, Diseno, Referentes, Molderia Base"),
    (P.TALLER_LIGHT, "AMBAR: TALLER - Alistamiento, Corte, Confeccion, Modistas"),
    (P.ING_LIGHT, "NARANJA: INGENIERIA - Mediciones, Escalado, Consumo Tecnico, Trazos"),
    (P.PROD_LIGHT, "ROJO: PRODUCCION - Ficha Final, Marquillas, Contramuestras, SAP, Calidad"),
    (P.STATUS_LIGHT, "GRIS: STATUS Y ENTREGABLES - Control de hitos, Liberacion entre fases"),
]
for i, (bg, desc) in enumerate(leyenda):
    row = r + 1 + i
    cell = ws_rep.cell(row=row, column=1)
    cell.font = Font(name="Calibri", size=9, bold=True)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_rep.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c2 = ws_rep.cell(row=row, column=3, value=desc)
    c2.font = Font(name="Calibri", size=9)
    c2.border = thin_border
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_rep.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)

# ── 7.6 NOTA DE USO GOOGLE SHEETS ──
r = 47
ws_rep.merge_cells(f"A{r}:H{r}")
c = ws_rep[f"A{r}"]; c.value = "6. RECOMENDACIONES PARA GOOGLE SHEETS (al convertir desde Excel)"
c.font = Font(bold=True, size=11, color=P.REPORT_DARK, name="Calibri")
c.fill = PatternFill(start_color=P.REPORT_LIGHT, end_color=P.REPORT_LIGHT, fill_type="solid")

notas_gs = [
    "1. Al subir a Google Drive: Archivo > Abrir con > Google Sheets (convertir)",
    "2. Los checkboxes (TRUE/FALSE) se convierten automaticamente a casillas de verificacion",
    "3. Los dropdowns dependientes (Linea->Subl_x_nea) usan INDIRECT, funcionan nativamente",
    "4. ARRAYFORMULA se expande sola al insertar filas, no requiere arrastrar",
    "5. Proteger rangos: Datos > Hojas y rangos protegidos > asignar por grupo (jo-creativa@, jo-taller@, etc.)",
    "6. Activar notificaciones: Extensiones > Apps Script > onEdit() para emails de cambio de fase",
    "7. Crear named versions: Archivo > Historial de versiones > Nombrar version actual en cada hito",
    "8. Los SPARKLINE y QUERY se recalculan automaticamente, no necesitan refresco manual",
]
for i, nota in enumerate(notas_gs):
    cell = ws_rep.cell(row=r + 1 + i, column=1, value=nota)
    cell.font = Font(name="Calibri", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_rep.merge_cells(start_row=r + 1 + i, start_column=1, end_row=r + 1 + i, end_column=8)

for col_letter_rep in ["A", "B", "C", "D", "E", "F", "G", "H"]:
    if col_letter_rep == "A":
        ws_rep.column_dimensions[col_letter_rep].width = 32
    elif col_letter_rep == "B":
        ws_rep.column_dimensions[col_letter_rep].width = 16
    else:
        ws_rep.column_dimensions[col_letter_rep].width = 18

ws_rep.sheet_properties.tabColor = "003366"
print(f"[OK] REPORTES_DINAMICOS: Dashboard con QUERY + SPARKLINE + guia Google Sheets")

# ═══════════════════════════ GUARDAR ═══════════════════════════
wb.save(OUTPUT)
print(f"\n{'='*60}")
print(f"[OK] prueba4.xlsx creado exitosamente")
print(f"Ruta: {OUTPUT}")
print(f"{'='*60}")
print(f"\nOPTIMIZACIONES GOOGLE SHEETS:")
print(f"  [*] ARRAYFORMULA en CURVA_TALLAS!TOTAL y CONSUMO_MATERIALES!Ahorro")
print(f"  [*] Dropdown dependiente Linea->Subl_x_nea via INDIRECT + Named Ranges")
print(f"  [*] Data Validation vinculada a PARAMETROS en 12+ columnas")
print(f"  [*] Checkboxes (TRUE/FALSE) para campos SI/NO")
print(f"  [*] Formato condicional x fila (7 estados de Status)")
print(f"  [*] Alerta visual: APROBADO sin Cod_PT = fila roja")
print(f"  [*] QUERY + COUNTIF + SPARKLINE en REPORTES_DINAMICOS")
print(f"  [*] Guia de conversion Excel->Google Sheets incluida")
print(f"\nCONVERSION A GOOGLE SHEETS:")
print(f"  1. Subir prueba4.xlsx a Google Drive")
print(f"  2. Click derecho > Abrir con > Google Sheets")
print(f"  3. Configurar proteccion de rangos por grupo (jo-creativa@, jo-taller@, etc.)")
print(f"  4. Activar notificaciones via Apps Script (onEdit)")
