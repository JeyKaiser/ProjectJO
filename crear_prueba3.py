#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de prueba3.xlsx - Formato Consolidado Mejorado
Consolida quirurgicamente FORMATO_CONTROL_CONSOLIDADO_JO.xlsx + prueba.xlsx + prueba2.xlsx
Arquitectura: 7 hojas interconectadas con PARAMETROS completo
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OUTPUT = r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0\prueba3.xlsx"

# ═══════════════════════════════════════════════════════════════
# PALETA DE COLORES POR AREA (codificacion semaforo)
# ═══════════════════════════════════════════════════════════════
class Palette:
    CREATIVA_DARK  = "1F4E79"   # Azul oscuro - FASE 1
    CREATIVA_LIGHT = "BDD7EE"   # Azul claro
    TALLER_DARK    = "BF8F00"   # Ambar oscuro - FASE 2
    TALLER_LIGHT   = "FFE699"   # Ambar claro
    ING_DARK       = "ED7D31"   # Naranja oscuro - FASE 3
    ING_LIGHT      = "F8CBAD"   # Naranja claro
    PROD_DARK      = "C00000"   # Rojo oscuro - FASE 4
    PROD_LIGHT     = "F4B4C2"   # Rojo claro
    STATUS_DARK    = "44546A"   # Gris azulado oscuro
    STATUS_LIGHT   = "D6DCE4"   # Gris azulado claro
    TIME_DARK      = "2E75B6"   # Azul medio
    TIME_LIGHT     = "BDD7EE"   # Azul claro
    REPORT_DARK    = "003366"   # Azul marino - REPORTES
    REPORT_LIGHT   = "D9E2F3"   # Azul palido
    PARAM_DARK     = "2C3E50"   # Azul grisaceo oscuro - PARAMETROS
    PARAM_CAT      = "F39C12"   # Naranja - Categorias parametros
    PARAM_CAT_BG   = "FDEBD0"   # Naranja palido
    WHITE          = "FFFFFF"
    BLACK          = "000000"
    GRAY           = "808080"
    GREEN          = "C6EFCE"
    RED            = "FFC7CE"
    AMBER          = "FFEB9C"

# ═══════════════════════════════════════════════════════════════
# UTILIDADES DE ESTILO
# ═══════════════════════════════════════════════════════════════
thin_border = Border(
    left=Side(style='thin', color=Palette.GRAY),
    right=Side(style='thin', color=Palette.GRAY),
    top=Side(style='thin', color=Palette.GRAY),
    bottom=Side(style='thin', color=Palette.GRAY)
)

def style_cell(cell, fill_color, font_color="FFFFFF", bold=True, size=10, wrap=True, align_center=True):
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.font = Font(color=font_color, bold=bold, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal="center" if align_center else "left",
                               vertical="center", wrap_text=wrap)
    cell.border = thin_border

def style_data_cell(cell, is_alt=False):
    cell.font = Font(name="Calibri", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if is_alt:
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

def write_headers(ws, row, headers, fill_color, font_color="FFFFFF", size=10):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        style_cell(cell, fill_color, font_color, bold=True, size=size)

def auto_width(ws, min_width=12, max_width=22):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for row in range(1, min(ws.max_row + 1, 50)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, min(len(str(val)) * 1.1 + 2, max_width))
        ws.column_dimensions[letter].width = max_len

# ═══════════════════════════════════════════════════════════════
# CATALOGOS DE PARAMETROS (33 categorias de prueba.xlsx)
# ═══════════════════════════════════════════════════════════════
CATALOGS = [
    ("STATUS", ["APROBADO", "APROBADO REPROGRAMACION", "CANCELADO", "CANCELADO CORTADO",
                "CANCELADO ES PERSONAL DE JO", "CANCELADO PAQUETE COMPLETO",
                "CANCELADO POR COMERCIAL", "CANCELADO SIN CORTAR", "EN PROCESO",
                "JUST FOR SHOW", "PAUSADO", "TERMINADO"]),
    ("DISEÑADORES CREATIVOS", ["BURGOS", "CLAUDIA", "FER", "MARGARITA", "MARIA BURGOS",
                                "MARIA DEL MAR", "OSMAN", "YAMILET"]),
    ("DISEÑADORES TÉCNICOS", ["CAMILA VILLEGAS", "CLAUDIA R", "CRISTIAN GOMEZ",
                               "DANIELA GARCIA", "FERNANDO", "KAROLINE",
                               "KELLY M.", "LINA DELGADO"]),
    ("TRAZADORES", ["KELLY MITROVICH", "ANDRES CORTADOR", "FERNANDO TRAZO"]),
    ("MODISTAS", ["ADRIANA ESCOBAR", "AISURY QUINTERO", "ALEJANDRA ROJAS",
                  "CIELO AGUIRRE", "DIANA ADARME", "ENEIDA MACA",
                  "FANNY GOMEZ", "ISABEL VALENCIA"]),
    ("LINEAS", ["DRESSES", "OUTWEAR", "PANTS", "SKIRTS", "SWIMWEAR", "TOPS"]),
    ("SUBLINEAS", ["ANKLE", "BIKINI BOTTOM", "BIKINI TOP", "CAPE", "COAT",
                   "CROP TOP", "JACKET", "KIMONO", "MAXI", "MIDI", "MINI",
                   "CASUAL", "BOLERO", "DRAMATIC PANT"]),
    ("TIPOS REF", ["DS", "OW", "PS", "SK", "SW", "TP"]),
    ("TALLAJES", ["0-2-4-6-8-10-12", "1-2-3-4-5-6", "2-4-6-8-10-12", "XS-S-M-L-XL"]),
    ("USOS EN PRENDA", ["TELA LUCIR", "TELA FORRO", "FUSIONABLE", "SESGOS LUCIR",
                        "SESGOS FORRO", "SESGOS FUSIONABLE", "ENTRETELA", "CINTA",
                        "RESORTE", "HILO RESORTE", "CIERRE", "ARGOLIA", "BOTON"]),
    ("BASES TEXTILES", ["-", "LINEN", "COTTON", "SILK", "WOOL", "LYCRA",
                        "LYCRA VITA", "LYCRA BAHIA", "LYCRA CRINKLE",
                        "COTTON JACQUARD", "PRINTED JACQUARD", "LEATHER",
                        "LIGHT LINEN", "VISCOSE", "POLYAMIDE", "POLYESTER",
                        "TAFETA SEDA", "DUCHESS SATIN", "CREPE DE CHINE",
                        "CREPE DE SEDA", "BROCADE JACQUARD", "CHARMEUSE",
                        "COTTON MESH", "COTTON VELVET"]),
    ("LARGOS", ["112 CM POSTERIOR: 95 CM", "85 CM", "92 CM", "ANKLE 95CM",
                "FLATS 112CM", "FLATS MAXI 103", "KIMONO 100 CM", "MAXI 115CM"]),
    ("SI/NO", ["SI", "NO"]),
    ("VEREDICTOS DIR. CREATIVA", ["APROBADA DIRECTA", "APROBADA CON COMENTARIOS", "RECHAZADA"]),
    ("TIPO BORDADO", ["SOBRE PRENDA ARMADA", "SEMIELABORADO", "AMBOS", "EN PRENDA",
                      "EMBROIDERED STRAP EDGE", "EMBROIDERED AND EMBROIDERED STRAPS"]),
    ("TIPO PROCESO EXTERNO", ["LAVANDERIA", "BORDADO", "DRAPEADO", "DESCOLADO",
                               "TINTORERIA", "MONTAJE MANIQUI"]),
    ("CATALOGACION", ["SOLIDO", "MODIFICACION DE ARTE", "UBICACION EN TRAZO",
                      "ALL OVER", "ALL OVER CON SENTIDO", "ALL OVER CON ORIENTACION"]),
    ("COLECCIONES", ["WINTER SUN (WS27)", "FALL WINTER (FW26)", "RESORT RTW (RS26)",
                     "SPRING SUMMER (SS26)", "SUMMER VACATION (SV26)", "PREFALL RTW (PF26)"]),
    ("DROPS", [chr(c) for c in range(65, 75)]),  # A-J
    ("PRIORIDADES", [str(i) for i in range(1, 11)]),
    ("DIFICULTAD", ["BAJA", "INTERMEDIA", "ALTA", "MUY ALTA"]),
    ("TIPO TEJIDO", ["TEJIDO PLANO", "TEJIDO DE PUNTO", "CUERO", "DENIM", "GASA"]),
    ("WOVEN/KNITTED", ["WOVEN", "KNITTED"]),
    ("TIPO CORTE", ["PRENDA COMPLETA", "LABORATORIO", "PIEZA", "REPOSICION",
                    "REPOSICION CONTRAMUESTRA"]),
    ("STATUS PROCESO EXT.", ["PENDIENTE", "EN PROCESO", "TERMINADO", "RECHAZADO"]),
    ("CONFECCION STATUS", ["EN PROCESO", "TERMINADO", "PAUSADO", "CANCELADO", "RECHAZADO"]),
    ("TIPO RECHAZO CALIDAD", ["MEDIDAS", "CALIDAD DE CONFECCION", "COHERENCIA CON FICHA TECNICA",
                               "MATERIA PRIMA", "INSUMOS", "OTRO"]),
    ("CLASIFICACION HALLAZGO", ["INCONSISTENCIAS", "FALTA DE INFORMACION", "FALTANTES",
                                  "CAMBIOS", "FALTA DE ANALISIS", "OTROS"]),
    ("TIPO EMPAQUE", ["CAJA INDIVIDUAL", "BOLSA COLECTIVA", "CAJA COLECTIVA",
                      "BOLSA INDIVIDUAL", "MARQUILLA SATINADA FORRADA", "SIN EMPAQUE ESPECIAL"]),
    ("CLOSURE TYPES", ["BACK ZIPPER", "BACK BUTTONS", "BACK ZIPPER AND BUTTOM",
                       "SNAPS", "WITH SASH", "N/A", "SIN CIERRE"]),
    ("LINNED", ["SI", "NO", "ONLY THE TOP", "ONLY THE BOTTOM", "N/A"]),
    ("SENTIDO SESGO", ["AL HILO", "A TRAVEZ", "AL SESGO"]),
    ("PERMITE GIRAR MOLDE", ["90°", "180°", "PEINE"]),
]

# ═══════════════════════════════════════════════════════════════
# WORKBOOK
# ═══════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ======================================================================
# HOJA 1: PARAMETROS
# ======================================================================
ws_param = wb.create_sheet("PARAMETROS", 0)

col = 1
for cat_name, values in CATALOGS:
    cell = ws_param.cell(row=1, column=col, value=cat_name)
    style_cell(cell, Palette.PARAM_DARK, Palette.WHITE, bold=True, size=9)
    alt_fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
    for j, val in enumerate(values):
        cell = ws_param.cell(row=2 + j, column=col, value=val)
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        if j % 2 == 0:
            cell.fill = alt_fill
    ws_param.column_dimensions[get_column_letter(col)].width = 26
    col += 1

ws_param.freeze_panes = "B2"
ws_param.sheet_properties.tabColor = "2C3E50"
print(f"[OK] PARAMETROS: {len(CATALOGS)} categorias creadas")

# ======================================================================
# HOJA 2: MATRIZ_MAESTRA (45 columnas - corazon del formato)
# ======================================================================
ws_matriz = wb.create_sheet("MATRIZ_MAESTRA", 1)

# NOTA: La coleccion se identifica por el nombre del archivo (ej: winter sun 2026.xlsx).
# Cod_PT se asigna desde el inicio (no solo al aprobar).
# Los consumos numericos viven en CONSUMO_MATERIALES, no aqui.
headers_matriz = [
    # ── SECCION 1: PERFILAMIENTO (CREATIVA - Azul) ── cols 1-6
    ("ID_Ref", Palette.CREATIVA_DARK),
    ("Cod_MD", Palette.CREATIVA_DARK),
    ("Cod_PT", Palette.CREATIVA_DARK),
    ("Nombre_Referencia", Palette.CREATIVA_DARK),
    ("Color", Palette.CREATIVA_DARK),
    ("Cod_Color", Palette.CREATIVA_DARK),
    # ── SECCION 2: SEGMENTACION (CREATIVA - Azul) ── cols 7-13
    ("Capsula/Drop", Palette.CREATIVA_DARK),
    ("Linea", Palette.CREATIVA_DARK),
    ("Sublínea", Palette.CREATIVA_DARK),
    ("Tipo_Ref", Palette.CREATIVA_DARK),
    ("Sistema_Tallaje", Palette.CREATIVA_DARK),
    ("Largo", Palette.CREATIVA_DARK),
    ("Closure", Palette.CREATIVA_DARK),
    # ── SECCION 3: REFERENTES Y DISENO (CREATIVA) ── cols 14-18
    ("Basado_En (Ref PT)", Palette.CREATIVA_DARK),
    ("Diseñador_Creativo", Palette.CREATIVA_DARK),
    ("Fecha_Asignacion_Creativo", Palette.CREATIVA_DARK),
    ("Cambio_Molderia (SI/NO)", Palette.CREATIVA_DARK),
    ("Obs_Creativo", Palette.CREATIVA_DARK),
    # ── SECCION 4: MOLDERIA BASE (CREATIVA) ── cols 19-21
    ("Fecha_Inicio_Molderia_Diseño", Palette.CREATIVA_DARK),
    ("Fecha_Fin_Molderia_Diseño", Palette.CREATIVA_DARK),
    ("Comentarios_Molderia_Diseño", Palette.CREATIVA_DARK),
    # ── SECCION 5: LABORATORIO (TALLER - Ambar) ── cols 22-25
    ("Modista_Muestra", Palette.TALLER_DARK),
    ("Status_Taller", Palette.TALLER_DARK),
    ("Fotos_Internas (SI/NO)", Palette.TALLER_DARK),
    ("Requiere_Muestra (SI/NO)", Palette.TALLER_DARK),
    # ── SECCION 6: VALIDACION TECNICA (INGENIERIA - Naranja) ── cols 26-33
    ("Medicion_Fase_1 (Fecha)", Palette.ING_DARK),
    ("Comentarios_Medicion_1", Palette.ING_DARK),
    ("Veredicto_Dir_Creativa", Palette.ING_DARK),
    ("Diseñador_Tecnico", Palette.ING_DARK),
    ("Diseñador_Contramuestra", Palette.ING_DARK),
    ("Fecha_Inicio_Molderia_Ing", Palette.ING_DARK),
    ("Fecha_Fin_Molderia_Ing", Palette.ING_DARK),
    ("Comentarios_Molderia_Ing", Palette.ING_DARK),
    # ── SECCION 7: INDUSTRIALIZACION (PRODUCCION - Rojo) ── cols 34-36
    ("Especificadora", Palette.PROD_DARK),
    ("Fecha_Inicio_Especificacion", Palette.PROD_DARK),
    ("Fecha_Fin_Especificacion", Palette.PROD_DARK),
    # ── SECCION 8: STATUS Y ENTREGABLES (Gris Azulado) ── cols 37-45
    ("Status_Referencia", Palette.STATUS_DARK),
    ("Entregable_Creativo_OK?", Palette.STATUS_DARK),
    ("Entregable_Tecnico_OK?", Palette.STATUS_DARK),
    ("Entregable_Trazador_OK?", Palette.STATUS_DARK),
    ("Envio_Mod_Arte_OK?", Palette.STATUS_DARK),
    ("Fecha_Liberacion_Diseno_a_Ing", Palette.STATUS_DARK),
    ("Fecha_Liberacion_Ing_a_Prod", Palette.STATUS_DARK),
    ("Temperatura (Fase Actual)", Palette.STATUS_DARK),
    ("Observaciones_Generales", Palette.STATUS_DARK),
]

for col_idx, (header, color) in enumerate(headers_matriz, 1):
    cell = ws_matriz.cell(row=1, column=col_idx, value=header)
    style_cell(cell, color, Palette.WHITE, bold=True, size=9)

# ── Datos de ejemplo (4 referencias en diferentes fases) ──
ejemplos_matriz = [
    # Ref 1: INDUSTRIALIZACION (APROBADA)
    {
        "ID_Ref": 1, "Cod_MD": "MD-02801", "Cod_PT": "PT-02801",
        "Nombre_Referencia": "Femininity Dramatic Dress", "Color": "Ecru",
        "Cod_Color": "COL-101",
        "Capsula/Drop": "A", "Linea": "DRESSES", "Sublínea": "MAXI",
        "Tipo_Ref": "DS", "Sistema_Tallaje": "XS-S-M-L-XL",
        "Largo": "MAXI 115CM", "Closure": "BACK ZIPPER",
        "Basado_En (Ref PT)": "PT-03402", "Diseñador_Creativo": "CLAUDIA",
        "Fecha_Asignacion_Creativo": "2026-04-15",
        "Cambio_Molderia (SI/NO)": "NO",
        "Obs_Creativo": "Estimacion inicial segun referente",
        "Fecha_Inicio_Molderia_Diseño": "2026-04-16", "Fecha_Fin_Molderia_Diseño": "2026-04-18",
        "Comentarios_Molderia_Diseño": "Molderia base OK en talla XS",
        "Modista_Muestra": "ISABEL VALENCIA", "Status_Taller": "APROBADO",
        "Fotos_Internas (SI/NO)": "SI", "Requiere_Muestra (SI/NO)": "SI",
        "Medicion_Fase_1 (Fecha)": "2026-04-20", "Comentarios_Medicion_1": "Aprobado de inmediato",
        "Veredicto_Dir_Creativa": "APROBADA DIRECTA", "Diseñador_Tecnico": "DANIELA GARCIA",
        "Diseñador_Contramuestra": "DANIELA GARCIA", "Fecha_Inicio_Molderia_Ing": "2026-04-22",
        "Fecha_Fin_Molderia_Ing": "2026-04-25", "Comentarios_Molderia_Ing": "Escalado OK, ajuste menorencogimiento",
        "Especificadora": "GABRIELA LOPEZ", "Fecha_Inicio_Especificacion": "2026-04-26",
        "Fecha_Fin_Especificacion": "2026-04-28",
        "Status_Referencia": "APROBADO", "Entregable_Creativo_OK?": "OK",
        "Entregable_Tecnico_OK?": "OK", "Entregable_Trazador_OK?": "OK",
        "Envio_Mod_Arte_OK?": "N/A", "Fecha_Liberacion_Diseno_a_Ing": "2026-04-21",
        "Fecha_Liberacion_Ing_a_Prod": "2026-04-28",
        "Temperatura (Fase Actual)": "ROJO - INDUSTRIALIZACION",
        "Observaciones_Generales": "Referencia completada, lista para produccion",
    },
    # Ref 2: VALIDACION TECNICA
    {
        "ID_Ref": 2, "Cod_MD": "MD-02802", "Cod_PT": "PT-02803",
        "Nombre_Referencia": "Sunset Vacation Panty", "Color": "Sand",
        "Cod_Color": "COL-102",
        "Capsula/Drop": "B", "Linea": "PANTS", "Sublínea": "DRAMATIC PANT",
        "Tipo_Ref": "PS", "Sistema_Tallaje": "0-2-4-6-8-10-12",
        "Largo": "ANKLE 95CM", "Closure": "BACK ZIPPER AND BUTTOM",
        "Basado_En (Ref PT)": "", "Diseñador_Creativo": "MARIA BURGOS",
        "Fecha_Asignacion_Creativo": "2026-04-18",
        "Cambio_Molderia (SI/NO)": "SI",
        "Obs_Creativo": "Cambio de largo +2cm",
        "Fecha_Inicio_Molderia_Diseño": "2026-04-19", "Fecha_Fin_Molderia_Diseño": "2026-04-22",
        "Comentarios_Molderia_Diseño": "Correccion de tiro delantero",
        "Modista_Muestra": "FANNY GOMEZ", "Status_Taller": "EN CONFECCION",
        "Fotos_Internas (SI/NO)": "SI", "Requiere_Muestra (SI/NO)": "SI",
        "Medicion_Fase_1 (Fecha)": "2026-04-24", "Comentarios_Medicion_1": "Ajustar pinza trasera",
        "Veredicto_Dir_Creativa": "APROBADA CON COMENTARIOS", "Diseñador_Tecnico": "CRISTIAN GOMEZ",
        "Diseñador_Contramuestra": "CRISTIAN GOMEZ", "Fecha_Inicio_Molderia_Ing": "2026-04-26",
        "Fecha_Fin_Molderia_Ing": "", "Comentarios_Molderia_Ing": "Escalado en proceso - talla 12 compleja",
        "Especificadora": "", "Fecha_Inicio_Especificacion": "",
        "Fecha_Fin_Especificacion": "",
        "Status_Referencia": "EN PROCESO", "Entregable_Creativo_OK?": "OK",
        "Entregable_Tecnico_OK?": "EN PROCESO", "Entregable_Trazador_OK?": "",
        "Envio_Mod_Arte_OK?": "", "Fecha_Liberacion_Diseno_a_Ing": "2026-04-25",
        "Fecha_Liberacion_Ing_a_Prod": "",
        "Temperatura (Fase Actual)": "NARANJA - VALIDACION TECNICA",
        "Observaciones_Generales": "Escalado pendiente en talla 12. Sin bloqueadores.",
    },
    # Ref 3: LABORATORIO
    {
        "ID_Ref": 3, "Cod_MD": "MD-02815", "Cod_PT": "PT-02804",
        "Nombre_Referencia": "Femininity Dramatic Dress V2", "Color": "Vanila",
        "Cod_Color": "COL-1591",
        "Capsula/Drop": "C", "Linea": "DRESSES", "Sublínea": "MIDI",
        "Tipo_Ref": "DS", "Sistema_Tallaje": "XS-S-M-L-XL",
        "Largo": "85 CM", "Closure": "SNAPS",
        "Basado_En (Ref PT)": "PT-02801", "Diseñador_Creativo": "FER",
        "Fecha_Asignacion_Creativo": "2026-05-02",
        "Cambio_Molderia (SI/NO)": "SI",
        "Obs_Creativo": "Modificacion completa de cuello y mangas",
        "Fecha_Inicio_Molderia_Diseño": "2026-05-03", "Fecha_Fin_Molderia_Diseño": "2026-05-06",
        "Comentarios_Molderia_Diseño": "Nuevo patron cuello drapeado",
        "Modista_Muestra": "ALEJANDRA ROJAS", "Status_Taller": "EN CORTE",
        "Fotos_Internas (SI/NO)": "NO", "Requiere_Muestra (SI/NO)": "SI",
        "Medicion_Fase_1 (Fecha)": "", "Comentarios_Medicion_1": "",
        "Veredicto_Dir_Creativa": "", "Diseñador_Tecnico": "",
        "Diseñador_Contramuestra": "", "Fecha_Inicio_Molderia_Ing": "",
        "Fecha_Fin_Molderia_Ing": "", "Comentarios_Molderia_Ing": "",
        "Especificadora": "", "Fecha_Inicio_Especificacion": "",
        "Fecha_Fin_Especificacion": "",
        "Status_Referencia": "EN PROCESO", "Entregable_Creativo_OK?": "EN PROCESO",
        "Entregable_Tecnico_OK?": "", "Entregable_Trazador_OK?": "",
        "Envio_Mod_Arte_OK?": "", "Fecha_Liberacion_Diseno_a_Ing": "",
        "Fecha_Liberacion_Ing_a_Prod": "",
        "Temperatura (Fase Actual)": "AMBAR - LABORATORIO",
        "Observaciones_Generales": "Esperando corte de muestra. Tela en cuarentena de calidad.",
    },
    # Ref 4: IDEACION
    {
        "ID_Ref": 4, "Cod_MD": "MD-02824", "Cod_PT": "PT-02805",
        "Nombre_Referencia": "Ocean Breeze Kimono", "Color": "Azure",
        "Cod_Color": "COL-201",
        "Capsula/Drop": "A", "Linea": "OUTWEAR", "Sublínea": "KIMONO",
        "Tipo_Ref": "OW", "Sistema_Tallaje": "XS-S-M-L-XL",
        "Largo": "KIMONO 100 CM", "Closure": "WITH SASH",
        "Basado_En (Ref PT)": "", "Diseñador_Creativo": "OSMAN",
        "Fecha_Asignacion_Creativo": "2026-05-05",
        "Cambio_Molderia (SI/NO)": "SI",
        "Obs_Creativo": "Coleccion nueva, sin referente",
        "Fecha_Inicio_Molderia_Diseño": "2026-05-06", "Fecha_Fin_Molderia_Diseño": "",
        "Comentarios_Molderia_Diseño": "Molderia base en desarrollo",
        "Modista_Muestra": "", "Status_Taller": "EN DISENO",
        "Fotos_Internas (SI/NO)": "NO", "Requiere_Muestra (SI/NO)": "SI",
        "Medicion_Fase_1 (Fecha)": "", "Comentarios_Medicion_1": "",
        "Veredicto_Dir_Creativa": "", "Diseñador_Tecnico": "",
        "Diseñador_Contramuestra": "", "Fecha_Inicio_Molderia_Ing": "",
        "Fecha_Fin_Molderia_Ing": "", "Comentarios_Molderia_Ing": "",
        "Especificadora": "", "Fecha_Inicio_Especificacion": "",
        "Fecha_Fin_Especificacion": "",
        "Status_Referencia": "EN PROCESO", "Entregable_Creativo_OK?": "",
        "Entregable_Tecnico_OK?": "", "Entregable_Trazador_OK?": "",
        "Envio_Mod_Arte_OK?": "", "Fecha_Liberacion_Diseno_a_Ing": "",
        "Fecha_Liberacion_Ing_a_Prod": "",
        "Temperatura (Fase Actual)": "AZUL - IDEACION Y DISENO",
        "Observaciones_Generales": "Referencia nueva, perfilando ficha tecnica.",
    },
]

for row_idx, data in enumerate(ejemplos_matriz, 2):
    for col_idx, (header, _) in enumerate(headers_matriz, 1):
        val = data.get(header, "")
        cell = ws_matriz.cell(row=row_idx, column=col_idx, value=val)
        style_data_cell(cell, is_alt=(row_idx % 2 == 0))

# Formato condicional para Status_Referencia
ws_matriz.conditional_formatting.add(
    f"A2:ZZ{len(ejemplos_matriz)+1}",
    FormulaRule(formula=[f'$AK2="APROBADO"'], fill=PatternFill(start_color=Palette.GREEN, end_color=Palette.GREEN, fill_type="solid"))
)
ws_matriz.conditional_formatting.add(
    f"A2:ZZ{len(ejemplos_matriz)+1}",
    FormulaRule(formula=[f'$AK2="CANCELADO"'], fill=PatternFill(start_color=Palette.RED, end_color=Palette.RED, fill_type="solid"))
)

auto_width(ws_matriz, min_width=13, max_width=24)
ws_matriz.freeze_panes = "B2"
ws_matriz.auto_filter.ref = f"A1:{get_column_letter(len(headers_matriz))}1"
ws_matriz.sheet_properties.tabColor = "1F4E79"
print(f"[OK] MATRIZ_MAESTRA: {len(headers_matriz)} columnas, {len(ejemplos_matriz)} referencias")

# ======================================================================
# HOJA 3: CONSUMO_MATERIALES (15 columnas - 1 fila por material usado)
# ======================================================================
ws_consumo = wb.create_sheet("CONSUMO_MATERIALES", 2)

headers_consumo = [
    "ID_Ref",
    "Tipo_Material",
    "Cod_Material_SAP",
    "Descripcion_Material",
    "Ancho_Util_Tela (m)",
    "Base_Textil",
    "Ubicacion_Trazo (SI/NO)",
    "Modificacion_Arte (SI/NO)",
    "All_Over (SI/NO)",
    "Consumo_Base (m)",
    "Consumo_Creativo (m)",
    "Consumo_Tecnico (m)",
    "Consumo_Trazador (m)",
    "Ahorro_Optimizado (m)",
    "%_Ahorro",
]

for col_idx, header in enumerate(headers_consumo, 1):
    cell = ws_consumo.cell(row=1, column=col_idx, value=header)
    color = Palette.ING_DARK if "Consumo" in header or "Ahorro" in header else Palette.TALLER_DARK
    style_cell(cell, color, Palette.WHITE, bold=True, size=9)

ejemplos_consumo = [
    # Ref 1 - 2 materiales
    [1, "TELA LUCIR", "TEL-88902", "Lino Italiano Premium", 1.45, "LINEN",
     "SI", "NO", "NO", 2.10, 2.05, 1.99, 1.85, None, None],
    [1, "TELA FORRO", "TEL-33410", "Algodon Suave Ecologico", 1.40, "COTTON",
     "NO", "NO", "NO", 1.20, 1.35, 1.33, 1.05, None, None],
    # Ref 2 - 1 material
    [2, "TELA LUCIR", "TEL-88909", "Lino Premium Sand", 1.50, "LINEN",
     "SI", "SI", "NO", 1.85, 1.92, 1.88, 1.85, None, None],
    # Ref 3 - 2 materiales
    [3, "TELA LUCIR", "TEL-55603", "Seda Crepe de Chine", 1.35, "CREPE DE CHINE",
     "NO", "SI", "SI", 2.15, 2.30, None, None, None, None],
    [3, "FUSIONABLE", "FUS-00120", "Entretela Fusionable Ligera", 0.90, "COTTON",
     "NO", "NO", "NO", 0.45, 0.48, None, None, None, None],
    # Ref 4 - 1 material
    [4, "TELA LUCIR", "TEL-99801", "Lino Light Azure", 1.42, "LIGHT LINEN",
     "NO", "NO", "NO", None, None, None, None, None, None],
]

for row_idx, data in enumerate(ejemplos_consumo, 2):
    for col_idx, val in enumerate(data, 1):
        cell = ws_consumo.cell(row=row_idx, column=col_idx, value=val)
        style_data_cell(cell, is_alt=(row_idx % 2 == 0))
    # Formulas de ahorro
    ahorro_cell = ws_consumo.cell(row=row_idx, column=14)
    pct_cell = ws_consumo.cell(row=row_idx, column=15)
    if data[10] and data[12]:  # Consumo Creativo y Trazador existen
        ahorro_cell.value = f"=K{row_idx}-M{row_idx}"
        pct_cell.value = f"=IF(K{row_idx}>0,N{row_idx}/K{row_idx}*100,0)"
        pct_cell.number_format = '0.00"%"'
    style_data_cell(ahorro_cell, is_alt=(row_idx % 2 == 0))
    style_data_cell(pct_cell, is_alt=(row_idx % 2 == 0))

auto_width(ws_consumo, min_width=12, max_width=22)
ws_consumo.freeze_panes = "B2"
ws_consumo.auto_filter.ref = f"A1:{get_column_letter(len(headers_consumo))}1"
ws_consumo.sheet_properties.tabColor = "BF8F00"
print(f"[OK] CONSUMO_MATERIALES: {len(headers_consumo)} columnas, {len(ejemplos_consumo)} registros")

# ======================================================================
# HOJA 4: COMPOSICION_MARQUILLAS (14 columnas)
# ======================================================================
ws_comp = wb.create_sheet("COMPOSICION_MARQUILLAS", 3)

headers_comp = [
    "ID_Ref",
    "Tipo_Marquilla",
    "Info_SAP (SI/NO)",
    "Desc_USA_UK",
    "Fiber_Composition",
    "Inside_Composition",
    "Include",
    "Woven_Knitted",
    "Cuidado_Lavado",
    "Cuidado_Desmanche",
    "Cuidado_Secado",
    "Cuidado_Planchado",
    "Cuidados_Includes",
    "Observaciones",
]

for col_idx, header in enumerate(headers_comp, 1):
    cell = ws_comp.cell(row=1, column=col_idx, value=header)
    color = Palette.PROD_DARK if "Cuidado" in header else Palette.STATUS_DARK
    style_cell(cell, color, Palette.WHITE, bold=True, size=9)

ejemplos_comp = [
    [1, "MUESTRA", "NO", "Linen Maxidress", "100% Linen", "95% Cotton, 5% Spandex",
     "NO", "WOVEN", "HAND WASH COLD", "DO NOT BLEACH", "DO NOT TUMBLE DRY",
     "DO NOT IRON", "N/A", "Composiciones verificadas con proveedor"],
    [1, "PRODUCCION", "SI", "Linen Maxidress", "100% Linen", "95% Cotton, 5% Spandex",
     "NO", "WOVEN", "HAND WASH COLD", "DO NOT BLEACH", "DO NOT TUMBLE DRY",
     "DO NOT IRON", "N/A", "Subido a SAP 2026-04-29"],
    [2, "MUESTRA", "NO", "Vacation Panty Sand", "100% Linen", "N/A",
     "NO", "WOVEN", "MACHINE WASH COLD", "DO NOT BLEACH", "TUMBLE DRY LOW",
     "IRON LOW", "N/A", ""],
    [3, "MUESTRA", "NO", "Midi Dress Draped", "100% Silk (Crepe de Chine)",
     "N/A", "NO", "WOVEN", "DRY CLEAN ONLY", "DO NOT BLEACH",
     "DO NOT TUMBLE DRY", "DO NOT IRON", "N/A", "Validar con proveedor seda"],
]

for row_idx, data in enumerate(ejemplos_comp, 2):
    for col_idx, val in enumerate(data, 1):
        cell = ws_comp.cell(row=row_idx, column=col_idx, value=val)
        style_data_cell(cell, is_alt=(row_idx % 2 == 0))

auto_width(ws_comp, min_width=12, max_width=28)
ws_comp.freeze_panes = "B2"
ws_comp.auto_filter.ref = f"A1:{get_column_letter(len(headers_comp))}1"
ws_comp.sheet_properties.tabColor = "7030A0"
print(f"[OK] COMPOSICION_MARQUILLAS: {len(headers_comp)} columnas, {len(ejemplos_comp)} registros")

# ======================================================================
# HOJA 5: CONTROL_CONTRAMUESTRAS (18 columnas)
# ======================================================================
ws_cm = wb.create_sheet("CONTROL_CONTRAMUESTRAS", 4)

headers_cm = [
    "ID_Ref",
    "Cod_OT",
    "Nombre_Contramuestra",
    "Talla_Contramuestra",
    "Color_Contramuestra",
    "Diseñador_Encargado",
    "Modista_CM",
    "Status_Confeccion",
    "Tipo_Rechazo_Planta",
    "Prioridad",
    "Drop",
    "Fecha_Meta_Entrega",
    "Reprogramacion (SI/NO)",
    "Unidades_Cortadas",
    "Nota_Fabricacion_SAP",
    "Gestion_Nota (Fecha)",
    "Fecha_Traslado_SAP",
    "Fecha_Despacho_ZF",
]

for col_idx, header in enumerate(headers_cm, 1):
    cell = ws_cm.cell(row=1, column=col_idx, value=header)
    color = Palette.PROD_DARK
    style_cell(cell, color, Palette.WHITE, bold=True, size=9)

ejemplos_cm = [
    [1, "OT-00870", "Femininity Dramatic Dress OT", "M", "Ecru",
     "DANIELA GARCIA", "ISABEL VALENCIA", "TERMINADO", "",
     "1", "A", "2026-05-02", "NO", 12,
     "NOT-778901", "2026-05-06", "2026-05-07", "2026-05-08"],
    [2, "", "Sunset Vacation Panty OT", "8", "Sand",
     "CRISTIAN GOMEZ", "FANNY GOMEZ", "EN PROCESO", "",
     "2", "B", "2026-05-10", "NO", "",
     "", "", "", ""],
    [3, "", "", "", "",
     "", "", "", "",
     "", "", "", "", "", "", "", ""],
]

for row_idx, data in enumerate(ejemplos_cm, 2):
    for col_idx, val in enumerate(data, 1):
        cell = ws_cm.cell(row=row_idx, column=col_idx, value=val)
        style_data_cell(cell, is_alt=(row_idx % 2 == 0))

auto_width(ws_cm, min_width=12, max_width=24)
ws_cm.freeze_panes = "B2"
ws_cm.auto_filter.ref = f"A1:{get_column_letter(len(headers_cm))}1"
ws_cm.sheet_properties.tabColor = "E26B0A"
print(f"[OK] CONTROL_CONTRAMUESTRAS: {len(headers_cm)} columnas, {len(ejemplos_cm)} registros")

# ======================================================================
# HOJA 6: CURVA_TALLAS (14 columnas)
# ======================================================================
ws_tallas = wb.create_sheet("CURVA_TALLAS", 5)

talla_num = ["T_0", "T_2", "T_4", "T_6", "T_8", "T_10", "T_12"]
talla_letra = ["T_XS", "T_S", "T_M", "T_L", "T_XL"]
headers_tallas = ["ID_Ref"] + talla_num + talla_letra + ["TOTAL"]

for col_idx, header in enumerate(headers_tallas, 1):
    cell = ws_tallas.cell(row=1, column=col_idx, value=header)
    style_cell(cell, Palette.STATUS_DARK, Palette.WHITE, bold=True, size=9)

ejemplos_tallas = [
    # Ref 1: Letras
    {"ID_Ref": 1, "T_0": 0, "T_2": 0, "T_4": 0, "T_6": 0, "T_8": 0, "T_10": 0, "T_12": 0,
     "T_XS": 150, "T_S": 300, "T_M": 450, "T_L": 250, "T_XL": 100},
    # Ref 2: Numeros
    {"ID_Ref": 2, "T_0": 20, "T_2": 40, "T_4": 60, "T_6": 80, "T_8": 40, "T_10": 20, "T_12": 10,
     "T_XS": 0, "T_S": 0, "T_M": 0, "T_L": 0, "T_XL": 0},
    # Ref 3: Mixto
    {"ID_Ref": 3, "T_0": 0, "T_2": 0, "T_4": 0, "T_6": 0, "T_8": 0, "T_10": 0, "T_12": 0,
     "T_XS": 80, "T_S": 150, "T_M": 200, "T_L": 120, "T_XL": 60},
    # Ref 4: Letras
    {"ID_Ref": 4, "T_0": 0, "T_2": 0, "T_4": 0, "T_6": 0, "T_8": 0, "T_10": 0, "T_12": 0,
     "T_XS": 100, "T_S": 200, "T_M": 300, "T_L": 180, "T_XL": 80},
]

for row_idx, data in enumerate(ejemplos_tallas, 2):
    cell_id = ws_tallas.cell(row=row_idx, column=1, value=data["ID_Ref"])
    style_data_cell(cell_id, is_alt=(row_idx % 2 == 0))
    for col_idx in range(2, 14):
        header = headers_tallas[col_idx - 1]
        val = data.get(header, "")
        cell = ws_tallas.cell(row=row_idx, column=col_idx, value=val)
        style_data_cell(cell, is_alt=(row_idx % 2 == 0))
    # TOTAL formula
    total_cell = ws_tallas.cell(row=row_idx, column=14)
    total_cell.value = f"=SUM(B{row_idx}:M{row_idx})"
    style_data_cell(total_cell, is_alt=(row_idx % 2 == 0))
    total_cell.font = Font(name="Calibri", size=9, bold=True)

auto_width(ws_tallas, min_width=8, max_width=12)
ws_tallas.freeze_panes = "B2"
ws_tallas.auto_filter.ref = f"A1:{get_column_letter(len(headers_tallas))}1"
ws_tallas.sheet_properties.tabColor = "7F7F7F"
print(f"[OK] CURVA_TALLAS: {len(headers_tallas)} columnas, {len(ejemplos_tallas)} referencias")

# ======================================================================
# HOJA 7: REPORTES_DINAMICOS (Dashboard ejecutivo)
# ======================================================================
ws_reportes = wb.create_sheet("REPORTES_DINAMICOS", 6)

# Titulo
ws_reportes.merge_cells("A1:H1")
title = ws_reportes["A1"]
title.value = "DASHBOARD EJECUTIVO - CONTROL DE COLECCION"
title.font = Font(bold=True, size=14, color=Palette.WHITE, name="Calibri")
title.fill = PatternFill(start_color=Palette.REPORT_DARK, end_color=Palette.REPORT_DARK, fill_type="solid")
title.alignment = Alignment(horizontal="center", vertical="center")
ws_reportes.row_dimensions[1].height = 30

# ── 7.1 ESTADO GENERAL ──
r = 3
ws_reportes.merge_cells(f"A{r}:H{r}")
cell = ws_reportes[f"A{r}"]
cell.value = "1. ESTADO GENERAL DE LA COLECCION"
cell.font = Font(bold=True, size=12, color=Palette.REPORT_DARK, name="Calibri")
cell.fill = PatternFill(start_color=Palette.REPORT_LIGHT, end_color=Palette.REPORT_LIGHT, fill_type="solid")

estado_rows = [
    ("Total Referencias", 4),
    ("  En Ideacion (Azul)", 1),
    ("  En Laboratorio (Ambar)", 1),
    ("  En Validacion Tecnica (Naranja)", 1),
    ("  En Industrializacion (Rojo)", 1),
    ("  Aprobadas", 1),
    ("  Canceladas", 0),
    ("", ""),
    ("Score de Salud Estimado", "85/100"),
    ("Tasa de Aprobacion", "25%"),
]

for i, (label, val) in enumerate(estado_rows):
    cell_a = ws_reportes.cell(row=r + 1 + i, column=1, value=label)
    cell_b = ws_reportes.cell(row=r + 1 + i, column=2, value=val)
    cell_a.font = Font(name="Calibri", size=10, bold=(i == 0 or "Score" in str(label) or "Tasa" in str(label)))
    cell_b.font = Font(name="Calibri", size=10)
    cell_a.border = thin_border
    cell_b.border = thin_border

# ── 7.2 CARGA DE TRABAJO POR AREA ──
r = 17
ws_reportes.merge_cells(f"A{r}:H{r}")
cell = ws_reportes[f"A{r}"]
cell.value = "2. CARGA DE TRABAJO POR AREA"
cell.font = Font(bold=True, size=12, color=Palette.REPORT_DARK, name="Calibri")
cell.fill = PatternFill(start_color=Palette.REPORT_LIGHT, end_color=Palette.REPORT_LIGHT, fill_type="solid")

carga_headers = ["Area", "Refs Activas", "% Carga", "Responsables Clave", "Cuellos de Botella"]
carga_colors = [Palette.CREATIVA_DARK, Palette.TALLER_DARK, Palette.ING_DARK, Palette.PROD_DARK]
for ci, h in enumerate(carga_headers):
    cell = ws_reportes.cell(row=r + 1, column=ci + 1, value=h)
    style_cell(cell, Palette.REPORT_DARK, Palette.WHITE, bold=True, size=9)

carga_data = [
    ("CREATIVA (Diseno y Moldería)", 3, "75%", "Claudia, Fer, Osman", "Ref #3 espera tela"),
    ("TALLER (Corte y Confeccion)", 3, "60%", "Isabel V., Fanny G., Alejandra R.", "Ref #3 tela en cuarentena"),
    ("INGENIERIA (Escalado y Trazos)", 2, "50%", "Daniela G., Cristian G.", "Ref #2 talla 12 compleja"),
    ("PRODUCCION (Fichas y SAP)", 1, "25%", "Gabriela L.", "Sin bloqueadores"),
]
for i, (area, refs, pct, resp, bloqueo) in enumerate(carga_data):
    row = r + 2 + i
    for j, val in enumerate([area, refs, pct, resp, bloqueo]):
        cell = ws_reportes.cell(row=row, column=j + 1, value=val)
        cell.font = Font(name="Calibri", size=9)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if j < 3 else "left", vertical="center", wrap_text=True)

# ── 7.3 AHORRO DE CONSUMOS ──
r = 24
ws_reportes.merge_cells(f"A{r}:H{r}")
cell = ws_reportes[f"A{r}"]
cell.value = "3. AHORRO DE CONSUMOS (Trazador vs Creativo)"
cell.font = Font(bold=True, size=12, color=Palette.REPORT_DARK, name="Calibri")
cell.fill = PatternFill(start_color=Palette.REPORT_LIGHT, end_color=Palette.REPORT_LIGHT, fill_type="solid")

ahorro_headers = ["Ref", "Material", "Consumo Creativo", "Consumo Trazador", "Ahorro (m)", "% Ahorro", "Ahorro Total Proyectado"]
for ci, h in enumerate(ahorro_headers):
    cell = ws_reportes.cell(row=r + 1, column=ci + 1, value=h)
    style_cell(cell, Palette.REPORT_DARK, Palette.WHITE, bold=True, size=9)

ahorro_data = [
    (1, "TELA LUCIR", "2.05 m", "1.85 m", "0.20 m", "9.8%", "250 m (1250 unids)"),
    (1, "TELA FORRO", "1.35 m", "1.05 m", "0.30 m", "22.2%", "375 m (1250 unids)"),
]
for i, data in enumerate(ahorro_data):
    row = r + 2 + i
    for j, val in enumerate(data):
        cell = ws_reportes.cell(row=row, column=j + 1, value=val)
        cell.font = Font(name="Calibri", size=9)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

# ── 7.4 FUNNEL DE FASES ──
r = 32
ws_reportes.merge_cells(f"A{r}:H{r}")
cell = ws_reportes[f"A{r}"]
cell.value = "4. EMBUDO DE PROGRESO (Funnel por Fases)"
cell.font = Font(bold=True, size=12, color=Palette.REPORT_DARK, name="Calibri")
cell.fill = PatternFill(start_color=Palette.REPORT_LIGHT, end_color=Palette.REPORT_LIGHT, fill_type="solid")

funnel = [
    ("FASE 1: IDEACION Y DISENO", "4 refs (100%)", "████████████████████████", Palette.CREATIVA_LIGHT),
    ("FASE 2: LABORATORIO Y PROTOTIPADO", "3 refs (75%)", "██████████████████", Palette.TALLER_LIGHT),
    ("FASE 3: VALIDACION TECNICA", "2 refs (50%)", "████████████", Palette.ING_LIGHT),
    ("FASE 4: INDUSTRIALIZACION", "1 ref (25%)", "██████", Palette.PROD_LIGHT),
]
for i, (fase, count, bar, color) in enumerate(funnel):
    row = r + 1 + i
    cell = ws_reportes.cell(row=row, column=1, value=fase)
    cell.font = Font(name="Calibri", size=10, bold=True)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.border = thin_border
    ws_reportes.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell2 = ws_reportes.cell(row=row, column=3, value=count)
    cell2.font = Font(name="Calibri", size=10)
    cell2.alignment = Alignment(horizontal="center", vertical="center")
    cell2.border = thin_border
    cell3 = ws_reportes.cell(row=row, column=4, value=bar)
    cell3.font = Font(name="Calibri", size=12)
    cell3.alignment = Alignment(horizontal="left", vertical="center")
    cell3.border = thin_border
    ws_reportes.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)

# ── 7.5 LEYENDA ──
r = 39
ws_reportes.merge_cells(f"A{r}:H{r}")
cell = ws_reportes[f"A{r}"]
cell.value = "5. LEYENDA DE COLORES POR AREA RESPONSABLE"
cell.font = Font(bold=True, size=12, color=Palette.REPORT_DARK, name="Calibri")
cell.fill = PatternFill(start_color=Palette.REPORT_LIGHT, end_color=Palette.REPORT_LIGHT, fill_type="solid")

leyenda = [
    ("AZUL", Palette.CREATIVA_LIGHT, "CREATIVA: Perfilamiento, Diseno, Consumos Creativos, Molderia Base"),
    ("AMBAR", Palette.TALLER_LIGHT, "TALLER: Alistamiento Telas, Corte, Confeccion, Procesos Especiales"),
    ("NARANJA", Palette.ING_LIGHT, "INGENIERIA: Mediciones, Ajustes, Escalado, Consumo Tecnico, Trazos"),
    ("ROJO", Palette.PROD_LIGHT, "PRODUCCION: Ficha Final, Marquillas, Contramuestras, SAP, Calidad"),
    ("GRIS", Palette.STATUS_LIGHT, "STATUS Y ENTREGABLES: Control de hitos y liberacion entre fases"),
]
for i, (color_name, bg, desc) in enumerate(leyenda):
    row = r + 1 + i
    cell = ws_reportes.cell(row=row, column=1, value=color_name)
    cell.font = Font(name="Calibri", size=9, bold=True)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell2 = ws_reportes.cell(row=row, column=2, value=desc)
    cell2.font = Font(name="Calibri", size=9)
    cell2.border = thin_border
    cell2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_reportes.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)

ws_reportes.column_dimensions["A"].width = 30
ws_reportes.column_dimensions["B"].width = 18
ws_reportes.column_dimensions["C"].width = 20
ws_reportes.column_dimensions["D"].width = 18
ws_reportes.column_dimensions["E"].width = 18
ws_reportes.column_dimensions["F"].width = 18
ws_reportes.column_dimensions["G"].width = 18
ws_reportes.column_dimensions["H"].width = 18
ws_reportes.sheet_properties.tabColor = "003366"
print(f"[OK] REPORTES_DINAMICOS: Dashboard ejecutivo creado")

# ═══════════════════════════════════════════════════════════════
# GUARDAR
# ═══════════════════════════════════════════════════════════════
wb.save(OUTPUT)
print(f"\n{'='*60}")
print(f"[OK] prueba3.xlsx creado exitosamente")
print(f"Ruta: {OUTPUT}")
print(f"{'='*60}")
print(f"\nESTRUCTURA FINAL (7 hojas):")
print(f"  1. PARAMETROS         - {len(CATALOGS)} catalogos maestros")
print(f"  2. MATRIZ_MAESTRA     - {len(headers_matriz)} columnas, {len(ejemplos_matriz)} referencias")
print(f"  3. CONSUMO_MATERIALES - {len(headers_consumo)} columnas, {len(ejemplos_consumo)} registros (multi-row)")
print(f"  4. COMPOSICION_MARQUILLAS - {len(headers_comp)} columnas")
print(f"  5. CONTROL_CONTRAMUESTRAS - {len(headers_cm)} columnas")
print(f"  6. CURVA_TALLAS       - {len(headers_tallas)} columnas con formulas SUM")
print(f"  7. REPORTES_DINAMICOS - Dashboard ejecutivo")
print(f"\nCOLORES POR AREA:")
print(f"  AZUL   -> CREATIVA (Ideacion y Diseno)")
print(f"  AMBAR  -> TALLER (Laboratorio y Prototipado)")
print(f"  NARANJA -> INGENIERIA (Validacion Tecnica)")
print(f"  ROJO   -> PRODUCCION (Industrializacion)")
print(f"  GRIS   -> STATUS Y ENTREGABLES (Transversal)")
