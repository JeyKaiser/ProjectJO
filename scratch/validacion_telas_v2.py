"""
VALIDACION DE TELAS v2.0 - Hoja de captura mejorada
Compatible 100% con migracion.gs (posiciones de columna identicas)
"""

import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

BASE = r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0"
DIST = os.path.join(BASE, "dist")

# ── Color Palette ─────────────────────────────────────────
TITLE_BG   = "1A1A2E"  # dark navy
TITLE_FG   = "FFFFFF"
PHASE1_BG  = "5DADE2"  # blue - Fase 1
PHASE2_BG  = "F4D03F"  # amber/gold - Fase 2
PHASE3_BG  = "E67E22"  # orange - Fase 3
PHASE4_BG  = "E74C3C"  # red - Fase 4
PHASE1_FG  = "FFFFFF"
PHASE2_FG  = "2C3E50"  # dark text on amber
PHASE3_FG  = "FFFFFF"
PHASE4_FG  = "FFFFFF"
ALT_ROW_BG = "F8F9FA"
WHITE      = "FFFFFF"
CANCEL_BG  = "D5D8DC"
ALERT_BG   = "FCF3CF"
INSTR_BG   = "EBF5FB"
BORDER_C   = "000000"

# ── Reusable styles ──────────────────────────────────────
thin_black = Border(
    left=Side(style='thin', color=BORDER_C),
    right=Side(style='thin', color=BORDER_C),
    top=Side(style='thin', color=BORDER_C),
    bottom=Side(style='thin', color=BORDER_C))
center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
center = Alignment(horizontal='center', vertical='center')
data_font = Font(name='Calibri', size=9, color='2C3E50')
title_font = Font(name='Calibri', size=16, bold=True, color=TITLE_FG)
subtitle_font = Font(name='Calibri', size=9, color='2C3E50')
instr_font = Font(name='Calibri', size=9, color='7F8C8D', italic=True)

# ── Column definitions (A=1, B=2, ... BN=66) ─────────────
# Phase sections: (name, bg_color, fg_color, [column letters])
PHASES = [
    ("FASE 1 - INFORMACION BASICA",  PHASE1_BG, PHASE1_FG,
     "A","B","C","D","E","F","G","H","I","J"),
    ("FASE 2 - MATERIALES (1 fila por tela/insumo)", PHASE2_BG, PHASE2_FG,
     "K","L","M","N","O"),
    ("FASE 3 - CREATIVO + TECNICO",  PHASE3_BG, PHASE3_FG,
     "P","Q","R","S","T","U","V","W","X","Y","Z",
     "AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL",
     "AM","AN","AO"),
    ("FASE 4 - TRAZO, CORTE Y CONTRAMUESTRA", PHASE4_BG, PHASE4_FG,
     "AP","AQ","AR","AS","AT","AU","AV","AW","AX",
     "AY","AZ","BA","BB","BC","BD","BE","BF","BG",
     "BH","BI","BJ","BK","BL","BM","BN","BO","BP"),
]

# Sub-headers for each column (A through BN = 66 columns)
SUB_HEADERS = {
    # FASE 1 - cols A-J (1-10)
    1: "REF", 2: "FOTO", 3: "REFERENCIA", 4: "STATUS",
    5: "MOD.ARTE", 6: "UBI.TRAZO", 7: "VAR.COLOR 1",
    8: "VAR.COLOR 2", 9: "LARGO",
    # FASE 2 - cols K-O (11-15)
    11: "USO PRENDA", 12: "COD TELA", 13: "DESC TELA",
    14: "ANCHO", 15: "CONSUMO BASE",
    # FASE 3a CREATIVO - cols P-U (16-21)
    16: "CREATIVO DIS", 17: "CAMBIO MOLD",
    18: "CREAT C1", 19: "CREAT C2", 20: "CREAT C3",
    21: "CREAT OBS",
    # FASE 3b TECNICO - cols V-AL (22-38)
    22: "TECNICO DIS",
    23: "SOL_TALLA", 24: "SOL_UND",
    25: "SOL_C1", 26: "SOL_C2", 27: "SOL_C3",
    28: "MA_TALLA", 29: "MA_UND",
    30: "MA_C1", 31: "MA_C2", 32: "MA_C3",
    33: "UT_TALLA", 34: "UT_UND",
    35: "UT_C1", 36: "UT_C2", 37: "UT_C3",
    38: "TEC OBS",
    # FASE 4a TRAZO - cols AM-BB (39-54)
    39: "TR_SOL_T", 40: "TR_SOL_U",
    41: "TR_SOL_C1", 42: "TR_SOL_C2", 43: "TR_SOL_C3", 44: "TR_SOL_C4",
    45: "TR_MA_T", 46: "TR_MA_U",
    47: "TR_MA_C1", 48: "TR_MA_C2", 49: "TR_MA_C3", 50: "TR_MA_C4",
    51: "TR_UT_T", 52: "TR_UT_U", 53: "TR_UT_C1",
    54: "TR_UT_OBS",
    # FASE 4b CONTRAMUESTRA - cols BC-BO (55-67)
    55: "CM_SOL_T", 56: "CM_SOL_U", 57: "CM_SOL_C1",
    58: "CM_MA_T", 59: "CM_MA_U", 60: "CM_MA_C1",
    61: "CM_UT_T", 62: "CM_UT_U", 63: "CM_UT_C1",
    64: "CM_OBS",
    # Extra costeo
    65: "TEC", 66: "TRAZADOR"
}

# ── Data Validation values (hardcoded for .xlsx) ──────────
# These will be the dropdown options. When uploaded to Sheets,
# user can redirect them to IMPORTRANGE.
DROPDOWNS = {
    4: ['APROBADO','CANCELADO','EN PROCESO','PENDIENTE','RECHAZADO',
        'PAUSADO','TERMINADO'],
    5: ['SI','NO'],  # MOD_ARTE
    6: ['SI','NO'],  # UBI_TRAZO
    11: ['TELA LUCIR','TELA FORRO','FUSIONABLE','SESGOS LUCIR',
         'SESGOS FORRO','SESGOS FUSIONABLE','ENTRETELA','CINTA',
         'RESORTE','HILO RESORTE','CIERRE','ARGOLIA','BOTON',
         'CORDON','BORLA','MARQUILLA','ELASTICO',
         'CINTA BOLILLO','CINTA ILUSION','CINTA RASSO',
         'CINTA RAZO','ELASTIC TRANSPARENTE','FRAMILON',
         'ESCALERILLA','HEBILLA JO','INSUMOS CON RECORRIDOS',
         'VARILLA','TENSORES 8MM'],
    16: ['CLAUDIA','FER','MARGARITA','MARIA BURGOS',
         'MARIA DEL MAR','OSMAN','YAMILETH'],
    17: ['SI','NO'],  # CAMBIO_MOLDERIA
    22: ['CAMILA VILLEGAS','CLAUDIA R','CRISTIAN GOMEZ',
         'DANIELA GARCIA','FERNANDO','KAROLINE',
         'KELLY M.','LINA DELGADO','LINA P.',
         'MARIA BURGOS','MARISOL R.'],
    65: ['CAMILA VILLEGAS','CRISTIAN GOMEZ','DANIELA GARCIA',
         'FERNANDO','KAROLINE','LINA DELGADO'],
    66: ['EQUIPO TRAZO','EQUIPO CORTE','EXTERNO'],
}

# ── Conditional formatting rules ──────────────────────────
CF_RULES = [
    # R1: CANCELADO row -> gray
    ('formula', '$D5="CANCELADO"', CANCEL_BG, '808080', True),
    # R2: STATUS empty + row has data -> alert
    ('formula', 'AND($A5<>"",$D5="")', ALERT_BG, '000000', False),
    # R3: EN PROCESO without designer -> alert
    ('formula', 'AND($D5="EN PROCESO",$P5="")', ALERT_BG, '000000', False),
    # R4: Material without tela code -> alert
    ('formula', 'AND($K5<>"",$L5="")', ALERT_BG, '000000', False),
]


def create_workbook():
    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════
    # SHEET: 1.VALIDACION DE TELAS
    # ══════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "1.VALIDACION DE TELAS"

    total_cols = 66  # A to BN

    # ── Row 1: TITLE ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(1, 1, "VALIDACION DE TELAS - COLECCION WS27")
    title_cell.font = title_font
    title_cell.fill = PatternFill(start_color=TITLE_BG, end_color=TITLE_BG, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = Border(bottom=Side(style='medium', color=TITLE_BG))
    ws.row_dimensions[1].height = 36

    # ── Row 2: QUICK INSTRUCTIONS ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    instr_cell = ws.cell(2, 1,
        "Complete las columnas de su area (FASE). Las demas columnas son de SOLO LECTURA para usted. "
        "Cada referencia ocupa 19 filas. NO elimine filas ni reordene columnas.")
    instr_cell.font = instr_font
    instr_cell.fill = PatternFill(start_color=INSTR_BG, end_color=INSTR_BG, fill_type="solid")
    instr_cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # ── Row 3: PHASE SECTION HEADERS ──
    ws.row_dimensions[3].height = 28
    for name, bg, fg, *cols in PHASES:
        first_col = openpyxl.utils.column_index_from_string(cols[0])
        last_col = openpyxl.utils.column_index_from_string(cols[-1])
        ws.merge_cells(start_row=3, start_column=first_col,
                       end_row=3, end_column=last_col)
        cell = ws.cell(3, first_col, name)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.font = Font(name='Calibri', size=10, bold=True, color=fg)
        cell.alignment = center_wrap
        cell.border = thin_black

    # ── Row 4: SUB-HEADERS ──
    ws.row_dimensions[4].height = 32
    for ci in range(1, total_cols + 1):
        name = SUB_HEADERS.get(ci, '')
        cell = ws.cell(4, ci, name)
        cell.font = Font(name='Calibri', size=8, bold=True, color='2C3E50')
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                    wrap_text=True, text_rotation=90)
        cell.border = thin_black
        ws.column_dimensions[get_column_letter(ci)].width = 7 if len(
            name) > 6 else 8.5

    # Override width for text-heavy columns
    wide_cols = {1: 7, 3: 35, 9: 10, 11: 18, 12: 14, 13: 30,
                 16: 16, 21: 40, 38: 40, 54: 40, 64: 40}
    for ci, w in wide_cols.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Row 5-99: DATA ROWS ──
    DATA_START = 5
    DATA_END = 99  # ~5 referencias * 19 filas = 95 filas
    for ri in range(DATA_START, DATA_END + 1):
        for ci in range(1, total_cols + 1):
            cell = ws.cell(ri, ci)
            cell.border = thin_black
            cell.font = data_font
            cell.alignment = center
            # Alternating row colors
            if (ri - DATA_START) % 2 == 0:
                cell.fill = PatternFill(start_color=ALT_ROW_BG,
                                         end_color=ALT_ROW_BG,
                                         fill_type="solid")

    # ── DATA VALIDATIONS (dropdowns) ──
    print("  Configurando data validations...")
    for col_num, options in DROPDOWNS.items():
        col_letter = get_column_letter(col_num)
        range_str = f"{col_letter}{DATA_START}:{col_letter}{DATA_END}"
        formula_str = '"' + ','.join(options) + '"'
        dv = DataValidation(type="list", formula1=formula_str, allow_blank=True)
        dv.error = "Valor no valido. Seleccione de la lista."
        dv.errorTitle = "Error de validacion"
        dv.prompt = "Seleccione una opcion de la lista"
        dv.promptTitle = "Valores permitidos"
        ws.add_data_validation(dv)
        dv.add(range_str)

    # ── CONDITIONAL FORMATTING ──
    print("  Configurando formato condicional...")
    data_range = f"A{DATA_START}:{get_column_letter(total_cols)}{DATA_END}"

    # R1: CANCELADO
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=['$D5="CANCELADO"'],
        fill=PatternFill(start_color=CANCEL_BG, end_color=CANCEL_BG, fill_type="solid"),
        font=Font(name='Calibri', size=9, color='808080', strike=True)))

    # R2: STATUS empty but REF has data
    ws.conditional_formatting.add(f"D{DATA_START}:D{DATA_END}", FormulaRule(
        formula=['AND($A5<>"",$D5="")'],
        fill=PatternFill(start_color=ALERT_BG, end_color=ALERT_BG, fill_type="solid"),
        font=Font(name='Calibri', size=9, color='C0392B', bold=True)))

    # R3: Material filled but no tela code
    ws.conditional_formatting.add(f"L{DATA_START}:L{DATA_END}", FormulaRule(
        formula=['AND($K5<>"",$L5="")'],
        fill=PatternFill(start_color=ALERT_BG, end_color=ALERT_BG, fill_type="solid"),
        font=Font(name='Calibri', size=9, color='C0392B', bold=True)))

    # R4: Approved -> green accent on REF
    ws.conditional_formatting.add(f"A{DATA_START}:A{DATA_END}", FormulaRule(
        formula=['$D5="APROBADO"'],
        fill=PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type="solid"),
        font=Font(name='Calibri', size=9, color='1E8449', bold=True)))

    # ── FREEZE PANES ──
    ws.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════
    # SHEET: INSTRUCCIONES
    # ══════════════════════════════════════════════════════
    print("  Creando hoja INSTRUCCIONES...")
    ws2 = wb.create_sheet("INSTRUCCIONES")
    ws2.sheet_properties.tabColor = "2980B9"
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 80

    r = 1
    ws2.cell(r, 2, "GUIA DE USO - VALIDACION DE TELAS v2.0").font = Font(
        name='Calibri', size=16, bold=True, color='1A1A2E')
    r += 2

    guia = [
        ("OBJETIVO", "Esta hoja captura los datos de telas, insumos y consumos de cada referencia. Los datos se migran automaticamente al archivo AtelierData para analisis e informes."),
        ("ESTRUCTURA", "Cada referencia ocupa 19 FILAS. La primera contiene datos generales (REF, STATUS, DISENADOR). Las siguientes 13 filas son para TELAS (una por cada material). Las ultimas 5 filas son para INSUMOS."),
        ("COLORES", "Cada FASE tiene un color distintivo:\n  🔵 Azul = FASE 1 (Info basica, Status)\n  🟡 Ambar = FASE 2 (Telas, Insumos)\n  🟠 Naranja = FASE 3 (Creativo + Tecnico)\n  🔴 Rojo = FASE 4 (Trazo, Corte, Contramuestra)"),
        ("QUIEN EDITA QUE", "Coordinadora: Columnas A-J\n  Creativo: Columnas K-U\n  Tecnico: Columnas V-AL\n  Trazador: Columnas AM-BB\n  Contramuestras: Columnas BC-BO"),
        ("FORMATO CONDICIONAL", "Filas en GRIS = Referencia CANCELADA\n  Celdas en AMARILLO = Campo requerido sin llenar\n  REF en VERDE = Referencia APROBADA"),
        ("VALIDACION", "Las columnas con dropdown (STATUS, DISENADOR, USO PRENDA) tienen valores predefinidos. No escriba valores manualmente: seleccione de la lista."),
        ("NO HACER", "NO elimine filas\n  NO inserte columnas\n  NO reordene las columnas\n  NO modifique las filas de encabezado (1-4)\n  Esto romperia la migracion automatica."),
        ("COMPATIBILIDAD", "Este formato es 100% compatible con migracion.gs. Las posiciones de columna son identicas a la version original."),
    ]

    for titulo, contenido in guia:
        ws2.cell(r, 2, titulo).font = Font(name='Calibri', size=11, bold=True,
                                            color='E74C3C')
        ws2.cell(r+1, 2, contenido).font = Font(name='Calibri', size=10,
                                                  color='2C3E50')
        ws2.cell(r+1, 2).alignment = Alignment(wrap_text=True, vertical='top')
        ws2.row_dimensions[r+1].height = max(30, contenido.count('\n') * 20 + 20)
        r += 3

    # ── Mapa de columnas ──
    r += 1
    ws2.cell(r, 2, "MAPA DE COLUMNAS (para referencia de migracion.gs)").font = Font(
        name='Calibri', size=11, bold=True, color='2980B9')
    r += 1

    ws2.cell(r, 2, "Col LETRA | Col NUM | Nombre | migracion.gs var").font = Font(
        name='Consolas', size=9, bold=True)
    r += 1

    key_vars = [
        ("A","1","REF","C.REF: 0"),
        ("B","2","FOTO","C.FOTO: 1"),
        ("C","3","REFERENCIA","C.REFERENCIA: 2"),
        ("D","4","STATUS","C.STATUS: 3"),
        ("E","5","MOD ARTE","C.MOD_ARTE: 4"),
        ("F","6","UBI TRAZO","C.UBI_TRAZO: 5"),
        ("K","11","USO PRENDA","C.USO_PRENDA: 9"),
        ("L","12","CODIGO TELA","C.CODIGO_TELA: 10"),
        ("M","13","DESC TELA","C.DESC_TELA: 11"),
        ("P","16","CREATIVO DIS","C.CREATIVO_DIS: 15"),
        ("R","18","CREATIVO C1","C.CREATIVO_C1: 17"),
        ("V","22","TECNICO DIS","C.TECNICO_DIS: 21"),
        ("Y","25","TEC SOL C1","C.TECNICO_SOL_C1: 24"),
        ("AO","41","TR SOL C1","C.TRAZO_SOL_C1: 40"),
        ("BE","57","CM SOL C1","C.CONTRA_SOL_C1: 56"),
    ]

    for letra, num, nombre, var in key_vars:
        ws2.cell(r, 2, f"{letra}       {num}       {nombre}       {var}").font = Font(
            name='Consolas', size=8, color='7F8C8D')
        r += 1

    # ── SAVE ──
    output = os.path.join(DIST, "VALIDACION_DE_TELAS_v2.xlsx")
    wb.save(output)
    print(f"\nArchivo guardado: {output}")
    print(f"Pestañas: {wb.sheetnames}")
    print(f"Columnas: {total_cols} (A-BN)")
    print(f"Filas de datos: {DATA_END - DATA_START + 1}")
    print("VALIDACION DE TELAS v2.0 COMPLETADO")


create_workbook()
