import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_consolidated_excel():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Typography & Styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    param_header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    
    # Custom harmonious pastel fills for sheet headers
    fills = {
        "MATRIZ_MAESTRA": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"), # Deep Blue
        "CONSUMO_MATERIALES": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"), # Steel Blue
        "COMPOSICION_MARQUILLAS": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"), # Purple
        "CONTROL_CONTRAMUESTRAS": PatternFill(start_color="E26B0A", end_color="E26B0A", fill_type="solid"), # Amber/Orange
        "CURVA_TALLAS": PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid"), # Slate Gray
        "PARAMETROS": PatternFill(start_color="375623", end_color="375623", fill_type="solid"), # Dark Green
    }
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. CREATE PARAMETROS SHEET FIRST (Since others will refer to it)
    ws_param = wb.create_sheet(title="PARAMETROS")
    ws_param.views.sheetView[0].showGridLines = True
    
    # Parameters columns
    params_data = {
        "Diseñadores_Creativos": ["Juan Perez", "Gabriela Lopez", "Johanna Ortiz", "Sofia Martinez", "Carlos Castano"],
        "Diseñadores_Técnicos": ["Maria Gomez", "Carlos Ruiz", "Andres Silva", "Luisa Fernandez"],
        "Modistas": ["Elena Rojas", "Lucia Diaz", "Maria Flores", "Juana Ramirez", "Rosa Gomez", "Patricia Soto"],
        "Especificadoras": ["Gabriela Lopez", "Sofia Torres", "Ana Maria Rendon"],
        "Líneas": ["Vestidos", "Pantalones", "Tops", "Jackets", "Enterizos", "Faldas"],
        "Bases_Textiles": ["Lino", "Seda", "Algodón", "Viscosa", "Poliéster", "Lana", "Encaje", "Dril", "Forro Tafeta"],
        "Status_Taller": ["En Corte", "En Confección", "En Lavado", "En Bordado", "Terminada", "Pausado"],
        "Veredictos": ["Aprobada Directa", "Aprobada con Comentarios", "Rechazada", "En Proceso"],
        "Tipos_Material": ["Tela Lucir", "Tela Forro", "Tela Fusionable", "Sesgo Lucir", "Sesgo Forro", "Sesgo Fusionable", "Insumo"],
        "Tipos_Rechazo": ["Medidas", "Calidad Confección", "Coherencia Ficha", "Tono Tela", "Defecto Tela"]
    }

    # Write parameter headers
    for col_idx, col_name in enumerate(params_data.keys(), 1):
        cell = ws_param.cell(row=1, column=col_idx, value=col_name.replace("_", " "))
        cell.font = param_header_font
        cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid") # Distinct Sage Green
        cell.alignment = align_center
        cell.border = thin_border
        
        # Write list items
        for row_idx, item in enumerate(params_data[col_name], 2):
            cell_item = ws_param.cell(row=row_idx, column=col_idx, value=item)
            cell_item.font = data_font
            cell_item.border = thin_border
            cell_item.alignment = align_left

    ws_param.row_dimensions[1].height = 28
    
    # Auto-adjust column widths for PARAMETROS
    for col in ws_param.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_param.column_dimensions[col_letter].width = max(max_len + 4, 16)

    # 2. DEFINE OTHER SHEETS
    sheets_data = {
        "MATRIZ_MAESTRA": [
            "ID_Ref", "Cod_MD", "Cod_PT", "Nombre_Referencia", "Color", "Cod_Color", 
            "Línea", "Sublínea", "Tipo_Ref", "Sistema_Tallaje", "Basado_En", 
            "Diseñador_Creativo", "Modista_Muestra", "Status_Taller", "Fotos_Internas", 
            "Medicion_Fase_1", "Comentarios_Med1", "Veredicto_Medicion", "Diseñador_Tecnico", 
            "Fecha_Inicio_Molderia", "Fecha_Fin_Molderia", "Comentarios_Molderia", 
            "Especificadora", "Fecha_Fin_Ficha", "Status_Referencia"
        ],
        "CONSUMO_MATERIALES": [
            "ID_Ref", "Tipo_Material", "Cod_Material_SAP", "Descripción_Material", 
            "Ancho_Util_Tela", "Base_Textil", "Ubicación_Trazo", "Modificación_Arte", 
            "Consumo_Base", "Consumo_Creativo", "Consumo_Tecnico", "Consumo_Trazador", 
            "Ahorro_Optimizado"
        ],
        "COMPOSICION_MARQUILLAS": [
            "ID_Ref", "Tipo_Marquilla", "Desc_USA_UK", "Fiber_Composition", 
            "Inside_Composition", "Woven_Knitted", "Cuidado_Lavado", "Cuidado_Secado"
        ],
        "CONTROL_CONTRAMUESTRAS": [
            "ID_Ref", "Cod_OT", "Talla_Contramuestra", "Diseñador_Encargado", 
            "Status_Confección", "Tipo_Rechazo_Planta", "Nota_Fabricacion_SAP", 
            "Fecha_Traslado_SAP", "Fecha_Despacho_ZF"
        ],
        "CURVA_TALLAS": [
            "ID_Ref", "T_0", "T_2", "T_4", "T_6", "T_8", "T_10", "T_12", 
            "T_XS", "T_S", "T_M", "T_L", "T_XL", "TOTAL"
        ]
    }

    # Populate sheets
    for name, cols in sheets_data.items():
        ws = wb.create_sheet(title=name)
        ws.views.sheetView[0].showGridLines = True
        
        # Write headers
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = title_font
            cell.fill = fills[name]
            cell.alignment = align_center
            cell.border = thin_border
            
        # Add sample mock data rows
        if name == "MATRIZ_MAESTRA":
            sample_rows = [
                [1, "MD-02801", "PT-02801", "Femininity Dramatic Dress", "Ecru", "COL-101", 
                 "Vestidos", "Maxidress", "Compleja", "Letras (XS-XL)", "", "Juan Perez", 
                 "Elena Rojas", "Aprobado", "SI", "2026-05-01", "Aprobado de inmediato con modelo", 
                 "Aprobada Directa", "Maria Gomez", "2026-05-02", "2026-05-04", 
                 "Escalado corregido por encogimiento", "Gabriela Lopez", "2026-05-05", "Aprobada"],
                [2, "MD-02802", "", "Sunset Vacation Pant", "Sand", "COL-102", 
                 "Pantalones", "Dramatic Pant", "Básica", "Números (0-12)", "PT-03402", "Juan Perez", 
                 "Lucia Diaz", "En Confección", "NO", "", "", "En Proceso", "", "", "", "", "", "", "En Desarrollo"]
            ]
            for row in sample_rows:
                ws.append(row)
        elif name == "CONSUMO_MATERIALES":
            sample_rows = [
                [1, "Tela Lucir", "TEL-88902", "Lino Italiano Premium", 1.45, "Lino", "SI", "NO", 2.10, 2.05, 1.95, 1.85, None],
                [1, "Tela Forro", "TEL-33410", "Algodón Suave Ecológico", 1.40, "Algodón", "NO", "NO", 1.20, 1.15, 1.10, 1.05, None]
            ]
            for row in sample_rows:
                ws.append(row)
            for r in range(2, 4):
                ws.cell(row=r, column=13, value=f"=J{r}-L{r}")
        elif name == "COMPOSICION_MARQUILLAS":
            sample_rows = [
                [1, "Producción", "Linen Maxidress", "100% Linen", "95% Cotton, 5% Spandex", "Woven (Plano)", "HAND WASH COLD", "DO NOT TUMBLE DRY"]
            ]
            for row in sample_rows:
                ws.append(row)
        elif name == "CONTROL_CONTRAMUESTRAS":
            sample_rows = [
                [1, "OT-00870", "M", "Maria Gomez", "Aprobada", "", "NOT-778901", "2026-05-06", "2026-05-07"]
            ]
            for row in sample_rows:
                ws.append(row)
        elif name == "CURVA_TALLAS":
            sample_rows = [
                [1, 0, 0, 0, 0, 0, 0, 0, 150, 300, 450, 250, 100, None]
            ]
            for row in sample_rows:
                ws.append(row)
            ws.cell(row=2, column=14, value="=SUM(B2:M2)")

        # Style data rows and make cells empty for subsequent rows to let users use validations
        max_r = max(ws.max_row, 100) # Pre-format 100 rows for smooth select usage
        for row_idx in range(2, max_r + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, len(cols) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if cell.value is None:
                    # Maintain alignments on empty cells
                    if col_idx in [1, 2, 3, 5, 6, 9, 10, 11, 14, 15, 16, 18, 20, 21, 23, 24, 25]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
                else:
                    if col_idx in [1, 2, 3, 5, 6, 9, 10, 11, 14, 15, 16, 18, 20, 21, 23, 24, 25]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
        
        ws.row_dimensions[1].height = 28
        
        # Column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # 3. APPLY EXCEL DATA VALIDATIONS (SELECT DROPDOWNS)
    ws_maestra = wb["MATRIZ_MAESTRA"]
    ws_consumo = wb["CONSUMO_MATERIALES"]
    ws_contramuestra = wb["CONTROL_CONTRAMUESTRAS"]

    # Validations rules referencing PARAMETROS sheet
    validations = {
        # MATRIZ_MAESTRA
        "maestra_linea": (ws_maestra, "G2:G100", "=PARAMETROS!$E$2:$E$10", "Selecciona una Línea", "Línea"),
        "maestra_tallaje": (ws_maestra, "J2:J100", '"Números (0-12),Letras (XS-XL)"', "Selecciona el tallaje", "Sistema Tallaje"), # Direct list
        "maestra_creativo": (ws_maestra, "L2:L100", "=PARAMETROS!$A$2:$A$10", "Selecciona un Diseñador Creativo", "Diseñador Creativo"),
        "maestra_modista": (ws_maestra, "M2:M100", "=PARAMETROS!$C$2:$C$10", "Selecciona una Modista", "Modista Muestra"),
        "maestra_status_taller": (ws_maestra, "N2:N100", "=PARAMETROS!$G$2:$G$10", "Selecciona un estado de taller", "Status Taller"),
        "maestra_fotos": (ws_maestra, "O2:O100", '"SI,NO"', "Indica si tiene fotos", "Fotos Internas"),
        "maestra_veredicto": (ws_maestra, "R2:R100", "=PARAMETROS!$H$2:$H$10", "Selecciona un veredicto de fit", "Veredicto Medición"),
        "maestra_tecnico": (ws_maestra, "S2:S100", "=PARAMETROS!$B$2:$B$10", "Selecciona un Diseñador Técnico", "Diseñador Técnico"),
        "maestra_especificadora": (ws_maestra, "W2:W100", "=PARAMETROS!$D$2:$D$10", "Selecciona una Especificadora", "Especificadora"),
        "maestra_status_ref": (ws_maestra, "Y2:Y100", '"En Desarrollo,Aprobada,Cancelada"', "Selecciona el estado final de la ref", "Status Referencia"),
        
        # CONSUMO_MATERIALES
        "consumo_tipo": (ws_consumo, "B2:B100", "=PARAMETROS!$I$2:$I$10", "Selecciona el tipo de material", "Tipo Material"),
        "consumo_base_textil": (ws_consumo, "F2:F100", "=PARAMETROS!$F$2:$F$12", "Selecciona la base textil", "Base Textil"),
        "consumo_trazo": (ws_consumo, "G2:G100", '"SI,NO"', "Indica ubicación en trazo", "Ubicación en Trazo"),
        "consumo_arte": (ws_consumo, "H2:H100", '"SI,NO"', "Indica modificación de arte", "Modificación Arte"),
        
        # CONTROL_CONTRAMUESTRAS
        "contra_status": (ws_contramuestra, "E2:E100", "=PARAMETROS!$H$2:$H$10", "Selecciona estatus de confección", "Status Confección"),
        "contra_rechazo": (ws_contramuestra, "F2:F100", "=PARAMETROS!$J$2:$J$10", "Selecciona tipo de rechazo si aplica", "Tipo Rechazo Planta")
    }

    for key, (ws, cell_range, formula, prompt_msg, title) in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "El valor ingresado no está en la lista de opciones válidas."
        dv.errorTitle = "Entrada no válida"
        dv.prompt = prompt_msg
        dv.promptTitle = title
        ws.add_data_validation(dv)
        dv.add(cell_range)

    # Save workbook
    filepath = "c:\\Users\\jchacon\\Documents\\JEFERSON STUDY\\Analisis de secuencia de coleccion\\version 2.0\\FORMATO_CONTROL_CONSOLIDADO_JO.xlsx"
    wb.save(filepath)
    print(f"Excel successfully updated with PARAMETROS sheet and dropdown selectors at: {filepath}")

if __name__ == "__main__":
    create_consolidated_excel()
