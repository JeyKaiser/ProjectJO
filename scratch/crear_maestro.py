"""
F2: Creacion del libro MAESTRO PARAMETROS JO.

Fusiona datos de los 3 silos en un unico libro con pestanas normalizadas:
  - COLORES: fusion A+C, deduplicado por CODIGO
  - TELAS: fusion A+C, deduplicado por CODIGO TELA
  - MATERIALES: desde B (estructura 1NF)
  - CATALOGOS: todos los catalogos nombrados extraidos de datos reales
  - COLECCIONES: definicion de colecciones
  - FOTOS_MAESTRO: estructura para IDs de fotos
  - CHANGE_LOG: plantilla vacia para trazabilidad
"""

import sys, io, os, re
from pathlib import Path
from datetime import datetime
from collections import OrderedDict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0")
DOCS = BASE / "Documentos"
DIST = BASE / "dist"
OUTPUT = DIST / "MAESTRO_PARAMETROS_JO.xlsx"

nan = 'nan'

thin_border = Border(
    left=Side(style='thin', color='808080'),
    right=Side(style='thin', color='808080'),
    top=Side(style='thin', color='808080'),
    bottom=Side(style='thin', color='808080')
)

header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
sub_header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
sub_header_font = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
alt_fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="Calibri", size=9)
accent_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")


def safe(v):
    if v is None or pd.isna(v):
        return ''
    s = str(v).strip()
    if s in ('', 'nan', '#N/A', '#REF!', '#REF! ', '#\u00a1REF!'):
        return ''
    return s


def fix_encoding(s):
    replacements = {
        'CAFÃ‰': 'CAFE', 'CAFÃ‰ CLARO': 'CAFE CLARO',
        'REPROGRAMACIÃ“N': 'REPROGRAMACION',
        'PEQCAMBIOS': 'PEQUENOS CAMBIOS',
        'ESPECIFICACIÃ“N': 'ESPECIFICACION',
        'PRODUCCIÃ“N': 'PRODUCCION',
        'CONFECCIÃ“N': 'CONFECCION',
        'INFORMACIÃ“N': 'INFORMACION',
        'SELECCIÃ“N': 'SELECCION',
        'VALIDACIÃ“N': 'VALIDACION',
        'CLASIFICACIÃ“N': 'CLASIFICACION',
        'COMPOSICIONES': 'COMPOSICIONES',
        'Ã‘': 'N',
    }
    for bad, good in replacements.items():
        s = s.replace(bad, good)
    return s


def write_header_row(ws, row, headers, fill=None, font=None):
    f = fill or header_fill
    fn = font or header_font
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.fill = f
        cell.font = fn
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def write_data_row(ws, row, values, start_col=1):
    for ci, v in enumerate(values, start_col):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.alignment = data_align
        cell.border = thin_border
        cell.font = data_font
        if (row - 1) % 2 == 0:
            cell.fill = alt_fill


def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


# ═══════════════════════════════════════════════════════════
# 1. CARGAR DATOS DE PARAMETROS (A y C)
# ═══════════════════════════════════════════════════════════

print("=" * 60)
print("CARGANDO DATOS DE PARAMETROS")
print("=" * 60)

dfA = pd.read_html(str(DOCS / "Copia de FO... APPAREL 2026" / "PARAMETROS.html"))[0]
dfC = pd.read_html(str(DOCS / "Copia de CONTRAM WS27" / "PARAMETROS.html"))[0]

hA = {}
for ci in range(1, len(dfA.columns)):
    v = str(dfA.iloc[1, ci]).strip() if pd.notna(dfA.iloc[1, ci]) else ''
    if v and v != nan:
        hA[ci] = v

hC = {}
for ci in range(1, len(dfC.columns)):
    v = str(dfC.iloc[1, ci]).strip() if pd.notna(dfC.iloc[1, ci]) else ''
    if v and v != nan:
        hC[ci] = v

dataA = []
for ri in range(2, len(dfA)):
    row = {}
    for ci, name in hA.items():
        val = dfA.iloc[ri, ci]
        if pd.notna(val):
            row[name] = fix_encoding(str(val).strip())
    if row:
        dataA.append(row)

dataC = []
for ri in range(2, len(dfC)):
    row = {}
    for ci, name in hC.items():
        val = dfC.iloc[ri, ci]
        if pd.notna(val):
            row[name] = fix_encoding(str(val).strip())
    if row:
        dataC.append(row)

print(f"Archivo A: {len(dataA)} filas, columnas: {list(hA.values())}")
print(f"Archivo C: {len(dataC)} filas, columnas: {list(hC.values())}")

# ═══════════════════════════════════════════════════════════
# 2. FUSIONAR COLORES (A + C, deduplicar por CODIGO)
# ═══════════════════════════════════════════════════════════

print("\nFusionando COLORES...")

colores_cols = ['CODIGO', 'DESCRIPCION']
colores_seen = {}
colores_merged = []

for src_label, data in [("A", dataA), ("C", dataC)]:
    for row in data:
        codigo = row.get('CODIGO', '').strip()
        descripcion = row.get('DESCRIPCION', '').strip()
        if not codigo:
            continue
        if codigo not in colores_seen:
            colores_seen[codigo] = {
                'CODIGO': codigo,
                'DESCRIPCION': descripcion,
                'FUENTE': src_label,
            }
            colores_merged.append(colores_seen[codigo])
        else:
            existing = colores_seen[codigo]
            if descripcion and not existing.get('DESCRIPCION', ''):
                existing['DESCRIPCION'] = descripcion
            existing['FUENTE'] = existing.get('FUENTE', '') + '+' + src_label

print(f"  COLORES fusionados: {len(colores_merged)} (deduplicados de {len(dataA)+len(dataC)} filas)")

# ═══════════════════════════════════════════════════════════
# 3. FUSIONAR TELAS (A + C, deduplicar por CODIGO TELA)
# ═══════════════════════════════════════════════════════════

print("\nFusionando TELAS...")

telas_cols = ['CODIGO TELA', 'DESCRIPCION TELA', 'ANCHO']
telas_seen = {}
telas_merged = []

for src_label, data in [("A", dataA), ("C", dataC)]:
    for row in data:
        cod_tela = row.get('CODIGO TELA', '').strip()
        desc_tela = row.get('DESCRIPCION TELA', '').strip()
        ancho = row.get('ANCHO', '').strip()
        obs = row.get('OBSERVACIONES', '').strip()
        coleccion = row.get('COLECCION', '').strip()

        if not cod_tela:
            continue
        if cod_tela not in telas_seen:
            telas_seen[cod_tela] = {
                'CODIGO_TELA': cod_tela,
                'DESCRIPCION_TELA': desc_tela,
                'ANCHO': ancho,
                'OBSERVACIONES': obs,
                'COLECCION': coleccion,
                'FUENTE': src_label,
            }
            telas_merged.append(telas_seen[cod_tela])
        else:
            existing = telas_seen[cod_tela]
            if desc_tela and not existing.get('DESCRIPCION_TELA', ''):
                existing['DESCRIPCION_TELA'] = desc_tela
            if ancho and not existing.get('ANCHO', ''):
                existing['ANCHO'] = ancho
            if coleccion and not existing.get('COLECCION', ''):
                existing['COLECCION'] = coleccion
            existing['FUENTE'] = existing.get('FUENTE', '') + '+' + src_label

print(f"  TELAS fusionadas: {len(telas_merged)} (deduplicadas de {len(dataA)+len(dataC)} filas)")

# ═══════════════════════════════════════════════════════════
# 4. EXTRAER MATERIALES DESDE B (1NF)
# ═══════════════════════════════════════════════════════════

print("\nExtrayendo MATERIALES desde Archivo B...")

csv_b = DOCS / "Copia de COLECCION WS27 - 1.VALIDACION DE TELAS.csv"
materiales = []

with open(csv_b, 'r', encoding='utf-8-sig') as f:
    import csv
    all_rows = list(csv.reader(f))

header_idx = 3
headers_b = [str(h).strip() for h in all_rows[header_idx]]

CSV_COL = {
    'REF': 0, 'REFERENCIA': 2, 'STATUS': 3,
    'MOD_ARTE': 4, 'UBI_TRAZO': 5, 'VARIACION_COLOR': 6,
    'LARGO': 8, 'USO_PRENDA': 9, 'CODIGO_TELA': 10,
    'DESCRIPCION_TELA': 11, 'ANCHO': 12, 'CONSUMO_BASE': 14,
    'DISENADOR_CREATIVO': 15, 'CAMBIO_MOLDERIA': 16,
    'CONSUMO1': 17, 'CONSUMO2': 18, 'CONSUMO3': 19,
    'DISENADOR_TECNICO': 21,
    'TALLA_SOLIDO': 22, 'CONSUMO1_SOLIDO': 24,
    'TALLA_MOD': 27, 'CONSUMO1_MOD': 29,
    'TALLA_UBI': 32, 'CONSUMO1_UBI': 34,
    'OBS_TECNICO': 37, 'TRAZADOR': 65,
    'TERMINADO_TALLER': 66,
}

current_ref = None
current_nombre = None
current_status = None
current_mod_arte = None
current_ubi_trazo = None
current_disenador_c = None

for i, csv_row in enumerate(all_rows[header_idx + 1:], header_idx + 1):
    def getv(key):
        idx = CSV_COL.get(key, -1)
        if idx >= 0 and idx < len(csv_row):
            return safe(csv_row[idx])
        return ''

    ref_val = getv('REF')
    uso = getv('USO_PRENDA')
    cod_tela = getv('CODIGO_TELA')

    if ref_val and ref_val.isdigit():
        current_ref = ref_val
        current_nombre = getv('REFERENCIA')
        current_status = getv('STATUS')
        current_mod_arte = getv('MOD_ARTE')
        current_ubi_trazo = getv('UBI_TRAZO')
        current_disenador_c = getv('DISENADOR_CREATIVO')

    if not current_ref:
        continue
    if not uso and not cod_tela:
        continue

    materiales.append({
        'REF': current_ref,
        'NOMBRE_REFERENCIA': current_nombre,
        'STATUS': current_status,
        'MOD_ARTE': current_mod_arte,
        'UBI_TRAZO': current_ubi_trazo,
        'DISENADOR_CREATIVO': current_disenador_c,
        'USO_EN_PRENDA': uso,
        'CODIGO_TELA': cod_tela,
        'DESCRIPCION_TELA': getv('DESCRIPCION_TELA'),
        'ANCHO': getv('ANCHO'),
        'CONSUMO_BASE': getv('CONSUMO_BASE'),
        'CONSUMO_CREATIVO_1': getv('CONSUMO1'),
        'CONSUMO_CREATIVO_2': getv('CONSUMO2'),
        'CONSUMO_CREATIVO_3': getv('CONSUMO3'),
        'DISENADOR_TECNICO': getv('DISENADOR_TECNICO'),
        'CONSUMO_SOLIDO': getv('CONSUMO1_SOLIDO'),
        'CONSUMO_MOD_ARTE': getv('CONSUMO1_MOD'),
        'CONSUMO_UBI_TRAZO': getv('CONSUMO1_UBI'),
        'TRAZADOR': getv('TRAZADOR'),
        'TERMINADO_TALLER': getv('TERMINADO_TALLER'),
        'COLECCION': 'WINTER SUN (WS27)',
    })

print(f"  MATERIALES extraidos: {len(materiales)} filas")

# ═══════════════════════════════════════════════════════════
# 5. EXTRAER CATALOGOS DE DATOS REALES
# ═══════════════════════════════════════════════════════════

print("\nExtrayendo catalogos de datos reales...")


def extract_csv_catalogs(csv_path, header_row_idx, catalog_cols, label):
    if not os.path.exists(str(csv_path)):
        print(f"  [{label}] No encontrado: {csv_path}")
        return {}
    df = pd.read_csv(str(csv_path), encoding='utf-8-sig', low_memory=False, header=None)
    data = df.iloc[header_row_idx + 1:]
    catalogs = {}
    for name, col_idx in catalog_cols.items():
        if col_idx < len(data.columns):
            vals = data.iloc[:, col_idx].dropna().astype(str)
            vals = sorted(set(
                fix_encoding(v.strip()) for v in vals
                if v.strip() not in ('', 'nan', '#N/A', '#REF!', '0', '0.0')
            ))
            if vals:
                catalogs[name] = vals
    return catalogs


cats_B = extract_csv_catalogs(
    DOCS / "Copia de COLECCION WS27 - 1.VALIDACION DE TELAS.csv", 3,
    {'STATUS': 3, 'USO_PRENDA': 9, 'DISENADOR_CREATIVO': 15,
     'CAMBIO_MOLDERIA': 16, 'DISENADOR_TECNICO': 21},
    "B"
)

cats_A_ws27 = extract_csv_catalogs(
    DOCS / "Copia de FO... APPAREL 2026 - WS27.csv", 2,
    {'STATUS': 10, 'DISENADOR_CREATIVO': 11, 'LINEA': 15, 'SUBLINEA': 16,
     'TIPO_REF': 17, 'TALLAJE': 18, 'LARGO': 19, 'CLOSURE': 20,
     'BASE_TEXTIL': 28, 'MOD_ARTE': 33, 'UBI_TRAZO': 34, 'ALL_OVER': 35,
     'STATUS_TALLER': 12, 'MODISTA': 13, 'LINNED': 21, 'INCLUDES': 22},
    "A-WS27"
)

cats_A_fw26 = extract_csv_catalogs(
    DOCS / "Copia de FO... APPAREL 2026 - FW26.csv", 3,
    {'STATUS': 10, 'DISENADOR_CREATIVO': 11, 'LINEA': 15, 'SUBLINEA': 16,
     'TIPO_REF': 17, 'TALLAJE': 18, 'LARGO': 19, 'CLOSURE': 20,
     'BASE_TEXTIL': 28, 'MOD_ARTE': 33, 'UBI_TRAZO': 34, 'ALL_OVER': 35,
     'STATUS_TALLER': 12, 'MODISTA': 13, 'LINNED': 21, 'INCLUDES': 22},
    "A-FW26"
)

# Merge and normalize
all_sources = {'B-COLECCION': cats_B, 'A-WS27': cats_A_ws27, 'A-FW26': cats_A_fw26}

NAME_NORMALIZE = {
    'BURGOS': 'MARIA BURGOS',
    'YAMILET': 'YAMILETH',
}

merged_catalogs = OrderedDict()
all_keys = set()
for cats in all_sources.values():
    all_keys.update(cats.keys())

for key in sorted(all_keys):
    all_vals = set()
    for src, cats in all_sources.items():
        if key in cats:
            for v in cats[key]:
                nv = NAME_NORMALIZE.get(v, v)
                all_vals.add(nv)
    merged_catalogs[key] = sorted(all_vals)

# Add supplementary catalogs not found in data
merged_catalogs['SI_NO'] = ['NO', 'SI']
merged_catalogs['VEREDICTO_DIR_CREATIVA'] = [
    'APROBADA DIRECTA', 'APROBADA CON COMENTARIOS', 'RECHAZADA'
]
merged_catalogs['TIPO_BORDADO'] = [
    'SOBRE PRENDA ARMADA', 'SEMIELABORADO', 'AMBOS', 'EN PRENDA',
    'EMBROIDERED STRAP EDGE', 'EMBROIDERED AND EMBROIDERED STRAPS'
]
merged_catalogs['TIPO_PROCESO_EXTERNO'] = [
    'LAVANDERIA', 'BORDADO', 'DRAPEADO', 'DESCOLADO', 'TINTORERIA',
    'MONTAJE MANIQUI'
]
merged_catalogs['CATALOGACION'] = [
    'SOLIDO', 'MODIFICACION DE ARTE', 'UBICACION EN TRAZO', 'ALL OVER',
    'ALL OVER CON SENTIDO', 'ALL OVER CON ORIENTACION',
    'REPROGRAMACION MOD ARTE', 'APOYO ESPECIAL'
]
merged_catalogs['DROPS'] = [chr(c) for c in range(65, 91)]
merged_catalogs['PRIORIDADES'] = [str(i) for i in range(1, 11)]
merged_catalogs['DIFICULTAD'] = ['BAJA', 'INTERMEDIA', 'ALTA', 'MUY ALTA']
merged_catalogs['TIPO_TEJIDO'] = ['TEJIDO PLANO', 'TEJIDO DE PUNTO', 'CUERO', 'DENIM', 'GASA']
merged_catalogs['WOVEN_KNITTED'] = ['WOVEN', 'KNITTED']
merged_catalogs['TIPO_CORTE'] = [
    'PRENDA COMPLETA', 'LABORATORIO', 'PIEZA', 'REPOSICION',
    'REPOSICION CONTRAMUESTRA'
]
merged_catalogs['CONFECCION_STATUS'] = [
    'EN PROCESO', 'TERMINADO', 'PAUSADO', 'CANCELADO', 'RECHAZADO'
]
merged_catalogs['TIPO_RECHAZO'] = [
    'MEDIDAS', 'CALIDAD DE CONFECCION', 'COHERENCIA CON FICHA TECNICA',
    'MATERIA PRIMA', 'INSUMOS', 'OTRO'
]
merged_catalogs['CLASIFICACION_HALLAZGO'] = [
    'INCONSISTENCIAS', 'FALTA DE INFORMACION', 'FALTANTES',
    'CAMBIOS', 'FALTA DE ANALISIS', 'OTROS'
]
merged_catalogs['SENTIDO_SESGO'] = ['AL HILO', 'A TRAVES', 'AL SESGO']
merged_catalogs['STATUS_PROCESO_EXT'] = ['PENDIENTE', 'EN PROCESO', 'TERMINADO', 'RECHAZADO']

print(f"  Catalogos fusionados: {len(merged_catalogs)}")
for k, v in merged_catalogs.items():
    print(f"    {k}: {len(v)} valores")

# ═══════════════════════════════════════════════════════════
# 6. CREAR LIBRO EXCEL MAESTRO
# ═══════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("CREANDO LIBRO MAESTRO PARAMETROS JO")
print("=" * 60)

wb = openpyxl.Workbook()
ws_first = wb.active
ws_first.title = "INICIO"

# ── Pestaña INICIO ──
ws_first.cell(row=1, column=1, value="MAESTRO PARAMETROS JO").font = Font(
    size=16, bold=True, name="Calibri", color="2C3E50")
ws_first.cell(row=2, column=1, value=f"Creado: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(
    size=10, name="Calibri", color="7F8C8D")
ws_first.cell(row=4, column=1, value="INDICE DE PESTANAS").font = Font(
    size=12, bold=True, name="Calibri")

index_items = [
    ("COLORES", "Codigos de color y descripcion (fusion A+C, deduplicado)"),
    ("TELAS", "Codigos de tela, descripcion, ancho (fusion A+C, deduplicado)"),
    ("MATERIALES", "Relacion REF-Tela/Insumo en estructura 1NF (desde B)"),
    ("CATALOGOS", "Todos los catalogos desplegables extraidos de datos reales"),
    ("DISENADORES", "Disenadores creativos y tecnicos (normalizados)"),
    ("MODISTAS", "Modistas asignadas"),
    ("LINEAS", "Lineas y sublineas de producto"),
    ("COLECCIONES", "Definicion de colecciones y temporadas"),
    ("FOTOS_MAESTRO", "Registro centralizado de IDs de fotos en Drive"),
    ("CHANGE_LOG", "Trazabilidad de cambios (rellenado por Apps Script)"),
]

for i, (name, desc) in enumerate(index_items, 5):
    ws_first.cell(row=i, column=1, value=name).font = Font(
        bold=True, size=10, name="Calibri", color="2980B9")
    ws_first.cell(row=i, column=2, value=desc).font = Font(
        size=10, name="Calibri", color="7F8C8D")

ws_first.column_dimensions['A'].width = 25
ws_first.column_dimensions['B'].width = 60

# ── Pestaña COLORES ──
print("Creando COLORES...")
ws_col = wb.create_sheet("COLORES")
col_headers = ['CODIGO', 'DESCRIPCION', 'FUENTE_ORIGEN']
write_header_row(ws_col, 1, col_headers)
for ri, row_data in enumerate(colores_merged, 2):
    write_data_row(ws_col, ri, [row_data.get('CODIGO', ''), row_data.get('DESCRIPCION', ''), row_data.get('FUENTE', '')])
auto_width(ws_col)
print(f"  {len(colores_merged)} colores")

# ── Pestaña TELAS ──
print("Creando TELAS...")
ws_tel = wb.create_sheet("TELAS")
tel_headers = ['CODIGO_TELA', 'DESCRIPCION_TELA', 'ANCHO', 'OBSERVACIONES', 'COLECCION', 'FUENTE_ORIGEN']
write_header_row(ws_tel, 1, tel_headers)
for ri, row_data in enumerate(telas_merged, 2):
    write_data_row(ws_tel, ri, [
        row_data.get('CODIGO_TELA', ''), row_data.get('DESCRIPCION_TELA', ''),
        row_data.get('ANCHO', ''), row_data.get('OBSERVACIONES', ''),
        row_data.get('COLECCION', ''), row_data.get('FUENTE', '')
    ])
auto_width(ws_tel)
print(f"  {len(telas_merged)} telas")

# ── Pestaña MATERIALES ──
print("Creando MATERIALES...")
ws_mat = wb.create_sheet("MATERIALES")
mat_headers = [
    'REF', 'NOMBRE_REFERENCIA', 'STATUS', 'MOD_ARTE', 'UBI_TRAZO',
    'DISENADOR_CREATIVO', 'USO_EN_PRENDA', 'CODIGO_TELA',
    'DESCRIPCION_TELA', 'ANCHO', 'CONSUMO_BASE',
    'CONSUMO_CREATIVO_1', 'CONSUMO_CREATIVO_2', 'CONSUMO_CREATIVO_3',
    'DISENADOR_TECNICO', 'CONSUMO_SOLIDO', 'CONSUMO_MOD_ARTE',
    'CONSUMO_UBI_TRAZO', 'TRAZADOR', 'TERMINADO_TALLER', 'COLECCION'
]
write_header_row(ws_mat, 1, mat_headers)
for ri, row_data in enumerate(materiales, 2):
    vals = [row_data.get(h, '') for h in mat_headers]
    write_data_row(ws_mat, ri, vals)
auto_width(ws_mat)
print(f"  {len(materiales)} materiales")

# ── Pestaña CATALOGOS ──
print("Creando CATALOGOS...")
ws_cat = wb.create_sheet("CATALOGOS")

cat_names = sorted(merged_catalogs.keys())
max_rows = max(len(merged_catalogs[c]) for c in cat_names) + 1

write_header_row(ws_cat, 1, cat_names)
for ci, cat_name in enumerate(cat_names, 1):
    cell = ws_cat.cell(row=1, column=ci, value=cat_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

    for ri, val in enumerate(merged_catalogs[cat_name], 2):
        cell = ws_cat.cell(row=ri, column=ci, value=val)
        cell.alignment = data_align
        cell.border = thin_border
        cell.font = data_font
        if ri % 2 == 0:
            cell.fill = alt_fill

for ci in range(1, len(cat_names) + 1):
    ws_cat.column_dimensions[get_column_letter(ci)].width = 28

ws_cat.freeze_panes = "A2"
print(f"  {len(cat_names)} catalogos")

# ── Pestaña DISENADORES ──
print("Creando DISENADORES...")
ws_dis = wb.create_sheet("DISENADORES")
dis_headers = ['NOMBRE', 'ROL', 'ACTIVO']
write_header_row(ws_dis, 1, dis_headers)

disenadores_all = set()
for src in [cats_B, cats_A_ws27, cats_A_fw26]:
    if 'DISENADOR_CREATIVO' in src:
        for v in src['DISENADOR_CREATIVO']:
            disenadores_all.add(NAME_NORMALIZE.get(v, v))
    if 'DISENADOR_TECNICO' in src:
        for v in src['DISENADOR_TECNICO']:
            disenadores_all.add(NAME_NORMALIZE.get(v, v))

creativos = set()
tecnicos = set()
for src in [cats_B, cats_A_ws27, cats_A_fw26]:
    if 'DISENADOR_CREATIVO' in src:
        for v in src['DISENADOR_CREATIVO']:
            creativos.add(NAME_NORMALIZE.get(v, v))
    if 'DISENADOR_TECNICO' in src:
        for v in src['DISENADOR_TECNICO']:
            tecnicos.add(NAME_NORMALIZE.get(v, v))

ri = 2
for nombre in sorted(disenadores_all):
    roles = []
    if nombre in creativos:
        roles.append('CREATIVO')
    if nombre in tecnicos:
        roles.append('TECNICO')
    write_data_row(ws_dis, ri, [nombre, ' / '.join(roles), 'SI'])
    ri += 1

auto_width(ws_dis)
print(f"  {len(disenadores_all)} disenadores")

# ── Pestaña MODISTAS ──
print("Creando MODISTAS...")
ws_mod = wb.create_sheet("MODISTAS")
mod_headers = ['NOMBRE', 'ACTIVO']
write_header_row(ws_mod, 1, mod_headers)

modistas_all = set()
for src in [cats_A_ws27]:
    if 'MODISTA' in src:
        for v in src['MODISTA']:
            modistas_all.add(v)

ri = 2
for nombre in sorted(modistas_all):
    write_data_row(ws_mod, ri, [nombre, 'SI'])
    ri += 1

auto_width(ws_mod)
print(f"  {len(modistas_all)} modistas")

# ── Pestaña LINEAS ──
print("Creando LINEAS...")
ws_lin = wb.create_sheet("LINEAS")
lin_headers = ['LINEA', 'SUBLINEA']
write_header_row(ws_lin, 1, lin_headers)

lineas_all = set()
sublineas_all = set()
for src in [cats_A_ws27, cats_A_fw26]:
    if 'LINEA' in src:
        lineas_all.update(src['LINEA'])
    if 'SUBLINEA' in src:
        sublineas_all.update(src['SUBLINEA'])

ri = 2
for linea in sorted(lineas_all):
    write_data_row(ws_lin, ri, [linea, ''])
    ri += 1

ri_sub_start = ri + 1
cell = ws_lin.cell(row=ri, column=1, value="SUBLINEAS (todas)")
cell.fill = accent_fill
cell.font = Font(bold=True, size=10, name="Calibri")
ri += 1
for sub in sorted(sublineas_all):
    write_data_row(ws_lin, ri, ['', sub])
    ri += 1

auto_width(ws_lin)
print(f"  {len(lineas_all)} lineas, {len(sublineas_all)} sublineas")

# ── Pestaña COLECCIONES ──
print("Creando COLECCIONES...")
ws_col2 = wb.create_sheet("COLECCIONES")
col2_headers = ['CODIGO_COLECCION', 'NOMBRE_COLECCION', 'TEMPORADA', 'ANIO', 'DRIVE_ID', 'ACTIVA']
write_header_row(ws_col2, 1, col2_headers)

colecciones = [
    ('WS27', 'WINTER SUN', 'WINTER SUN', '2027', '', 'SI'),
    ('FW26', 'FALL WINTER', 'FALL WINTER', '2026', '', 'SI'),
    ('RS26', 'RESORT', 'RESORT RTW', '2026', '', 'SI'),
    ('SS26', 'SPRING SUMMER', 'SPRING SUMMER', '2026', '', 'SI'),
    ('SV26', 'SUMMER VACATION', 'SUMMER VACATION', '2026', '', 'SI'),
    ('PF26', 'PREFALL', 'PREFALL RTW', '2026', '', 'SI'),
]

for ri, row_data in enumerate(colecciones, 2):
    write_data_row(ws_col2, ri, list(row_data))

auto_width(ws_col2)
print(f"  {len(colecciones)} colecciones")

# ── Pestaña FOTOS_MAESTRO ──
print("Creando FOTOS_MAESTRO...")
ws_fot = wb.create_sheet("FOTOS_MAESTRO")
fot_headers = [
    'ID_FOTO', 'TIPO', 'CODIGO_ASOCIADO', 'NOMBRE_ARCHIVO',
    'DRIVE_FILE_ID', 'URL_PUBLICA', 'URL_THUMBNAIL',
    'COLECCION', 'FECHA_CARGA', 'CARGADO_POR'
]
write_header_row(ws_fot, 1, fot_headers)

# Pre-populate with existing color/tela references from C
foto_ri = 2
for row_data in dataC:
    codigo = row_data.get('CODIGO', '').strip()
    cod_tela = row_data.get('CODIGO TELA', '').strip()
    coleccion = row_data.get('COLECCION', '').strip()

    if codigo:
        write_data_row(ws_fot, foto_ri, [
            f'COLOR-{codigo}', 'COLOR', codigo, '', '', '', '', coleccion, '', ''
        ])
        foto_ri += 1

    for link_col in ['LINK FINAL']:
        link = row_data.get(link_col, '').strip()
        if link:
            break

# Add template rows for prenda photos
for i in range(5):
    write_data_row(ws_fot, foto_ri + i, [
        '', 'PRENDA', '', '', '', '', '', '', '', ''
    ])

auto_width(ws_fot)
print(f"  {foto_ri - 2} registros pre-cargados (colores)")

# ── Pestaña CHANGE_LOG ──
print("Creando CHANGE_LOG...")
ws_log = wb.create_sheet("CHANGE_LOG")
log_headers = [
    'FECHA', 'USUARIO', 'HOJA_ORIGEN', 'ARCHIVO_ORIGEN',
    'REF', 'CELDA', 'COLUMNA_NOMBRE', 'VALOR_ANTERIOR',
    'VALOR_NUEVO', 'FASE', 'OBSERVACIONES'
]
write_header_row(ws_log, 1, log_headers)

# Add example row
write_data_row(ws_log, 2, [
    '2026-05-23 18:00', 'usuario@jo.com', 'WS27', 'APPAREL 2026',
    '1', 'K5', 'STATUS', 'EN PROCESO', 'APROBADO', 'FASE 1', ''
])

auto_width(ws_log)
print("  Plantilla CHANGE_LOG creada")

# ── Guardar ──
print(f"\nGuardando: {OUTPUT}")
wb.save(str(OUTPUT))

print(f"\n{'=' * 60}")
print(f"MAESTRO PARAMETROS JO - RESUMEN")
print(f"{'=' * 60}")
print(f"Archivo: {OUTPUT}")
print(f"Pestanas: {len(wb.sheetnames)}")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row} filas x {ws.max_column} cols")
print(f"\nCOLORES: {len(colores_merged)} registros")
print(f"TELAS: {len(telas_merged)} registros")
print(f"MATERIALES: {len(materiales)} registros")
print(f"CATALOGOS: {len(cat_names)} catalogos")
print(f"DISENADORES: {len(disenadores_all)} personas")
print(f"MODISTAS: {len(modistas_all)} personas")
print(f"LINEAS: {len(lineas_all)} / SUBLINEAS: {len(sublineas_all)}")
print(f"COLECCIONES: {len(colecciones)}")
print(f"='*60")
print("F2 COMPLETADO")
