"""VALIDACION DE TELAS v2.1 - Con fases COSTEO/CONTRAMUESTRA, celdas combinadas y mini-estructura por referencia"""
import sys, io, os, csv
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

BASE = r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0"
DIST = os.path.join(BASE, "dist")
CSV_PATH = os.path.join(BASE, "Documentos", "Copia de COLECCION WS27 - 1.VALIDACION DE TELAS.csv")

# ── Colors ───────────────────────────────────────────────
TITLE_BG = "1A1A2E"; TITLE_FG = "FFFFFF"
FASE1_BG = "2E86C1"; FASE1_FG = "FFFFFF"
FASE2_BG = "D4AC0D"; FASE2_FG = "2C3E50"
COSTEO_BG = "E67E22"; COSTEO_FG = "FFFFFF"
CONTRA_BG = "C0392B"; CONTRA_FG = "FFFFFF"
ALT_ROW   = "F8F9FA"; WHITE = "FFFFFF"
CANCEL_BG = "D5D8DC"; ALERT_BG = "FCF3CF"
INSTR_BG  = "EBF5FB"; SEP_BG = "DEE2E6"
INSUMO_BG = "E8F8F5"
THIN_C = "808080"; THICK_C = "000000"

thin = Border(left=Side('thin',THIN_C),right=Side('thin',THIN_C),top=Side('thin',THIN_C),bottom=Side('thin',THIN_C))
thick_bottom = Border(left=Side('thin',THIN_C),right=Side('thin',THIN_C),top=Side('thin',THIN_C),bottom=Side('medium',THICK_C))
thick_top = Border(left=Side('thin',THIN_C),right=Side('thin',THIN_C),top=Side('medium',THICK_C),bottom=Side('thin',THIN_C))
none_right = Border(left=Side('thin',THIN_C),bottom=Side('thin',THIN_C),top=Side('thin',THIN_C),right=Side(None))
c_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
c_center = Alignment(horizontal='center', vertical='center')
df = Font(name='Calibri', size=9, color='2C3E50')
dfs = Font(name='Calibri', size=8, color='2C3E50')

def col_idx(l): return openpyxl.utils.column_index_from_string(l)
def col_ltr(n): return get_column_letter(n)

# ── Column definitions ───────────────────────────────────
# Phase: (name, bg, fg, first_col_letter, last_col_letter)
PHASES = [
    ("FASE 1 - INFO BASICA",       FASE1_BG, FASE1_FG, "A","I"),
    ("FASE 2 - MATERIALES",        FASE2_BG, FASE2_FG, "J","O"),
    ("FASE 3 - COSTEO (Creativo + Tecnico + Trazo y Corte)", COSTEO_BG, COSTEO_FG, "P","BB"),
    ("FASE 4 - CONTRAMUESTRA (TEC + Trazador)", CONTRA_BG, CONTRA_FG, "BC","BN"),
]

# Columns to merge vertically per reference block (1-indexed spreadsheet cols)
MERGE_COLS = [1,2,3,4,5,6,7,8,9, 16,17,21, 22,38, 54,64]

# Data validation dropdowns (col number -> values)
DROPDOWNS = {
    4: ['APROBADO','CANCELADO','EN PROCESO','RECHAZADO','PENDIENTE','PAUSADO','TERMINADO'],
    5: ['SI','NO'], 6: ['SI','NO'],
    10: ['TELA LUCIR','TELA FORRO','FUSIONABLE','SESGOS LUCIR','SESGOS FORRO',
         'SESGOS FUSIONABLE','ENTRETELA','CINTA','RESORTE','HILO RESORTE',
         'CIERRE','ARGOLIA','BOTON','CORDON','BORLA','MARQUILLA','ELASTICO',
         'CINTA BOLILLO','CINTA ILUSION','CINTA RASSO','FRAMILON','VARILLA'],
    16: ['CLAUDIA','FER','MARGARITA','MARIA BURGOS','MARIA DEL MAR','OSMAN','YAMILETH'],
    17: ['SI','NO'],
    22: ['CAMILA VILLEGAS','CLAUDIA R','CRISTIAN GOMEZ','DANIELA GARCIA',
         'FERNANDO','KAROLINE','KELLY M.','LINA DELGADO','LINA P.','MARIA BURGOS'],
}

TOTAL_COLS = 66

# CSV to spreadsheet: identity map (CSV col 0->sheet 1, 1->2, ...)
CSV_MAP = {i: i+1 for i in range(66)}

# Column sub-headers (spreadsheet 1-indexed -> name)
SUB_HEADERS = {
    1:"REF",2:"FOTO",3:"REFERENCIA",4:"STATUS",5:"MOD.ARTE",6:"UBI.TRAZO",
    7:"VAR.COLOR 1",8:"VAR.COLOR 2",9:"LARGO",
    10:"USO PRENDA",11:"CODIGO TELA",12:"DESC TELA",13:"ANCHO",14:"TELA FOTO",15:"CONS BASE",
    16:"CREAT DIS",17:"CAMBIO MOLD",18:"CREAT C1",19:"CREAT C2",20:"CREAT C3",21:"CREAT OBS",
    22:"TECN DIS",
    23:"SOL TAL",24:"SOL UND",25:"SOL C1",26:"SOL C2",27:"SOL C3",
    28:"MA TAL",29:"MA UND",30:"MA C1",31:"MA C2",32:"MA C3",
    33:"UT TAL",34:"UT UND",35:"UT C1",36:"UT C2",37:"UT C3",38:"TEC OBS",
    39:"TR SOL T",40:"TR SOL U",41:"TR SOL C1",42:"TR SOL C2",43:"TR SOL C3",44:"TR SOL C4",
    45:"TR MA T",46:"TR MA U",47:"TR MA C1",48:"TR MA C2",49:"TR MA C3",50:"TR MA C4",
    51:"TR UT T",52:"TR UT U",53:"TR UT C1",54:"TR OBS",
    55:"CM SOL T",56:"CM SOL U",57:"CM SOL C1",
    58:"CM MA T",59:"CM MA U",60:"CM MA C1",
    61:"CM UT T",62:"CM UT U",63:"CM UT C1",
    64:"CM OBS",65:"TEC",66:"TRAZADOR",
}

# Columns to merge vertically per reference block (1-indexed spreadsheet cols)
MERGE_COLS = [1,2,3,4,5,6,7,8,9, 16,17,21, 22,38, 54,64]

# Data validation dropdowns
DROPDOWNS = {
    4: ['APROBADO','CANCELADO','EN PROCESO','RECHAZADO','PENDIENTE','PAUSADO','TERMINADO'],
    5: ['SI','NO'], 6: ['SI','NO'],
    10: ['TELA LUCIR','TELA FORRO','FUSIONABLE','SESGOS LUCIR','SESGOS FORRO',
         'SESGOS FUSIONABLE','ENTRETELA','CINTA','RESORTE','HILO RESORTE',
         'CIERRE','ARGOLIA','BOTON','CORDON','BORLA','MARQUILLA','ELASTICO',
         'CINTA BOLILLO','CINTA ILUSION','CINTA RASSO','FRAMILON','VARILLA'],
    16: ['CLAUDIA','FER','MARGARITA','MARIA BURGOS','MARIA DEL MAR','OSMAN','YAMILETH'],
    17: ['SI','NO'],
    22: ['CAMILA VILLEGAS','CLAUDIA R','CRISTIAN GOMEZ','DANIELA GARCIA',
         'FERNANDO','KAROLINE','KELLY M.','LINA DELGADO','LINA P.','MARIA BURGOS'],
}

# ── Load sample data from CSV ─────────────────────────────
def load_csv_samples():
    with open(CSV_PATH, encoding='utf-8-sig') as f:
        all_rows = list(csv.reader(f))
    refs = {}
    current_ref = None
    for i in range(4, len(all_rows)):
        r = all_rows[i]
        ref_val = (r[0] or '').strip()
        if ref_val and ref_val.isdigit():
            current_ref = ref_val
            refs[current_ref] = [r]
        elif current_ref and any((c or '').strip() for c in r if c):
            refs[current_ref].append(r)

    # Pick 3 diverse examples
    samples = []
    picked = set()
    for ref in ['1','10']:
        if ref in refs:
            samples.append((ref, refs[ref][:19]))
            picked.add(ref)
    # Add one more with data
    for ref in sorted(refs.keys(), key=lambda x: int(x)):
        if ref not in picked and len(refs[ref]) > 3 and any((refs[ref][0][c] or '').strip() for c in [9,10]):
            samples.append((ref, refs[ref][:19]))
            if len(samples) >= 3:
                break
    return samples

def safe(v):
    if v is None or str(v).strip() in ('','nan','#N/A','N/A','#REF!','#¡REF!'):
        return ''
    return str(v).strip()

# ── BUILD WORKBOOK ────────────────────────────────────────
print("Cargando datos de ejemplo...")
samples = load_csv_samples()
print(f"  {len(samples)} referencias de ejemplo cargadas")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "1.VALIDACION DE TELAS"

BLOQUE_FILAS = 19
DATA_START = 5  # Row 5 = first data row

# ── ROW 1: TITLE ──────────────────────────────────────────
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
c = ws.cell(1, 1, "VALIDACION DE TELAS - COLECCION WS27  |  v2.1")
c.font = Font(name='Calibri', size=16, bold=True, color=TITLE_FG)
c.fill = PatternFill(start_color=TITLE_BG, end_color=TITLE_BG, fill_type="solid")
c.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[1].height = 38

# ── ROW 2: QUICK INSTRUCTIONS ─────────────────────────────
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
c = ws.cell(2, 1,
    "Complete solo las columnas de SU area. Cada referencia ocupa 19 filas (13 telas + 1 separador + 5 insumos). "
    "Las celdas en gris son de la referencia anterior. NO elimine filas ni columnas.")
c.font = Font(name='Calibri', size=9, color='546E7A', italic=True)
c.fill = PatternFill(start_color=INSTR_BG, end_color=INSTR_BG, fill_type="solid")
c.alignment = Alignment(vertical='center', wrap_text=True)
ws.row_dimensions[2].height = 24

# ── ROW 3: PHASE HEADERS ──────────────────────────────────
ws.row_dimensions[3].height = 30
for name, bg, fg, first, last in PHASES:
    fc = col_idx(first); lc = col_idx(last)
    ws.merge_cells(start_row=3, start_column=fc, end_row=3, end_column=lc)
    c = ws.cell(3, fc, name)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.font = Font(name='Calibri', size=10, bold=True, color=fg)
    c.alignment = c_wrap
    c.border = thin

# ── ROW 4: SUB-HEADERS ────────────────────────────────────
ws.row_dimensions[4].height = 40
for ci in range(1, TOTAL_COLS + 1):
    name = SUB_HEADERS.get(ci, '')
    c = ws.cell(4, ci, name)
    c.font = Font(name='Calibri', size=7, bold=True, color='1A1A2E')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=90)
    c.border = thin
    c.fill = PatternFill(start_color="E5E8E8", end_color="E5E8E8", fill_type="solid")

# Column widths
for ci in range(1, TOTAL_COLS + 1):
    ws.column_dimensions[col_ltr(ci)].width = 5.5
wide_cols = {1:6, 2:5, 3:32, 9:8, 10:16, 12:28, 16:13, 21:30, 22:13, 38:30, 54:30, 64:30}
for ci, w in wide_cols.items():
    ws.column_dimensions[col_ltr(ci)].width = w

# ── DATA ROWS with reference blocks ───────────────────────
TOTAL_REFS = 5  # 3 sample + 2 empty = 95 rows

for ref_idx in range(TOTAL_REFS):
    block_start = DATA_START + ref_idx * BLOQUE_FILAS
    block_end = block_start + BLOQUE_FILAS - 1
    sample_data = None
    if ref_idx < len(samples):
        sample_data = samples[ref_idx][1]  # list of csv rows

    for local_row in range(BLOQUE_FILAS):
        ri = block_start + local_row
        ws.row_dimensions[ri].height = 16

        for ci in range(1, TOTAL_COLS + 1):
            cell = ws.cell(ri, ci)
            cell.font = dfs
            cell.alignment = c_center
            cell.border = thin

            # Fill with sample data if available
            if sample_data and local_row < len(sample_data):
                csv_row = sample_data[local_row]
                for csv_col, sheet_col in CSV_MAP.items():
                    if csv_col < len(csv_row):
                        val = safe(csv_row[csv_col])
                        if val:
                            ws.cell(ri, sheet_col).value = val

            # Row coloring
            is_insumo_row = local_row >= 14
            is_separator = local_row == 13

            if is_separator:
                cell.fill = PatternFill(start_color=SEP_BG, end_color=SEP_BG, fill_type="solid")
                cell.border = Border(top=Side('dotted', THIN_C), bottom=Side('dotted', THIN_C))
            elif is_insumo_row:
                if (local_row % 2) == 0:
                    cell.fill = PatternFill(start_color=INSUMO_BG, end_color=INSUMO_BG, fill_type="solid")
            else:  # tela rows
                if (local_row % 2) == 1:
                    cell.fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")

        # Separator label
        if local_row == 13:
            ws.cell(ri, 10).value = "── INSUMOS ──"
            ws.cell(ri, 10).font = Font(name='Calibri', size=8, italic=True, color='95A5A6')

    # ── MERGE CELLS for this reference block ──────────────
    for col_num in MERGE_COLS:
        ws.merge_cells(start_row=block_start, start_column=col_num,
                       end_row=block_end, end_column=col_num)

    # ── Border between blocks ─────────────────────────────
    for ci in range(1, TOTAL_COLS + 1):
        c = ws.cell(block_start, ci)
        c.border = Border(top=Side('medium', '2C3E50'), bottom=c.border.bottom,
                          left=c.border.left, right=c.border.right)

# Add final bottom border
last_row = DATA_START + TOTAL_REFS * BLOQUE_FILAS - 1
for ci in range(1, TOTAL_COLS + 1):
    c = ws.cell(last_row, ci)
    c.border = Border(bottom=Side('medium', THICK_C), left=c.border.left,
                      right=c.border.right, top=c.border.top)

# ── DATA VALIDATIONS ──────────────────────────────────────
print("  Configurando data validations...")
for col_num, options in DROPDOWNS.items():
    col_l = col_ltr(col_num)
    formula = '"' + ','.join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Valor no valido. Seleccione de la lista."
    dv.errorTitle = "Error"
    dv.prompt = "Seleccione una opcion"
    dv.promptTitle = "Valores permitidos"
    ws.add_data_validation(dv)
    # Apply to all rows
    for ref_idx in range(TOTAL_REFS):
        bs = DATA_START + ref_idx * BLOQUE_FILAS
        be = bs + BLOQUE_FILAS - 1
        dv.add(f"{col_l}{bs}:{col_l}{be}")

# ── CONDITIONAL FORMATTING ────────────────────────────────
print("  Configurando formato condicional...")
data_rng = f"A{DATA_START}:{col_ltr(TOTAL_COLS)}{last_row}"

ws.conditional_formatting.add(data_rng, FormulaRule(
    formula=['$D5="CANCELADO"'],
    fill=PatternFill(start_color=CANCEL_BG, end_color=CANCEL_BG, fill_type="solid"),
    font=Font(name='Calibri', size=8, color='808080', strike=True)))

ws.conditional_formatting.add(f"D{DATA_START}:D{last_row}", FormulaRule(
    formula=['AND($A5<>"",$D5="")'],
    fill=PatternFill(start_color=ALERT_BG, end_color=ALERT_BG, fill_type="solid"),
    font=Font(name='Calibri', size=9, color='C0392B', bold=True)))

ws.conditional_formatting.add(f"A{DATA_START}:A{last_row}", FormulaRule(
    formula=['$D5="APROBADO"'],
    fill=PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type="solid"),
    font=Font(name='Calibri', size=9, color='1E8449', bold=True)))

# ── FREEZE ────────────────────────────────────────────────
ws.freeze_panes = "A5"
ws.sheet_properties.tabColor = "1A1A2E"

# ── INSTRUCCIONES SHEET ───────────────────────────────────
print("  Creando INSTRUCCIONES...")
ws2 = wb.create_sheet("INSTRUCCIONES")
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 85
ws2.sheet_properties.tabColor = "2980B9"

r = 1
ws2.cell(r,2,"GUIA DE USO - VALIDACION DE TELAS v2.1").font = Font(size=16,bold=True,color='1A1A2E',name='Calibri')
r += 2

items = [
    ("ESTRUCTURA","Cada referencia ocupa un BLOQUE de 19 filas:\n  Filas 1-13: TELAS (una fila por cada tela o material usado)\n  Fila 14: SEPARADOR 'INSUMOS'\n  Filas 15-19: INSUMOS (cintas, elasticos, argollas, etc.)\n\nSi una referencia necesita MAS telas o insumos, inserte filas DENTRO del bloque."),
    ("FASES Y COLORES","  FASE 1 (Azul): Informacion basica - REF, STATUS, MOD ARTE, UBI TRAZO\n  FASE 2 (Amarillo): Materiales - USO EN PRENDA, CODIGO TELA, DESCRIPCION\n  FASE 3 (Naranja): COSTEO - Creativo + Tecnico + Trazo y Corte\n  FASE 4 (Rojo): CONTRAMUESTRA - Consumos finales para produccion"),
    ("CELDAS COMBINADAS","Las columnas REF, FOTO, REFERENCIA, STATUS, MOD ARTE, UBI TRAZO, DISENADOR CREATIVO, DISENADOR TECNICO y OBSERVACIONES estan COMBINADAS verticalmente en las 19 filas del bloque. Su valor se comparte en todo el bloque."),
    ("COMPATIBILIDAD","Este formato es 100% compatible con migracion.gs.\nPosiciones de columna identicas: C.REF:0, C.FOTO:1, C.STATUS:3, etc."),
    ("NO HACER","NO reordene columnas.\nNO elimine filas del encabezado (1-4).\nNO inserte columnas entre A y BN.\nPara agregar telas/insumos a una referencia, inserte filas DENTRO del bloque."),
]
for title, content in items:
    ws2.cell(r,2,title).font = Font(name='Calibri',size=11,bold=True,color='E74C3C')
    ws2.cell(r+1,2,content).font = Font(name='Calibri',size=10,color='2C3E50')
    ws2.cell(r+1,2).alignment = Alignment(wrap_text=True, vertical='top')
    ws2.row_dimensions[r+1].height = max(40, content.count('\n')*18 + 25)
    r += 3

# ── SAVE ──────────────────────────────────────────────────
output = os.path.join(DIST, "VALIDACION_DE_TELAS_v2.1.xlsx")
wb.save(output)
print(f"\nGuardado: {output}")
print(f"Pestañas: {wb.sheetnames}")
print(f"Referencias de ejemplo: {len(samples)}")
print(f"Total filas datos: {TOTAL_REFS * BLOQUE_FILAS}")
print("VALIDACION DE TELAS v2.1 COMPLETADO")
