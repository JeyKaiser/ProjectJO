"""
OPCION 1: Crear PARAMETROS_NUEVA para los 3 archivos.
Cada .xlsx tiene 2 pestanas: INSTRUCCIONES + PARAMETROS_NUEVA.
La pestana PARAMETROS_NUEVA contiene la estructura de columnas
con placeholders para IMPORTRANGE (que se activan al subir a Google Sheets).
"""

import sys, io
from pathlib import Path
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0")
DIST = BASE / "dist"
OP1 = DIST / "opcion1-solo-parametros"
MAESTRO = DIST / "MAESTRO_PARAMETROS_JO.xlsx"

thin = Border(
    left=Side(style='thin', color='808080'),
    right=Side(style='thin', color='808080'),
    top=Side(style='thin', color='808080'),
    bottom=Side(style='thin', color='808080')
)
hdr_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
alt_fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
accent_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
imp_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

def safe(v):
    if v is None or str(v).strip() in ('', 'nan', '#N/A', '#REF!', '#REF! '):
        return ''
    return str(v).strip()

def write_instrucciones(ws, archivo_nombre, maestro_id_placeholder):
    ws.sheet_properties.tabColor = "2980B9"
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 80

    r = 1
    ws.cell(r, 2, "INSTRUCCIONES - PARAMETROS NUEVA").font = Font(size=14, bold=True, color="2C3E50", name="Calibri")
    r += 1
    ws.cell(r, 2, f"Archivo: {archivo_nombre}").font = Font(size=10, color="7F8C8D", name="Calibri")
    r += 1
    ws.cell(r, 2, f"Fecha: {datetime.now().strftime('%Y-%m-%d')}").font = Font(size=10, color="7F8C8D", name="Calibri")
    r += 2

    pasos = [
        ("PASO 1", "Sube este archivo .xlsx a Google Drive (junto al archivo original)"),
        ("PASO 2", "Abre el archivo original de Google Sheets (el que ya existe)"),
        ("PASO 3", "Renombra la pestana PARAMETROS actual a PARAMETROS_OLD"),
        ("PASO 4", "Abre este .xlsx en Google Sheets (se convertira automaticamente)"),
        ("PASO 5", "Copia TODA la pestana PARAMETROS_NUEVA (click derecho > Copiar a > Hoja de calculo existente)"),
        ("PASO 6", "Selecciona el archivo original como destino. La pestana se agrega automaticamente"),
        ("PASO 7", "En la nueva PARAMETROS_NUEVA, reemplaza 'PEGA_ID_MAESTRO_AQUI' por el ID real del Maestro"),
        ("PASO 8", "Cada celda con formula IMPORTRANGE mostrara #REF! - haz click y presiona 'Conceder acceso'"),
        ("PASO 9", "Actualiza las validaciones de datos (dropdowns) en la MATRIZ para que apunten a PARAMETROS_NUEVA"),
        ("PASO 10", "Verifica que los datos del Maestro aparecen correctamente"),
    ]
    for paso, desc in pasos:
        ws.cell(r, 2, paso).font = Font(bold=True, size=10, color="E74C3C", name="Calibri")
        ws.cell(r+1, 2, desc).font = Font(size=10, name="Calibri")
        ws.cell(r+1, 2).alignment = Alignment(wrap_text=True)
        r += 3

    r += 1
    ws.cell(r, 2, "IMPORTANTE: No elimines PARAMETROS_OLD hasta verificar que todo funciona.").font = Font(
        bold=True, size=10, color="E74C3C", name="Calibri")
    r += 2
    ws.cell(r, 2, "ID del Maestro (pegar aqui):").font = Font(bold=True, size=10, name="Calibri")
    ws.cell(r+1, 2, maestro_id_placeholder).font = Font(size=12, bold=True, color="E74C3C", name="Calibri")
    ws.cell(r+1, 2).fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")


def load_maestro_catalogos():
    wb = openpyxl.load_workbook(str(MAESTRO), read_only=True)
    ws = wb["CATALOGOS"]
    headers = []
    for cell in ws[1]:
        headers.append(cell.value or '')
    wb.close()
    return headers


def write_parametros_a(ws):
    ws.sheet_properties.tabColor = "27AE60"
    MID = 'PEGA_ID_MAESTRO_AQUI'

    sections = [
        ("A", "COLORES (desde Maestro)", [
            ("A", "CODIGO", f'=IMPORTRANGE("{MID}","COLORES!A:A")'),
            ("B", "DESCRIPCION", f'=IMPORTRANGE("{MID}","COLORES!B:B")'),
            ("C", "FUENTE_ORIGEN", f'=IMPORTRANGE("{MID}","COLORES!C:C")'),
        ]),
        ("D", "TELAS (desde Maestro)", [
            ("D", "CODIGO_TELA", f'=IMPORTRANGE("{MID}","TELAS!A:A")'),
            ("E", "DESCRIPCION_TELA", f'=IMPORTRANGE("{MID}","TELAS!B:B")'),
            ("F", "ANCHO", f'=IMPORTRANGE("{MID}","TELAS!C:C")'),
            ("G", "OBSERVACIONES", f'=IMPORTRANGE("{MID}","TELAS!D:D")'),
            ("H", "COLECCION", f'=IMPORTRANGE("{MID}","TELAS!E:E")'),
            ("I", "FUENTE_ORIGEN", f'=IMPORTRANGE("{MID}","TELAS!F:F")'),
        ]),
        ("J", "BASADO EN (desde Maestro)", [
            ("J", "BASADO_EN", f'=IMPORTRANGE("{MID}","COLORES!D:D")'),
            ("K", "FOTO_BASADO_EN", f'=IMPORTRANGE("{MID}","COLORES!E:E")'),
            ("L", "PT_BASADO_EN", f'=IMPORTRANGE("{MID}","COLORES!F:F")'),
        ]),
        ("M", "CATALOGOS DESPLEGABLES (desde Maestro)", [
            ("M", "ALL_OVER", f'=IMPORTRANGE("{MID}","CATALOGOS!A:A")'),
            ("N", "BASE_TEXTIL", f'=IMPORTRANGE("{MID}","CATALOGOS!B:B")'),
            ("O", "CAMBIO_MOLDERIA", f'=IMPORTRANGE("{MID}","CATALOGOS!C:C")'),
            ("P", "CATALOGACION", f'=IMPORTRANGE("{MID}","CATALOGOS!D:D")'),
            ("Q", "CLOSURE", f'=IMPORTRANGE("{MID}","CATALOGOS!E:E")'),
            ("R", "CLASIFICACION_HALLAZGO", f'=IMPORTRANGE("{MID}","CATALOGOS!F:F")'),
            ("S", "DISENADOR_CREATIVO", f'=IMPORTRANGE("{MID}","CATALOGOS!G:G")'),
            ("T", "DISENADOR_TECNICO", f'=IMPORTRANGE("{MID}","CATALOGOS!H:H")'),
            ("U", "MODISTA", f'=IMPORTRANGE("{MID}","CATALOGOS!I:I")'),
            ("V", "PRIORIDADES", f'=IMPORTRANGE("{MID}","CATALOGOS!J:J")'),
            ("W", "SI_NO", f'=IMPORTRANGE("{MID}","CATALOGOS!K:K")'),
            ("X", "STATUS_PROCESO_EXT", f'=IMPORTRANGE("{MID}","CATALOGOS!L:L")'),
            ("Y", "STATUS", f'=IMPORTRANGE("{MID}","CATALOGOS!X:X")'),
            ("Z", "STATUS_TALLER", f'=IMPORTRANGE("{MID}","CATALOGOS!Y:Y")'),
            ("AA", "SUBLINEA", f'=IMPORTRANGE("{MID}","CATALOGOS!M:M")'),
            ("AB", "TALLAJE", f'=IMPORTRANGE("{MID}","CATALOGOS!Z:Z")'),
            ("AC", "TIPO_BORDADO", f'=IMPORTRANGE("{MID}","CATALOGOS!AA:AA")'),
            ("AD", "TIPO_CORTE", f'=IMPORTRANGE("{MID}","CATALOGOS!AB:AB")'),
            ("AE", "TIPO_PROCESO_EXT", f'=IMPORTRANGE("{MID}","CATALOGOS!AC:AC")'),
            ("AF", "TIPO_RECHAZO", f'=IMPORTRANGE("{MID}","CATALOGOS!AD:AD")'),
            ("AG", "TIPO_REF", f'=IMPORTRANGE("{MID}","CATALOGOS!AE:AE")'),
            ("AH", "TIPO_TEJIDO", f'=IMPORTRANGE("{MID}","CATALOGOS!AF:AF")'),
            ("AI", "UBI_TRAZO", f'=IMPORTRANGE("{MID}","CATALOGOS!AG:AG")'),
            ("AJ", "USO_PRENDA", f'=IMPORTRANGE("{MID}","CATALOGOS!AH:AH")'),
            ("AK", "VEREDICTO_DIR", f'=IMPORTRANGE("{MID}","CATALOGOS!AI:AI")'),
            ("AL", "WOVEN_KNITTED", f'=IMPORTRANGE("{MID}","CATALOGOS!AJ:AJ")'),
            ("AM", "DIFICULTAD", f'=IMPORTRANGE("{MID}","CATALOGOS!O:O")'),
            ("AN", "DROPS", f'=IMPORTRANGE("{MID}","CATALOGOS!P:P")'),
            ("AO", "INCLUDES", f'=IMPORTRANGE("{MID}","CATALOGOS!Q:Q")'),
            ("AP", "LARGO", f'=IMPORTRANGE("{MID}","CATALOGOS!R:R")'),
            ("AQ", "LINEA", f'=IMPORTRANGE("{MID}","CATALOGOS!S:S")'),
            ("AR", "LINNED", f'=IMPORTRANGE("{MID}","CATALOGOS!T:T")'),
            ("AS", "CONFECCION_STATUS", f'=IMPORTRANGE("{MID}","CATALOGOS!N:N")'),
            ("AT", "SENTIDO_SESGO", f'=IMPORTRANGE("{MID}","CATALOGOS!W:W")'),
            ("AU", "DISENADORES", f'=IMPORTRANGE("{MID}","DISENADORES!A:C")'),
            ("AV", "MODISTAS", f'=IMPORTRANGE("{MID}","MODISTAS!A:B")'),
            ("AW", "LINEAS_SUBLINEAS", f'=IMPORTRANGE("{MID}","LINEAS!A:B")'),
            ("AX", "COLECCIONES", f'=IMPORTRANGE("{MID}","COLECCIONES!A:F")'),
            ("AY", "FOTOS_MAESTRO", f'=IMPORTRANGE("{MID}","FOTOS_MAESTRO!A:J")'),
        ]),
    ]

    for sec_start, sec_name, cols in sections:
        r1 = 1
        start_col_idx = openpyxl.utils.column_index_from_string(sec_start)
        end_col_idx = start_col_idx + len(cols) - 1
        end_col_letter = get_column_letter(end_col_idx)
        ws.merge_cells(start_row=1, start_column=start_col_idx, end_row=1, end_column=end_col_idx)
        cell = ws.cell(1, start_col_idx, sec_name)
        cell.fill = accent_fill
        cell.font = Font(bold=True, size=11, color="000000", name="Calibri")
        cell.alignment = center
        cell.border = thin

        for col_letter, col_name, formula in cols:
            ci = openpyxl.utils.column_index_from_string(col_letter)
            cell = ws.cell(2, ci, col_name)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = thin

            cell = ws.cell(3, ci, formula)
            cell.fill = imp_fill
            cell.font = Font(size=9, color="2980B9", name="Calibri")
            cell.alignment = left
            cell.border = thin

            ws.column_dimensions[col_letter].width = 22

    ws.freeze_panes = "A3"


def write_parametros_b(ws):
    ws.sheet_properties.tabColor = "27AE60"
    MID = 'PEGA_ID_MAESTRO_AQUI'

    sections = [
        ("A", "CATALOGOS DESPLEGABLES (desde Maestro)", [
            ("A", "ALL_OVER", f'=IMPORTRANGE("{MID}","CATALOGOS!A:A")'),
            ("B", "BASE_TEXTIL", f'=IMPORTRANGE("{MID}","CATALOGOS!B:B")'),
            ("C", "CAMBIO_MOLDERIA", f'=IMPORTRANGE("{MID}","CATALOGOS!C:C")'),
            ("D", "CATALOGACION", f'=IMPORTRANGE("{MID}","CATALOGOS!D:D")'),
            ("E", "CLOSURE", f'=IMPORTRANGE("{MID}","CATALOGOS!E:E")'),
            ("F", "CLASIFICACION_HALLAZGO", f'=IMPORTRANGE("{MID}","CATALOGOS!F:F")'),
            ("G", "DISENADOR_CREATIVO", f'=IMPORTRANGE("{MID}","CATALOGOS!G:G")'),
            ("H", "DISENADOR_TECNICO", f'=IMPORTRANGE("{MID}","CATALOGOS!H:H")'),
            ("I", "MODISTA", f'=IMPORTRANGE("{MID}","CATALOGOS!I:I")'),
            ("J", "PRIORIDADES", f'=IMPORTRANGE("{MID}","CATALOGOS!J:J")'),
            ("K", "SI_NO", f'=IMPORTRANGE("{MID}","CATALOGOS!K:K")'),
            ("L", "STATUS_PROCESO_EXT", f'=IMPORTRANGE("{MID}","CATALOGOS!L:L")'),
            ("M", "STATUS", f'=IMPORTRANGE("{MID}","CATALOGOS!X:X")'),
            ("N", "STATUS_TALLER", f'=IMPORTRANGE("{MID}","CATALOGOS!Y:Y")'),
            ("O", "SUBLINEA", f'=IMPORTRANGE("{MID}","CATALOGOS!M:M")'),
            ("P", "TALLAJE", f'=IMPORTRANGE("{MID}","CATALOGOS!Z:Z")'),
            ("Q", "TIPO_BORDADO", f'=IMPORTRANGE("{MID}","CATALOGOS!AA:AA")'),
            ("R", "TIPO_CORTE", f'=IMPORTRANGE("{MID}","CATALOGOS!AB:AB")'),
            ("S", "TIPO_PROCESO_EXT", f'=IMPORTRANGE("{MID}","CATALOGOS!AC:AC")'),
            ("T", "TIPO_RECHAZO", f'=IMPORTRANGE("{MID}","CATALOGOS!AD:AD")'),
            ("U", "TIPO_REF", f'=IMPORTRANGE("{MID}","CATALOGOS!AE:AE")'),
            ("V", "TIPO_TEJIDO", f'=IMPORTRANGE("{MID}","CATALOGOS!AF:AF")'),
            ("W", "UBI_TRAZO", f'=IMPORTRANGE("{MID}","CATALOGOS!AG:AG")'),
            ("X", "USO_PRENDA", f'=IMPORTRANGE("{MID}","CATALOGOS!AH:AH")'),
            ("Y", "VEREDICTO_DIR", f'=IMPORTRANGE("{MID}","CATALOGOS!AI:AI")'),
            ("Z", "WOVEN_KNITTED", f'=IMPORTRANGE("{MID}","CATALOGOS!AJ:AJ")'),
            ("AA", "DIFICULTAD", f'=IMPORTRANGE("{MID}","CATALOGOS!O:O")'),
            ("AB", "DROPS", f'=IMPORTRANGE("{MID}","CATALOGOS!P:P")'),
            ("AC", "INCLUDES", f'=IMPORTRANGE("{MID}","CATALOGOS!Q:Q")'),
            ("AD", "LARGO", f'=IMPORTRANGE("{MID}","CATALOGOS!R:R")'),
            ("AE", "LINEA", f'=IMPORTRANGE("{MID}","CATALOGOS!S:S")'),
            ("AF", "LINNED", f'=IMPORTRANGE("{MID}","CATALOGOS!T:T")'),
            ("AG", "CONFECCION_STATUS", f'=IMPORTRANGE("{MID}","CATALOGOS!N:N")'),
            ("AH", "SENTIDO_SESGO", f'=IMPORTRANGE("{MID}","CATALOGOS!W:W")'),
            ("AI", "DISENADORES", f'=IMPORTRANGE("{MID}","DISENADORES!A:C")'),
            ("AJ", "MODISTAS", f'=IMPORTRANGE("{MID}","MODISTAS!A:B")'),
            ("AK", "LINEAS_SUBLINEAS", f'=IMPORTRANGE("{MID}","LINEAS!A:B")'),
            ("AL", "COLECCIONES", f'=IMPORTRANGE("{MID}","COLECCIONES!A:F")'),
            ("AM", "MATERIALES", f'=IMPORTRANGE("{MID}","MATERIALES!A:U")'),
        ]),
    ]

    for sec_start, sec_name, cols in sections:
        start_col_idx = openpyxl.utils.column_index_from_string(sec_start)
        end_col_idx = start_col_idx + len(cols) - 1
        end_col_letter = get_column_letter(end_col_idx)
        ws.merge_cells(start_row=1, start_column=start_col_idx, end_row=1, end_column=end_col_idx)
        cell = ws.cell(1, start_col_idx, sec_name)
        cell.fill = accent_fill
        cell.font = Font(bold=True, size=11, color="000000", name="Calibri")
        cell.alignment = center
        cell.border = thin

        for col_letter, col_name, formula in cols:
            ci = openpyxl.utils.column_index_from_string(col_letter)
            cell = ws.cell(2, ci, col_name)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = thin

            cell = ws.cell(3, ci, formula)
            cell.fill = imp_fill
            cell.font = Font(size=9, color="2980B9", name="Calibri")
            cell.alignment = left
            cell.border = thin

            ws.column_dimensions[col_letter].width = 22

    ws.freeze_panes = "A3"


def write_parametros_c(ws):
    write_parametros_a(ws)


print("=" * 50)
print("OPCION 1: PARAMETROS_NUEVA para A, B, C")
print("=" * 50)

MID = 'PEGA_ID_MAESTRO_AQUI'

for nombre, label, writer_fn in [
    ("PARAMETROS_NUEVO_A_APPAREL.xlsx", "FO APPAREL 2026", write_parametros_a),
    ("PARAMETROS_NUEVO_B_COLECCION.xlsx", "COLECCION WS27", write_parametros_b),
    ("PARAMETROS_NUEVO_C_CONTRAM.xlsx", "CONTRAM WS27", write_parametros_c),
]:
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "INSTRUCCIONES"
    write_instrucciones(ws0, label, MID)

    ws1 = wb.create_sheet("PARAMETROS_NUEVA")
    writer_fn(ws1)

    path = OP1 / nombre
    wb.save(str(path))
    print(f"  Creado: {path.name}")

print("OPCION 1 COMPLETADA")
