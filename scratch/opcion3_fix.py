
"""OPCION 3: Estructura vacia + PARAMETROS completo"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE = Path(r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0")
OP3 = BASE / "dist" / "opcion3-estructura-vacia"
MID = "PEGA_ID_MAESTRO_AQUI"

thin = Border(left=Side(style='thin', color='808080'), right=Side(style='thin', color='808080'), top=Side(style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
hdr_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True, size=10, name='Calibri')
sec_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
sec_font = Font(color='FFFFFF', bold=True, size=11, name='Calibri')
accent_fill = PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid')
imp_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
data_font = Font(name='Calibri', size=9)

def ci(l): return openpyxl.utils.column_index_from_string(l)

HEADERS_FO = [
    ('A', 'INFORMACION BASICA', [('A','Ref'),('B','Imagen'),('C','Codigo MD'),('D','Codigo PT'),('E','Nombre referencia'),('F','Color'),('G','Cod color')]),
    ('H', 'REFERENTES', [('H','Foto basado en'),('I','PT referente'),('J','Basado en')]),
    ('K', 'STATUS COLECCION', [('K','Status'),('L','Disenador'),('M','Status taller'),('N','Modista'),('O','Fotos internas')]),
    ('P', 'CARACTERISTICAS', [('P','Linea'),('Q','Sublinea'),('R','Tipo ref'),('S','Tallaje'),('T','Largo'),('U','Closure'),('V','Largo 2')]),
    ('W', 'INCLUDES', [('W','Includes'),('X','Includes paq completo')]),
    ('Y', 'TELA LUCIR', [('Y','Cod tela lucir'),('Z','Foto tela'),('AA','Descr tela'),('AB','Ancho tela'),('AC','Base textil'),('AD','Ubi trazo'),('AE','Mod arte'),('AF','All over')]),
    ('AG', 'VARIACION COLOR', [('AG','Variacion color'),('AH','Ref variacion')]),
    ('AI', 'EMPAQUES', [('AI','Tipo empaque')]),
    ('AJ', 'BORDADO PRENDA', [('AJ','Aplica bordado'),('AK','Descr bordado')]),
    ('AL', 'SEMIELABORADOS', [('AL','Aplica semiel'),('AM','Descr semiel')]),
    ('AN', 'PROCESO EXTERNO', [('AN','Proveedor'),('AO','Proceso ext'),('AP','Costo')]),
    ('AQ', 'ESTADOS Y ENTREGAS', [('AQ','Bordado status'),('AR','Term taller costeo'),('AS','Entregable creativo'),('AT','Entregable tecnico'),('AU','Entregable trazador'),('AV','Envio MOD arte')]),
    ('AW', 'TELAS EXTRA', [('AW','Comentarios telas'),('AX','Bordado status t'),('AY','Term taller t')]),
    ('AZ', 'INSUMOS EXTRA', [('AZ','Proc ext insumos'),('BA','Costo insumos')]),
    ('BB', 'BORDADO EXTRA', [('BB','Bordado status 2'),('BC','Term taller bordado')]),
    ('BD', 'ENTREGAS PARCIALES', [('BD','Entrega parcial fecha')]),
    ('BE', 'TIME COLLECTION', [('BE','Espec confec #1'),('BF','Escalado mold #1'),('BG','Tiras continuas'),('BH','Dificultad prenda'),('BI','Dificultad bordado'),('BJ','Prioridad first buy'),('BK','Prioridad bordado'),('BL','Bordado tipo'),('BM','Prioridad textil'),('BN','Comentarios ing'),('BO','Comentarios trazo'),('BP','Comentarios costeo'),('BQ','Sugerencia MOD'),('BR','Requiere muestra'),('BS','Grupo estilo')]),
    ('BT', 'VALIDACION MP', [('BT','Fecha'),('BU','Area afectacion'),('BV','Clasif hallazgo'),('BW','MP'),('BX','Clasif MP'),('BY','Tipo ejecucion')]),
    ('BZ', 'COMP MUESTRA', [('BZ','Info SAP'),('CA','Desc USA-UK'),('CB','Fiber Comp'),('CC','Woven'),('CD','Inside'),('CE','Include'),('CF','Obs comp')]),
    ('CG', 'COMP PRODUCCION', [('CG','Info SAP prod'),('CH','Desc USA-UK prod'),('CI','Fiber Comp prod'),('CJ','Woven prod'),('CK','Inside prod'),('CL','Include prod'),('CM','Obs comp prod')]),
    ('CN', 'COMEX', [('CN','Comex')]),
    ('CO', 'CUIDADOS', [('CO','Proceso'),('CP','Lavado'),('CQ','Logo lav'),('CR','Desmanche'),('CS','Logo desm'),('CT','Secado'),('CU','Logo sec'),('CV','Planchado'),('CW','Logo plan'),('CX','Cuid includes')]),
    ('CY', 'UNIDADES PRODUCCION', [('CY','0'),('CZ','2'),('DA','4'),('DB','6'),('DC','8'),('DD','10'),('DE','12'),('DF','XS'),('DG','S'),('DH','M'),('DI','L'),('DJ','XL'),('DK','TOTAL')]),
    ('DL', 'MAQUILA', [('DL','Tipo tejido'),('DM','Compl corte'),('DN','Envio corte'),('DO','Compl confec'),('DP','Envio confec')]),
    ('DQ', 'MONTAJE MANIQUI', [('DQ','Tipo montaje'),('DR','(vacia)'),('DS','Proy montaje')]),
    ('DT', 'DIS CONTRA MUESTRA', [('DT','Dis tecnico CM'),('DU','Dis creativo CM')]),
    ('DV', 'MOLDERIA', [('DV','Fecha inicio'),('DW','Fecha fin'),('DX','Comentarios')]),
    ('DY', 'PROCESOS EXT', [('DY','Tipo proc'),('DZ','Fec recibido'),('EA','Fec entrega'),('EB','Status proc')]),
    ('EC', 'CORTE CM', [('EC','Fec corte 1'),('ED','Tipo corte 1'),('EE','Fec corte 2'),('EF','Tipo corte 2'),('EG','Fec corte 3'),('EH','Tipo corte 3'),('EI','Fec corte 4'),('EJ','Tipo corte 4'),('EK','Obs corte'),('EL','Total piezas'),('EM','Total muestras')]),
    ('EN', 'CONFECCION', [('EN','Modista'),('EO','Fec inicio'),('EP','Fec entrega'),('EQ','Status'),('ER','Obs modista'),('ES','Obs tec 1'),('ET','Obs tec 2'),('EU','Tiempo min'),('EV','Estado prenda'),('EW','Tipo rechazo'),('EX','Feedback prod')]),
    ('EY', 'MEDICION', [('EY','Fase 1'),('EZ','Com fase 1'),('FA','Fase 2'),('FB','Com fase 2'),('FC','Fase 3'),('FD','Com fase 3'),('FE','Fase 4'),('FF','Com fase 4'),('FG','Fase 5'),('FH','Com fase 5'),('FI','Foto CM'),('FJ','Revision'),('FK','Clasif'),('FL','Fec entrega Ficha')]),
    ('FM', 'FICHAS TEC PROD', [('FM','Especificadora'),('FN','Inicio espec'),('FO','Rev materiales'),('FP','Entr ficha bord'),('FQ','Fec ficha bord'),('FR','Fec final espec')]),
    ('FS', 'STATUS CM', [('FS','Prioridad'),('FT','Fec meta'),('FU','Drops'),('FV','Status CM')]),
    ('FX', 'ENTREGABLES FINALES', [('FX','Lib diseno a ing'),('FY','Lib ing a prod')]),
    ('FZ', 'CONTRAMUESTRAS', [('FZ','Foto CM'),('GA','Und cortadas'),('GB','Nombre CM'),('GC','Talla'),('GD','Color'),('GE','Reprogramacion'),('GF','Codigo OT'),('GG','Nota Fabricac'),('GH','Gestion Nota'),('GI','Fec traslado SAP'),('GJ','Fec despacho ZF')]),
    ('GK', 'FEEDBACK PROD', [('GK','Lib ficha fisica'),('GL','Hallazgos repro'),('GM','Clasif escalado'),('GN','Tipo escalado'),('GO','Obs escalado'),('GP','Clasif ficha'),('GQ','Variables ficha'),('GR','Obs ficha'),('GS','Clasif bordado'),('GT','Variables bord'),('GU','Obs bordado'),('GV','Clasif consumos'),('GW','Variables cons'),('GX','Obs consumos'),('GY','Clasif compras'),('GZ','Variables comp'),('HA','Area encargada'),('HB','Obs compras'),('HC','Correcc Audaces'),('HD','Situaciones corte')]),
]

HEADERS_B = [
    ('A', 'INFORMACION BASICA', [('A','REF'),('B','FOTO'),('C','REFERENCIA'),('D','STATUS')]),
    ('E', 'CATALOGACION', [('E','MOD ARTE'),('F','UBI TRAZO'),('G','VAR COLOR'),('H','(vacia)'),('I','LARGO')]),
    ('J', 'MATERIAL POR FILA', [('J','USO EN PRENDA'),('K','CODIGO TELA'),('L','DESCR TELA'),('M','ANCHO'),('N','TELA FOTO'),('O','CONSUMO BASE')]),
    ('P', 'EQUIPO CREATIVO', [('P','DIS CREATIVO'),('Q','CAMBIO MOLD'),('R','CONSUMO 1'),('S','CONSUMO 2'),('T','CONSUMO 3'),('U','OBS CREATIVO')]),
    ('V', 'EQUIPO TECNICO', [('V','DIS TECNICO'),('W','TALLA SOL'),('X','UND SOL'),('Y','CONS 1 SOL'),('Z','CONS 2 SOL'),('AA','CONS 3 SOL')]),
    ('AB', 'TEC MOD ARTE', [('AB','TALLA MOD'),('AC','UND MOD'),('AD','CONS 1 MOD'),('AE','CONS 2 MOD'),('AF','CONS 3 MOD')]),
    ('AG', 'TEC UBI TRAZO', [('AG','TALLA UBI'),('AH','UND UBI'),('AI','CONS 1 UBI'),('AJ','CONS 2 UBI')]),
    ('AK', 'OBS TECNICO', [('AK','OBS TECNICO'),('AL','A RIESGO')]),
    ('AM', 'TRAZO SOLIDO', [('AM','TALLA'),('AN','UND'),('AO','CONS 1'),('AP','CONS 2'),('AQ','CONS 3'),('AR','CONS 4')]),
    ('AS', 'TRAZO MOD ARTE', [('AS','TALLA'),('AT','UND'),('AU','CONS 1'),('AV','CONS 2'),('AW','CONS 3')]),
    ('AX', 'TRAZO UBI', [('AX','TALLA'),('AY','UND'),('AZ','CONS 1'),('BA','CONS 2')]),
    ('BB', 'COSTEO SOLIDO', [('BB','TALLA'),('BC','UND'),('BD','CONSUMO')]),
    ('BE', 'COSTEO MOD', [('BE','TALLA'),('BF','UND'),('BG','CONSUMO')]),
    ('BH', 'COSTEO UBI', [('BH','TALLA'),('BI','UND'),('BJ','CONSUMO')]),
    ('BK', 'OBS COSTEO', [('BK','OBS COSTEO')]),
    ('BL', 'EQUIPO', [('BL','TEC'),('BM','TRAZADOR'),('BN','TERMINADO TALLER')]),
]

CATALOGS = [
    ('A','ALL_OVER'),('B','BASE_TEXTIL'),('C','CAMBIO_MOLDERIA'),('D','CATALOGACION'),('E','CLOSURE'),
    ('F','CLASIFICACION_HALLAZGO'),('G','DISENADOR_CREATIVO'),('H','DISENADOR_TECNICO'),
    ('I','MODISTA'),('J','PRIORIDADES'),('K','SI_NO'),('L','STATUS_PROCESO_EXT'),
    ('M','STATUS'),('N','STATUS_TALLER'),('O','SUBLINEA'),('P','TALLAJE'),('Q','TIPO_BORDADO'),
    ('R','TIPO_CORTE'),('S','TIPO_PROCESO_EXT'),('T','TIPO_RECHAZO'),('U','TIPO_REF'),
    ('V','TIPO_TEJIDO'),('W','UBI_TRAZO'),('X','USO_PRENDA'),('Y','VEREDICTO_DIR'),
    ('Z','WOVEN_KNITTED'),('AA','DIFICULTAD'),('AB','DROPS'),('AC','INCLUDES'),
    ('AD','LARGO'),('AE','LINEA'),('AF','LINNED'),('AG','CONFECCION_STATUS'),('AH','SENTIDO_SESGO'),
]

def write_param(ws, extra_cols=None):
    ws.sheet_properties.tabColor = '27AE60'
    for col_l, col_name in CATALOGS:
        cidx = ci(col_l)
        c = ws.cell(1, cidx, col_name)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = thin
        c2 = ws.cell(2, cidx, f'=IMPORTRANGE("{MID}","CATALOGOS!{col_l}:{col_l}")')
        c2.fill = imp_fill; c2.font = data_font; c2.alignment = left_al; c2.border = thin
        ws.column_dimensions[col_l].width = 22
    if extra_cols:
        for col_l, col_name, formula in extra_cols:
            cidx = ci(col_l)
            c = ws.cell(1, cidx, col_name)
            c.fill = accent_fill; c.font = Font(bold=True, size=9, name='Calibri'); c.alignment = center; c.border = thin
            c2 = ws.cell(2, cidx, formula)
            c2.fill = imp_fill; c2.font = data_font; c2.alignment = left_al; c2.border = thin
            ws.column_dimensions[col_l].width = 28
    ws.freeze_panes = 'A3'

def write_instrucciones(ws, archivo_nombre):
    ws.sheet_properties.tabColor = '2980B9'
    ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 80
    r = 1
    ws.cell(r,2,'INSTRUCCIONES - PARAMETROS NUEVA').font = Font(size=14, bold=True, color='2C3E50', name='Calibri')
    r += 2
    ws.cell(r,2,f'Archivo: {archivo_nombre}').font = Font(size=10, color='7F8C8D', name='Calibri')
    r += 2
    for paso, desc in [
        ('PASO 1', 'Sube este .xlsx a Google Drive (junto al archivo original)'),
        ('PASO 2', 'Abrelo en Google Sheets (se convierte automaticamente)'),
        ('PASO 3', 'Renombra PARAMETROS actual a PARAMETROS_OLD en el archivo original'),
        ('PASO 4', 'Copia la pestana PARAMETROS_NUEVA al archivo original'),
        ('PASO 5', 'Reemplaza PEGA_ID_MAESTRO_AQUI por el ID real del Maestro'),
        ('PASO 6', 'Autoriza cada #REF! haciendo click > Conceder acceso'),
        ('PASO 7', 'Actualiza dropdowns en MATRIZ para apuntar a PARAMETROS_NUEVA'),
        ('PASO 8', 'Verifica y NO borres PARAMETROS_OLD hasta confirmar'),
    ]:
        ws.cell(r,2,paso).font = Font(bold=True, size=10, color='E74C3C', name='Calibri')
        ws.cell(r+1,2,desc).font = Font(size=10, name='Calibri'); r += 3

def write_matriz(ws, headers):
    for sec_start, sec_name, cols in headers:
        s = ci(sec_start); e = s + len(cols) - 1
        ws.merge_cells(start_row=3, start_column=s, end_row=3, end_column=e)
        c = ws.cell(3, s, sec_name)
        c.fill = sec_fill; c.font = sec_font; c.alignment = center; c.border = thin
        for cl, cn in cols:
            cx = ci(cl)
            c2 = ws.cell(4, cx, cn)
            c2.fill = hdr_fill; c2.font = hdr_font; c2.alignment = center; c2.border = thin
            ws.column_dimensions[cl].width = max(len(cn)+2, 10)
    ws.cell(1, 1, 'MATRIZ - ESTRUCTURA VACIA - Llenar con datos').font = Font(size=12, bold=True, color='7F8C8D', name='Calibri')
    ws.freeze_panes = 'A5'

# ── GENERAR ──
print("=" * 50)
print("OPCION 3: Estructura vacia + PARAMETROS completo")
print("=" * 50)

files = [
    ("FO_APPAREL_2026_ESTRUCTURA.xlsx", "FO APPAREL 2026", HEADERS_FO, [
        ('AI','DISENADORES',f'=IMPORTRANGE("{MID}","DISENADORES!A:C")'),
        ('AL','MODISTAS',f'=IMPORTRANGE("{MID}","MODISTAS!A:B")'),
        ('AO','LINEAS',f'=IMPORTRANGE("{MID}","LINEAS!A:B")'),
        ('AR','COLECCIONES',f'=IMPORTRANGE("{MID}","COLECCIONES!A:F")'),
    ]),
    ("COLECCION_WS27_ESTRUCTURA.xlsx", "COLECCION WS27", HEADERS_B, []),
    ("CONTRAM_WS27_ESTRUCTURA.xlsx", "CONTRAM WS27", HEADERS_FO, [
        ('AI','DISENADORES',f'=IMPORTRANGE("{MID}","DISENADORES!A:C")'),
        ('AL','MODISTAS',f'=IMPORTRANGE("{MID}","MODISTAS!A:B")'),
        ('AO','LINEAS',f'=IMPORTRANGE("{MID}","LINEAS!A:B")'),
    ]),
]

for fname, label, headers, extra in files:
    wb = openpyxl.Workbook()
    ws0 = wb.active; ws0.title = "INSTRUCCIONES"; write_instrucciones(ws0, label)
    ws1 = wb.create_sheet("MATRIZ"); write_matriz(ws1, headers)
    ws2 = wb.create_sheet("PARAMETROS_NUEVA"); write_param(ws2, extra)
    path = OP3 / fname
    wb.save(str(path))
    print(f"  {fname}")

print("OPCION 3 COMPLETADA")
