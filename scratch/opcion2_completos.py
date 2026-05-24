"""OPCION 2: Archivos completos desde CSV/HTML con PARAMETROS_NUEVA"""
import sys, io, os, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE = Path(r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0")
DOCS = BASE / "Documentos"
OP2 = BASE / "dist" / "opcion2-archivos-completos"
MID = "PEGA_ID_MAESTRO_AQUI"

thin = Border(left=Side(style='thin', color='808080'), right=Side(style='thin', color='808080'),
              top=Side(style='thin', color='808080'), bottom=Side(style='thin', color='808080'))
hdr_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
hdr_font = Font(color='FFFFFF', bold=True, size=10, name='Calibri')
imp_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
accent_fill = PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_al = Alignment(horizontal='left', vertical='center', wrap_text=True)
data_font = Font(name='Calibri', size=9)

def ci(l): return openpyxl.utils.column_index_from_string(l)

CATALOGS = [
    ('A','ALL_OVER'),('B','BASE_TEXTIL'),('C','CAMBIO_MOLDERIA'),('D','CATALOGACION'),('E','CLOSURE'),
    ('F','CLASIFICACION_HALLAZGO'),('G','DISENADOR_CREATIVO'),('H','DISENADOR_TECNICO'),
    ('I','MODISTA'),('J','PRIORIDADES'),('K','SI_NO'),('L','STATUS_PROCESO_EXT'),
    ('M','STATUS'),('N','STATUS_TALLER'),('O','SUBLINEA'),('P','TALLAJE'),('Q','TIPO_BORDADO'),
    ('R','TIPO_CORTE'),('S','TIPO_PROCESO_EXT'),('T','TIPO_RECHAZO'),('U','TIPO_REF'),
    ('V','TIPO_TEJIDO'),('W','UBI_TRAZO'),('X','USO_PRENDA'),('Y','VEREDICTO_DIR'),
    ('Z','WOVEN_KNITTED'),('AA','DIFICULTAD'),('AB','DROPS'),('AC','INCLUDES'),
    ('AD','LARGO'),('AE','LINEA'),('AF','LINNED'),('AG','CONFECCION_STATUS'),('AH','SENTIDO_SESGO'),
]

def html_to_sheet(wb, html_path, sheet_name):
    """Parse HTML table and write to a sheet in the workbook."""
    # Try lxml first, fall back to default parser
    try:
        df = pd.read_html(str(html_path), flavor='lxml')[0]
    except Exception:
        try:
            df = pd.read_html(str(html_path))[0]
        except Exception:
            try:
                df = pd.read_html(str(html_path), flavor='bs4')[0]
            except Exception:
                print(f"    SKIP {sheet_name}: no tables parseable")
                return
    ws_name = sheet_name[:31]  # Excel sheet name limit
    ws = wb.create_sheet(title=ws_name)
    # Write values (skip 'Unnamed: 0' column which is just row index)
    start_col = 1
    if df.columns[0] == 'Unnamed: 0':
        start_col = 2
    for ri_idx, (ri, row) in enumerate(df.iterrows(), 1):
        for ci_idx, col_name in enumerate(df.columns):
            if col_name == 'Unnamed: 0':
                continue
            val = row[col_name]
            if pd.notna(val):
                cell = ws.cell(row=ri_idx, column=ci_idx + start_col, value=str(val).strip()[:32767])
                cell.border = thin
                cell.font = data_font
                cell.alignment = center
    # Set column widths
    for ci in range(1, len(df.columns)):
        ws.column_dimensions[get_column_letter(ci)].width = 15
    ws.freeze_panes = 'A2'
    print(f'    {sheet_name}: {df.shape[0]} rows x {df.shape[1]} cols')

def write_parametros_nueva(wb):
    ws = wb.create_sheet("PARAMETROS_NUEVA")
    ws.sheet_properties.tabColor = '27AE60'
    for col_l, col_name in CATALOGS:
        cidx = ci(col_l)
        c = ws.cell(1, cidx, col_name)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = thin
        c2 = ws.cell(2, cidx, f'=IMPORTRANGE("{MID}","CATALOGOS!{col_l}:{col_l}")')
        c2.fill = imp_fill; c2.font = data_font; c2.alignment = left_al; c2.border = thin
        ws.column_dimensions[col_l].width = 22
    # Extra: DISENADORES, MODISTAS, LINEAS
    for col_l, col_name, formula in [
        ('AI','DISENADORES',f'=IMPORTRANGE("{MID}","DISENADORES!A:C")'),
        ('AL','MODISTAS',f'=IMPORTRANGE("{MID}","MODISTAS!A:B")'),
        ('AO','LINEAS',f'=IMPORTRANGE("{MID}","LINEAS!A:B")'),
    ]:
        cidx = ci(col_l)
        c = ws.cell(1, cidx, col_name)
        c.fill = accent_fill; c.font = Font(bold=True, size=9, name='Calibri'); c.alignment = center; c.border = thin
        c2 = ws.cell(2, cidx, formula)
        c2.fill = imp_fill; c2.font = data_font; c2.alignment = left_al; c2.border = thin
        ws.column_dimensions[col_l].width = 28
    ws.freeze_panes = 'A3'

def write_instrucciones(ws):
    ws.sheet_properties.tabColor = '2980B9'
    ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 80
    r=1; ws.cell(r,2,'INSTRUCCIONES').font=Font(size=14,bold=True,color='2C3E50',name='Calibri')
    r+=2; ws.cell(r,2,'Archivo completo con datos + PARAMETROS_NUEVA').font=Font(size=10,color='7F8C8D',name='Calibri')
    r+=2
    msgs = [
        'PASO 1: Sube este .xlsx a Google Drive (se convierte automaticamente a Sheets)',
        'PASO 2: Renombra PARAMETROS actual a PARAMETROS_OLD en el archivo existente (si aplica)',
        'PASO 3: Copia la pestana PARAMETROS_NUEVA de este archivo al archivo existente',
        'PASO 4: Reemplaza PEGA_ID_MAESTRO_AQUI por el ID real del Maestro',
        'PASO 5: Autoriza cada #REF! (click > Conceder acceso)',
        'PASO 6: ADVERTENCIA: Las formulas, formato condicional y fotos NO se conservan desde los archivos originales',
        'PASO 7: Este archivo es una reconstruccion desde datos exportados. Algunas celdas pueden estar vacias o desordenadas',
    ]
    for msg in msgs:
        ws.cell(r,2,msg).font=Font(size=10,name='Calibri'); r+=2

# ═══════════════════════════════════════════════════════════
# ARCHIVO A: FO APPAREL 2026
# ═══════════════════════════════════════════════════════════
print("=" * 50)
print("ARCHIVO A: FO APPAREL 2026")
print("=" * 50)
wbA = openpyxl.Workbook()
ws0 = wbA.active; ws0.title = "INSTRUCCIONES"; write_instrucciones(ws0)

dirA = DOCS / "Copia de FO... APPAREL 2026"
# Map sheet name to exact HTML filename
html_map_A = {
    "WS27": "WS27.html",
    "FW26": "FW26.html",
    "INGENIERIA WS27": "INGENIERIA WS27.html",
    "INGENIERIA FW26": "INGENIERIA FW26.html",
    "LABS WS27": "LABS WS27.html",
    "LABS FW26": "FW26 LABS.html",
    "BORDADO WS27": "BORDADO WS27.html",
    "FW26 BORDADO": "FW26 BORDADO .html",
    "LABS BORDADO WS27": "LABS BORDADO WS27.html",
    "FW26 LABS BORDADO": "FW26 LABS BORDADO.html",
}
for tab, fname in html_map_A.items():
    try:
        html_path = dirA / fname
        html_to_sheet(wbA, html_path, tab)
    except Exception as e:
        print(f"    ERROR {tab}: {str(e)[:100]}")

write_parametros_nueva(wbA)
pathA = OP2 / "FO_APPAREL_2026.xlsx"
wbA.save(str(pathA))
print(f"  Guardado: {pathA.name} ({len(wbA.sheetnames)} pestañas)")

# ═══════════════════════════════════════════════════════════
# ARCHIVO B: COLECCION WS27
# ═══════════════════════════════════════════════════════════
print("\n" + "=" * 50)
print("ARCHIVO B: COLECCION WS27")
print("=" * 50)
wbB = openpyxl.Workbook()
ws0 = wbB.active; ws0.title = "INSTRUCCIONES"; write_instrucciones(ws0)

dirB = DOCS / "Copia de COLECCION WS27"
html_tabs_B = [
    "1.VALIDACION DE TELAS", "CHECK LIST", "COSTEO INSUMOS",
    "INSUMOS", "LINESHEET FW26",
]
for tab in html_tabs_B:
    try:
        html_path = dirB / f"{tab}.html"
        html_to_sheet(wbB, html_path, tab)
    except Exception as e:
        print(f"    ERROR {tab}: {str(e)[:100]}")

write_parametros_nueva(wbB)
pathB = OP2 / "COLECCION_WS27.xlsx"
wbB.save(str(pathB))
print(f"  Guardado: {pathB.name} ({len(wbB.sheetnames)} pestañas)")

# ═══════════════════════════════════════════════════════════
# ARCHIVO C: CONTRAM WS27
# ═══════════════════════════════════════════════════════════
print("\n" + "=" * 50)
print("ARCHIVO C: CONTRAM WS27")
print("=" * 50)
wbC = openpyxl.Workbook()
ws0 = wbC.active; ws0.title = "INSTRUCCIONES"; write_instrucciones(ws0)

def normalize_name(s):
    return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))

dirC = DOCS / "Copia de CONTRAM WS27"
all_files = os.listdir(str(dirC))
html_files = [f for f in all_files if f.endswith('.html') and f not in ('PARAMETROS.html',)]
print(f"  HTML files in C: {html_files}")
for target in ["MATRIZ", "DESCRIPCION"]:
    matches = [f for f in html_files if normalize_name(f).upper().startswith(target.upper())]
    if matches:
        html_to_sheet(wbC, dirC / matches[0], target)
    else:
        print(f"    ERROR {target}: not found among {html_files}")

write_parametros_nueva(wbC)
pathC = OP2 / "CONTRAM_WS27.xlsx"
wbC.save(str(pathC))
print(f"  Guardado: {pathC.name} ({len(wbC.sheetnames)} pestañas)")

print("\n" + "=" * 50)
print("OPCION 2 COMPLETADA")
print("=" * 50)
