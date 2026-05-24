"""
OPCION 3: Estructura vacia + PARAMETROS completo.
Cada .xlsx tiene la estructura de headers de MATRIZ (sin datos) + PARAMETROS_NUEVA.
Los headers siguen la estructura definida en descripcion proyecto.md.
Incluye Data Validations en las columnas de catalogo.
"""

import sys, io
from pathlib import Path
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0")
DIST = BASE / "dist"
OP3 = DIST / "opcion3-estructura-vacia"

thin = Border(
    left=Side(style='thin', color='808080'),
    right=Side(style='thin', color='808080'),
    top=Side(style='thin', color='808080'),
    bottom=Side(style='thin', color='808080')
)
hdr_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
sec_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
sec_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
alt_fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)
MID = 'PEGA_ID_MAESTRO_AQUI'

def col_idx(letter):
    return openpyxl.utils.column_index_from_string(letter)

HEADERS_FO = [
    ("A", "INFORMACION BASICA", [
        ("A", "Ref"), ("B", "Imagen"), ("C", "Codigo MD"), ("D", "Codigo PT"),
        ("E", "Nombre referencia"), ("F", "Color"), ("G", "Cod color"),
    ]),
    ("H", "REFERENTES", [
        ("H", "Foto basado en"), ("I", "PT"), ("J", "Basado en"),
    ]),
    ("K", "STATUS DE COLECCION", [
        ("K", "Status"), ("L", "Disenador"), ("M", "Status taller"),
        ("N", "Modista"), ("O", "Fotos internas"),
    ]),
    ("P", "CARACTERISTICAS PRENDA", [
        ("P", "Linea"), ("Q", "Sublinea"), ("R", "Tipo ref"),
        ("S", "Tallaje"), ("T", "Largo"), ("U", "Closure"), ("V", "Largo 2"),
    ]),
    ("W", "INCLUDES", [
        ("W", "Includes"), ("X", "Includes paq completo"),
    ]),
    ("Y", "TELA DE LUCIR MUESTRA", [
        ("Y", "Codigo tela lucir"), ("Z", "Foto tela"), ("AA", "Descripcion tela"),
        ("AB", "Ancho tela"), ("AC", "Base textil"), ("AD", "Ubi en trazo"),
        ("AE", "Mod arte"), ("AF", "All over"),
    ]),
    ("AG", "VARIACION COLOR", [
        ("AG", "Variacion color"), ("AH", "Ref variacion"),
    ]),
    ("AI", "EMPAQUES", [("AI", "Tipo empaque")]),
    ("AJ", "BORDADO EN PRENDA", [
        ("AJ", "Bordado aplica"), ("AK", "Bordado descripcion"),
    ]),
    ("AL", "SEMIELABORADOS", [
        ("AL", "Semielaborado aplica"), ("AM", "Semielaborado descripcion"),
    ]),
    ("AN", "PROCESO EXTERNO", [
        ("AN", "Proveedor"), ("AO", "Proceso externo"), ("AP", "Costo"),
    ]),
    ("AQ", "ESTADOS Y ENTREGAS", [
        ("AQ", "Bordado status"), ("AR", "Terminado taller costeo"),
        ("AS", "Entregable creativo"), ("AT", "Entregable tecnico"),
        ("AU", "Entregable trazador"), ("AV", "Envio MOD arte"),
    ]),
    ("AW", "TELAS", [
        ("AW", "Comentarios telas"), ("AX", "Bordado status telas"),
        ("AY", "Terminado taller telas"),
    ]),
    ("AZ", "INSUMOS", [
        ("AZ", "Proceso externo insumos"), ("BA", "Costo insumos"),
    ]),
    ("BB", "BORDADO", [
        ("BB", "Bordado status 2"), ("BC", "Terminado taller bordado"),
    ]),
    ("BD", "ENTREGAS", [("BD", "Entrega parcial")]),
    ("BE", "TIME COLLECTION", [
        ("BE", "Especificacion confeccion 1"), ("BF", "Escalado molderia 1"),
        ("BG", "Tiras continuas"), ("BH", "Dificultad prenda"),
        ("BI", "Dificultad bordado"), ("BJ", "Prioridades first buy"),
        ("BK", "Prioridades bordado"), ("BL", "Bordado tipo"),
        ("BM", "Prioridades textil stock"), ("BN", "Comentarios ingenieria"),
        ("BO", "Comentarios trazo"), ("BP", "Comentarios costeo"),
        ("BQ", "Sugerencia MOD/UBC"), ("BR", "Requiere muestra"),
        ("BS", "Grupo/estilo"),
    ]),
    ("BT", "VALIDACION MP", [
        ("BT", "Fecha"), ("BU", "Area afectacion"), ("BV", "Clasificacion hallazgo"),
        ("BW", "MP"), ("BX", "Clasificacion MP"), ("BY", "Tipo ejecucion"),
    ]),
    ("BZ", "COMPOSICION MUESTRA", [
        ("BZ", "Info SAP"), ("CA", "Description USA-UK"),
        ("CB", "Fiber Composition"), ("CC", "Woven"), ("CD", "Inside"),
        ("CE", "Include"), ("CF", "Observaciones comp"),
    ]),
    ("CG", "COMPOSICION PRODUCCION", [
        ("CG", "Info SAP prod"), ("CH", "Description USA-UK prod"),
        ("CI", "Fiber Composition prod"), ("CJ", "Woven prod"),
        ("CK", "Inside prod"), ("CL", "Include prod"),
        ("CM", "Observaciones comp prod"),
    ]),
    ("CN", "COMEX", [("CN", "Comex")]),
    ("CO", "CUIDADOS PRENDA", [
        ("CO", "Proceso"), ("CP", "Lavado"), ("CQ", "Logo lavado"),
        ("CR", "Desmanche"), ("CS", "Logo desmanche"),
        ("CT", "Secado"), ("CU", "Logo secado"),
        ("CV", "Planchado"), ("CW", "Logo plancha"),
        ("CX", "Cuidados includes"),
    ]),
    ("CY", "UNIDADES PRODUCCION", [
        ("CY", "Talla 0"), ("CZ", "Talla 2"), ("DA", "Talla 4"),
        ("DB", "Talla 6"), ("DC", "Talla 8"), ("DD", "Talla 10"),
        ("DE", "Talla 12"), ("DF", "XS"), ("DG", "S"), ("DH", "M"),
        ("DI", "L"), ("DJ", "XL"), ("DK", "TOTAL"),
    ]),
    ("DL", "MAQUILA", [
        ("DL", "Tipo tejido"), ("DM", "Complejidad corte"),
        ("DN", "Envio corte maquilas"), ("DO", "Complejidad confeccion"),
        ("DP", "Envio confeccion maquilas"),
    ]),
    ("DQ", "MONTAJE MANIQUI", [
        ("DQ", "Tipo montaje maniqui"), ("DR", "(Vacia)"),
        ("DS", "Proyecto montaje"),
    ]),
    ("DT", "DIS. CONTRAMUESTRA", [
        ("DT", "Disenador tecnico CM"), ("DU", "Disenador creativo CM"),
    ]),
    ("DV", "MOLDERIA", [
        ("DV", "Fecha inicio molderia"), ("DW", "Fecha fin molderia"),
        ("DX", "Comentarios molderia"),
    ]),
    ("DY", "PROCESOS EXTERNOS", [
        ("DY", "Tipo proceso externo"), ("DZ", "Fecha recibido pieza"),
        ("EA", "Fecha entrega pieza"), ("EB", "Status proceso externo"),
    ]),
    ("EC", "CORTE CONTRAMUESTRAS", [
        ("EC", "Fecha corte 1"), ("ED", "Tipo corte 1"),
        ("EE", "Fecha corte 2"), ("EF", "Tipo corte 2"),
        ("EG", "Fecha corte 3"), ("EH", "Tipo corte 3"),
        ("EI", "Fecha corte 4"), ("EJ", "Tipo corte 4"),
        ("EK", "Observaciones corte"), ("EL", "Total cortes piezas"),
        ("EM", "Total cortes muestras"),
    ]),
    ("EN", "CONFECCION", [
        ("EN", "Modista confeccion"), ("EO", "Fecha inicio confeccion"),
        ("EP", "Fecha entrega confeccion"), ("EQ", "Status confeccion"),
        ("ER", "Observaciones modista"), ("ES", "Observaciones tecnico 1"),
        ("ET", "Observaciones tecnico 2"), ("EU", "Tiempo confeccion min"),
        ("EV", "Estado prenda planta"), ("EW", "Tipo rechazo"),
        ("EX", "Feedback produccion"),
    ]),
    ("EY", "MEDICION", [
        ("EY", "Medicion fase 1"), ("EZ", "Comentarios medicion 1"),
        ("FA", "Medicion fase 2"), ("FB", "Comentarios medicion 2"),
        ("FC", "Medicion fase 3"), ("FD", "Comentarios medicion 3"),
        ("FE", "Medicion fase 4"), ("FF", "Comentarios medicion 4"),
        ("FG", "Medicion fase 5"), ("FH", "Comentarios medicion 5"),
        ("FI", "Foto contramuestra"), ("FJ", "Revision"),
        ("FK", "Clasificacion"), ("FL", "Fecha entrega a Ficha"),
    ]),
    ("FM", "FICHAS TECNICAS PROD", [
        ("FM", "Especificadora"), ("FN", "Inicio especificacion"),
        ("FO", "Revision materiales"), ("FP", "Entrega ficha bordado"),
        ("FQ", "Fecha entrega ficha bordado"), ("FR", "Fecha final especificacion"),
    ]),
    ("FS", "STATUS CONTRAMUESTRA", [
        ("FS", "Prioridad contramuestras"), ("FT", "Fecha meta entrega"),
        ("FU", "Drops"), ("FV", "Status contramuestra"),
    ]),
    ("FX", "ENTREGABLES", [
        ("FX", "Fecha liberacion diseno"), ("FY", "Fecha liberacion ingenieria"),
    ]),
    ("FZ", "CONTRAMUESTRAS", [
        ("FZ", "Foto contramuestra"), ("GA", "Unidades cortadas"),
        ("GB", "Nombre contramuestra"), ("GC", "Talla contramuestra"),
        ("GD", "Color contramuestra"), ("GE", "Reprogramacion CM"),
        ("GF", "Codigo OT"), ("GG", "Nota de Fabricacion"),
        ("GH", "Gestion de Nota"), ("GI", "Fecha traslado SAP"),
        ("GJ", "Fecha despacho ZF"),
    ]),
    ("GK", "FEEDBACK PRODUCCION", [
        ("GK", "Liberacion ficha fisica"), ("GL", "Hallazgos reprogramacion"),
        ("GM", "Clasificacion situac prod"), ("GN", "Tipo situacion produccion"),
        ("GO", "Observaciones escalado"),
        ("GP", "Clasificacion situac ficha"), ("GQ", "Variables ficha"),
        ("GR", "Observaciones ficha"),
        ("GS", "Clasificacion situac bordado"), ("GT", "Variables bordado"),
        ("GU", "Observaciones bordado"),
        ("GV", "Clasificacion situac consumos"), ("GW", "Variables consumos"),
        ("GX", "Observaciones consumos"),
        ("GY", "Clasificacion situac compras"), ("GZ", "Variables compras"),
        ("HA", "Area encargada"), ("HB", "Observaciones compras"),
        ("HC", "Correccion realizada Audaces"), ("HD", "Situaciones en corte"),
    ]),
]

HEADERS_B = [
    ("A", "INFORMACION BASICA", [
        ("A", "REF"), ("B", "FOTO"), ("C", "REFERENCIA"), ("D", "STATUS"),
    ]),
    ("E", "CATALOGACION", [
        ("E", "MOD.ARTE"), ("F", "UBI. TRAZO"), ("G", "VARIACION COLOR"),
        ("H", "(vacia)"), ("I", "LARGO"),
    ]),
    ("J", "MATERIAL", [
        ("J", "USO EN PRENDA"), ("K", "CODIGO TELA"), ("L", "DESCRIPCION TELA"),
        ("M", "ANCHO"), ("N", "TELA FOTO"), ("O", "CONSUMO BASE"),
    ]),
    ("P", "EQUIPO CREATIVO", [
        ("P", "DISENADOR CREATIVO"), ("Q", "CAMBIO MOLDERIA"),
        ("R", "CONSUMO 1"), ("S", "CONSUMO 2"), ("T", "CONSUMO 3"),
        ("U", "OBSERVACIONES CREATIVO"),
    ]),
    ("V", "EQUIPO TECNICO - SOLIDO", [
        ("V", "DISENADOR TECNICO"), ("W", "TALLA"), ("X", "UND"),
        ("Y", "CONSUMO 1 SOLIDO"), ("Z", "CONSUMO 2 SOLIDO"), ("AA", "CONSUMO 3 SOLIDO"),
    ]),
    ("AB", "EQUIPO TECNICO - MOD ARTE", [
        ("AB", "TALLA MOD"), ("AC", "UND MOD"), ("AD", "CONSUMO 1 MOD"),
        ("AE", "CONSUMO 2 MOD"), ("AF", "CONSUMO 3 MOD"),
    ]),
    ("AG", "EQUIPO TECNICO - UBI TRAZO", [
        ("AG", "TALLA UBI"), ("AH", "UND UBI"), ("AI", "CONSUMO 1 UBI"),
        ("AJ", "CONSUMO 2 UBI"),
    ]),
    ("AK", "OBS TECNICO", [
        ("AK", "OBSERVACIONES TECNICO"), ("AL", "A RIESGO / VALIDADO"),
    ]),
    ("AM", "TRAZO SOLIDO", [
        ("AM", "TALLA"), ("AN", "UND"), ("AO", "CONSUMO 1"),
        ("AP", "CONSUMO 2"), ("AQ", "CONSUMO 3"), ("AR", "CONSUMO 4"),
    ]),
    ("AS", "TRAZO MOD ARTE", [
        ("AS", "TALLA"), ("AT", "UND"), ("AU", "CONSUMO 1 MOD"),
        ("AV", "CONSUMO 2 MOD"), ("AW", "CONSUMO 3 MOD"),
    ]),
    ("AX", "TRAZO UBI", [
        ("AX", "TALLA"), ("AY", "UND"), ("AZ", "CONSUMO 1 UBI"),
        ("BA", "CONSUMO 2 UBI"),
    ]),
    ("BB", "COSTEO SOLIDO", [
        ("BB", "TALLA"), ("BC", "UND"), ("BD", "CONSUMO SOLIDO"),
    ]),
    ("BE", "COSTEO MOD ARTE", [
        ("BE", "TALLA"), ("BF", "UND"), ("BG", "CONSUMO MOD"),
    ]),
    ("BH", "COSTEO UBI", [
        ("BH", "TALLA"), ("BI", "UND"), ("BJ", "CONSUMO UBI"),
    ]),
    ("BK", "COSTEO OBS", [("BK", "OBS COSTEO")]),
    ("BL", "EQUIPO", [
        ("BL", "TEC"), ("BM", "TRAZADOR"), ("BN", "TERMINADO TALLER"),
    ]),
]


def write_matriz_headers(ws, header_sections, data_start_row=5):
    for sec_start, sec_name, cols in header_sections:
        start_ci = col_idx(sec_start)
        end_ci = start_ci + len(cols) - 1
        end_letter = get_column_letter(end_ci)
        ws.merge_cells(start_row=3, start_column=start_ci, end_row=3, end_column=end_ci)
        cell = ws.cell(3, start_ci, sec_name)
        cell.fill = sec_fill
        cell.font = sec_font
        cell.alignment = center
        cell.border = thin

        for col_letter, col_name in cols:
            ci = col_idx(col_letter)
            cell = ws.cell(4, ci, col_name)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = center
            cell.border = thin
            ws.column_dimensions[col_letter].width = max(len(col_name) + 2, 12)

    ws.freeze_panes = f"A{data_start_row}"

    # Row 1: collection identifier
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_ci)
    ws.cell(1, 1, "ESTRUCTURA VACIA - Llenar con datos via IMPORTRANGE o manualmente").font = Font(
        size=12, bold=True, color="7F8C8D", name="Calibri")

    # Row 2: placeholder for IMPORTRANGE data source
    ws.cell(2, 1, f'Formula sugerida row {data_start_row}: =IMPORTRANGE("{MID}","MATRIZ!A:HD")').font = Font(
        size=9, italic=True, color="2980B9", name="Calibri")


def write_parametros_nueva_op3(ws, tipo):
    ws.sheet_properties.tabColor = "27AE60"

    if tipo == 'A':
        from scratch.opcion1_parametros import write_parametros_a
        write_parametros_a(ws)
    elif tipo == 'B':
        from scratch.opcion1_parametros import write_parametros_b
        write_parametros_b(ws)
    else:
        from scratch.opcion1_parametros import write_parametros_c
        write_parametros_c(ws)


print("=" * 50)
print("OPCION 3: Estructura vacia + PARAMETROS")
print("=" * 50)

for nombre, label, headers, tipo_param in [
    ("FO_APPAREL_2026_ESTRUCTURA.xlsx", "FO APPAREL 2026", HEADERS_FO, "A"),
    ("COLECCION_WS27_ESTRUCTURA.xlsx", "COLECCION WS27", HEADERS_B, "B"),
    ("CONTRAM_WS27_ESTRUCTURA.xlsx", "CONTRAM WS27", HEADERS_FO, "C"),
]:
    wb = openpyxl.Workbook()

    ws0 = wb.active
    ws0.title = "INSTRUCCIONES"
    from scratch.opcion1_parametros import write_instrucciones
    write_instrucciones(ws0, label, MID)
    ws0.cell(20, 2, "NOTA: La pestana MATRIZ esta vacia (solo headers). Los datos se importan via IMPORTRANGE o se llenan manualmente.").font = Font(
        bold=True, size=10, color="E74C3C", name="Calibri")

    ws1 = wb.create_sheet("MATRIZ")
    write_matriz_headers(ws1, headers)

    ws2 = wb.create_sheet("PARAMETROS_NUEVA")
    if tipo_param == 'A':
        from scratch.opcion1_parametros import write_parametros_a as wp
    elif tipo_param == 'B':
        from scratch.opcion1_parametros import write_parametros_b as wp
    else:
        from scratch.opcion1_parametros import write_parametros_a as wp
    wp(ws2)

    path = OP3 / nombre
    wb.save(str(path))
    print(f"  Creado: {path.name}")

print("OPCION 3 COMPLETADA")
