"""Generate VALIDACION_DE_TELAS_v2.2 from v2.1 with updated INSTRUCCIONES"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

BASE = r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0"
INPUT = os.path.join(BASE, "dist", "VALIDACION_DE_TELAS_v2.1.xlsx")
OUTPUT = os.path.join(BASE, "dist", "VALIDACION_DE_TELAS_v2.2.xlsx")

wb = load_workbook(INPUT)
ws2 = wb["INSTRUCCIONES"]

# Clear current instructions
for ri in range(1, 55):
    for ci in range(1, 3):
        ws2.cell(ri, ci).value = None
    ws2.row_dimensions[ri].height = 22

r = 1
ws2.cell(r, 2, "GUIA DE USO - VALIDACION DE TELAS v2.2").font = Font(
    size=16, bold=True, color='1A1A2E', name='Calibri')
ws2.row_dimensions[r].height = 30

items = [
    ("ESTRUCTURA",
     "Cada referencia ocupa un BLOQUE de 19 filas:\n"
     "  Filas 1-13: TELAS (una fila por cada tela o material usado)\n"
     "  Fila 14: SEPARADOR 'INSUMOS'\n"
     "  Filas 15-19: INSUMOS (cintas, elasticos, argollas, etc.)\n\n"
     "Si una referencia necesita MAS telas o insumos, inserte filas DENTRO del bloque."),

    ("FASES Y COLORES",
     "  FASE 1 (Azul, cols A-I): Info basica - REF, STATUS, MOD ARTE, UBI TRAZO\n"
     "  FASE 2 (Amarillo, cols J-O): Materiales - USO PRENDA, COD TELA, DESC\n"
     "  FASE 3 (Naranja, cols P-BB): COSTEO - Creativo + Tecnico + Trazo y Corte\n"
     "  FASE 4 (Rojo, cols BC-BN): CONTRAMUESTRA - Consumos finales + TEC + Trazador"),

    ("PROTECCION POR AREA (IMPORTANTE - LEER)",
     "Para evitar que alguien edite columnas que no le corresponden:\n\n"
     "1. En Google Sheets: Datos > Hojas y rangos protegidos\n"
     "2. Hacer clic en 'Agregar hoja o rango'\n"
     "3. Seleccionar el rango de columnas segun la tabla abajo\n"
     "4. En 'Editar permisos', elegir 'Personalizada'\n"
     "5. Agregar los emails de las personas de esa area\n\n"
     "RANGOS A PROTEGER (filas 5-99 para todos):\n"
     "  Columnas A-I   (1-9)   -> Coordinacion (edita REF, STATUS, LARGO)\n"
     "  Columnas J-O   (10-15) -> Bodega/Creativo (edita TELAS)\n"
     "  Columnas P-U   (16-21) -> Disenador Creativo (edita consumos creativos)\n"
     "  Columnas V-AL  (22-38) -> Disenador Tecnico (edita consumos tecnicos)\n"
     "  Columnas AM-BB (39-54) -> Trazador (edita trazos)\n"
     "  Columnas BC-BN (55-66) -> Contramuestras (edita CM + TEC + Trazador)\n\n"
     "NOTA: Esto NO requiere script. Es nativo de Google Sheets. "
     "Cada usuario solo puede editar las columnas de su area. "
     "Las demas columnas aparecen con candado."),

    ("SYNC AUTOMATICO CON ATELIERDATA",
     "Los datos de esta hoja se sincronizan con el archivo AtelierData:\n\n"
     "1. sync_incremental.gs se ejecuta cada 6 horas automaticamente\n"
     "2. Detecta REF nuevas -> las agrega en REFERENCIAS\n"
     "3. Detecta STATUS cambiados -> los actualiza\n"
     "4. Detecta telas/consumos nuevos -> los inserta en CONSUMOS\n\n"
     "NO necesita hacer nada manual. El sync es AUTOMATICO.\n"
     "Si necesita forzar un sync manual: Extensiones > Apps Script > "
     "ejecutar syncIncremental()"),

    ("CELDAS COMBINADAS",
     "Las columnas REF, FOTO, REFERENCIA, STATUS, MOD ARTE, UBI TRAZO, "
     "DISENADOR CREATIVO, DISENADOR TECNICO y OBSERVACIONES estan COMBINADAS "
     "verticalmente en las 19 filas del bloque. Su valor se comparte en todo el bloque."),

    ("COMPATIBILIDAD",
     "Este formato es 100% compatible con migracion.gs y sync_incremental.gs.\n"
     "Posiciones de columna identicas: C.REF:0, C.FOTO:1, C.STATUS:3, etc."),

    ("NO HACER",
     "NO reordene columnas.\n"
     "NO elimine filas del encabezado (1-4).\n"
     "NO inserte columnas entre A y BN.\n"
     "NO quite las protecciones de rango.\n"
     "Para mas telas/insumos en una referencia, inserte filas DENTRO del bloque."),
]

for title, content in items:
    r += 2
    c = ws2.cell(r, 2, title)
    c.font = Font(name='Calibri', size=11, bold=True, color='E74C3C')
    r += 1
    c2 = ws2.cell(r, 2, content)
    c2.font = Font(name='Calibri', size=10, color='2C3E50')
    c2.alignment = Alignment(wrap_text=True, vertical='top')
    line_count = content.count("\n") + 1
    ws2.row_dimensions[r].height = max(40, line_count * 18 + 25)

wb.save(OUTPUT)
ws_main = wb.active
print(f"v2.2 guardado: {OUTPUT}")
print(f"Pestanas: {wb.sheetnames}")
print(f"Filas de datos: {ws_main.max_row - 4}")
print(f"Celdas combinadas: {len(list(ws_main.merged_cells.ranges))}")
print("v2.2 COMPLETADO")
