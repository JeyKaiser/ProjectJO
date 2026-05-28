"""
extract_matriz_light.py — Reduce un xlsx pesado (312MB) a uno liviano (~100KB)
eliminando imagenes, formatos, estilos y columnas innecesarias.
Guarda solo las primeras 214 columnas con datos de texto/numero/fecha.

Uso:
  python migracion/extract_matriz_light.py
  python migracion/extract_matriz_light.py --input "ruta/archivo.xlsx" --output "ruta/output.xlsx"
"""
import sys, os, argparse
import openpyxl
from openpyxl.utils import get_column_letter

def main():
    parser = argparse.ArgumentParser(description="Reduce xlsx pesado a version liviana")
    parser.add_argument("--input", default=None,
        help="Archivo xlsx de entrada (default: Documentos/Copia de CONTRAM WS27.xlsx)")
    parser.add_argument("--output", default=None,
        help="Archivo de salida (default: mismo nombre + _LIGHT)")
    parser.add_argument("--max-col", type=int, default=214,
        help="Numero maximo de columnas a extraer (default: 214 = HD)")
    args = parser.parse_args()

    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file = args.input or os.path.join(base, "Documentos", "Copia de CONTRAM WS27.xlsx")

    if args.output:
        output_file = args.output
    else:
        name, ext = os.path.splitext(input_file)
        output_file = name + "_LIGHT" + ext

    if not os.path.exists(input_file):
        print(f"ERROR: No existe {input_file}")
        sys.exit(1)

    print(f"Input:  {input_file}")
    print(f"Output: {output_file}")
    print(f"Extrayendo columnas A-{get_column_letter(args.max_col)}...")
    print()

    # Leer con read_only (eficiente para archivos grandes)
    wb_in = openpyxl.load_workbook(input_file, read_only=True, data_only=True)

    # Crear workbook de salida (sin read_only, para poder guardar)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # Remove default sheet

    for sheet_name in wb_in.sheetnames:
        print(f"  Procesando hoja: {sheet_name}")
        ws_in = wb_in[sheet_name]
        ws_out = wb_out.create_sheet(title=sheet_name)

        row_count = 0
        for row in ws_in.iter_rows(min_row=1, max_col=args.max_col, values_only=True):
            # Escribir solo celdas con datos (no None)
            out_row_idx = row_count + 1
            for col_idx, value in enumerate(row, start=1):
                if value is not None:
                    # Si es string muy largo con base64 de imagen, saltarlo
                    if isinstance(value, str) and len(value) > 500:
                        continue
                    ws_out.cell(row=out_row_idx, column=col_idx, value=value)
            row_count += 1
            if row_count % 25 == 0:
                print(f"    {row_count} filas...")

        print(f"    Total: {row_count} filas extraidas")

    wb_in.close()
    wb_out.save(output_file)
    size_kb = os.path.getsize(output_file) / 1024
    print(f"\nArchivo liviano guardado: {output_file} ({size_kb:.0f} KB)")
    print("Listo para usar en la app (Importar) o con etl_matriz.py")

if __name__ == "__main__":
    main()
