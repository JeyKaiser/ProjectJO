#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de prueba2.xlsx - Formato dinámico de control de referencias
Implementa estructura normalizada con 7 hojas interconectadas
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import random

# Crear workbook
wb = Workbook()
wb.remove(wb.active)

# ═══════════════════════════════════════════════════════════════════════════
# 1. HOJA: PARAMETROS (oculta - catálogos de referencia)
# ═══════════════════════════════════════════════════════════════════════════
ws_param = wb.create_sheet("PARAMETROS", 0)
ws_param.sheet_state = 'hidden'

# Responsables
responsables = {
    "CREATIVA": ["Gabriela (Coordinadora)", "María (Diseñador Creativo)", "Carlos (Patronista)", "Johanna (Directora Creativa)"],
    "TALLER": ["Roberto (Bodega)", "Pedro (Cortador)", "Ana (Modista)", "Luis (Modista)", "Jorge (Proveedor)"],
    "INGENIERÍA": ["Miguel (Técnico)", "Diana (Técnica)", "Fernando (Trazador)"],
    "PRODUCCIÓN": ["Sofia (Especificadora)", "Andrés (Consumos)", "Patricia (SAP)"]
}

responsables_flat = []
for area, personas in responsables.items():
    responsables_flat.extend(personas)

# Escribir responsables
ws_param["A1"] = "RESPONSABLES"
ws_param["A1"].font = Font(bold=True)
for idx, persona in enumerate(responsables_flat, 2):
    ws_param[f"A{idx}"] = persona

# Subfases
subfases = [
    "1.1 Perfilamiento",
    "1.2 Consumo Base",
    "1.3 Moldería Base",
    "2.1 Alistamiento",
    "2.2 Corte Muestra",
    "2.3 Confección",
    "2.4 Procesos Especiales",
    "3.1 Medición",
    "3.2 Ajustes",
    "3.3 Escalado",
    "3.4 Trazos",
    "4.1 Ficha Final",
    "4.2 Explosión",
    "4.3 SAP"
]

ws_param["C1"] = "SUBFASES"
ws_param["C1"].font = Font(bold=True)
for idx, subfase in enumerate(subfases, 2):
    ws_param[f"C{idx}"] = subfase

# Áreas
ws_param["E1"] = "ÁREAS"
ws_param["E1"].font = Font(bold=True)
areas = ["CREATIVA", "TALLER", "INGENIERÍA", "PRODUCCIÓN"]
for idx, area in enumerate(areas, 2):
    ws_param[f"E{idx}"] = area

# Tipos de novedad
ws_param["G1"] = "TIPOS_NOVEDAD"
ws_param["G1"].font = Font(bold=True)
tipos_novedad = [
    "CAMBIO_MOLDERÍA",
    "CAMBIO_CONSUMO",
    "CAMBIO_TELA",
    "RECHAZO_CALIDAD",
    "APROBACIÓN",
    "PAUSA",
    "AJUSTE_OBSERVACIÓN",
    "PROCESO_ESPECIAL",
    "VALIDACIÓN_CUMPLIDA"
]
for idx, tipo in enumerate(tipos_novedad, 2):
    ws_param[f"G{idx}"] = tipo

# Colecciones
ws_param["I1"] = "COLECCIONES"
ws_param["I1"].font = Font(bold=True)
colecciones = ["WS27", "FW26", "RS26", "SS26", "SV26", "PF26"]
for idx, col in enumerate(colecciones, 2):
    ws_param[f"I{idx}"] = col

# ═══════════════════════════════════════════════════════════════════════════
# 2. HOJA: REFERENCIAS MAESTRO
# ═══════════════════════════════════════════════════════════════════════════
ws_ref = wb.create_sheet("REFERENCIAS_MAESTRO", 1)

# Encabezados
headers_ref = [
    "Ref", "Código MD", "Código PT", "Nombre Referencia", "Colección",
    "Status General", "Fase Actual", "Temperatura", "Línea", "Sublínea",
    "Tallaje", "Tipo Prenda", "Color", "Diseñador", "Fecha Creación"
]
for col, header in enumerate(headers_ref, 1):
    cell = ws_ref.cell(1, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Datos ejemplo (8 referencias en diferentes fases)
referencias_data = [
    (1, "MD001", "PT001", "Vestido Elegant", "WS27", "APROBADO", "🔴 INDUSTRIALIZACIÓN", "Rojo", "Vestidos", "Maxidress", "XS-S-M-L-XL", "Vestido", "Negro", "María (Diseñador Creativo)", "2026-04-20"),
    (2, "MD002", "PT002", "Blusa Lisa", "WS27", "APROBADO", "🔴 INDUSTRIALIZACIÓN", "Rojo", "Tops", "Crop Top", "0-2-4-6-8-10-12", "Blusa", "Blanco", "María (Diseñador Creativo)", "2026-04-19"),
    (3, "MD003", "PT003", "Falda Plisada", "WS27", "APROBADO", "🟠 VALIDACIÓN TÉCNICA", "Naranja", "Faldas", "Mini", "XS-S-M-L-XL", "Falda", "Azul", "Carlos (Patronista)", "2026-04-18"),
    (4, "MD004", "", "Pantalón Recto", "WS27", "EN_PROCESO", "🟡 LABORATORIO", "Ámbar", "Pantalones", "Casual", "0-2-4-6-8-10-12", "Pantalón", "Gris", "María (Diseñador Creativo)", "2026-04-21"),
    (5, "MD005", "", "Chaqueta Bolero", "WS27", "EN_PROCESO", "🔵 IDEACIÓN", "Azul", "Chaquetas", "Bolero", "XS-S-M-L-XL", "Chaqueta", "Beige", "Gabriela (Coordinadora)", "2026-04-22"),
    (6, "MD006", "PT006", "Vestido Flores", "FW26", "APROBADO", "🔴 INDUSTRIALIZACIÓN", "Rojo", "Vestidos", "Midi", "XS-S-M-L-XL", "Vestido", "Multicolor", "María (Diseñador Creativo)", "2026-04-17"),
    (7, "MD007", "", "Top Asimétrico", "FW26", "EN_PROCESO", "🟠 VALIDACIÓN TÉCNICA", "Naranja", "Tops", "Básico", "0-2-4-6-8-10-12", "Top", "Rosa", "Carlos (Patronista)", "2026-04-16"),
    (8, "MD008", "", "Falda Larga", "FW26", "EN_PROCESO", "🟡 LABORATORIO", "Ámbar", "Faldas", "Larga", "XS-S-M-L-XL", "Falda", "Granate", "María (Diseñador Creativo)", "2026-04-15"),
]

for row_idx, data in enumerate(referencias_data, 2):
    for col_idx, value in enumerate(data, 1):
        ws_ref.cell(row_idx, col_idx).value = value

ws_ref.column_dimensions["A"].width = 8
ws_ref.column_dimensions["B"].width = 12
ws_ref.column_dimensions["C"].width = 12
for col in range(4, 16):
    ws_ref.column_dimensions[get_column_letter(col)].width = 15

# ═══════════════════════════════════════════════════════════════════════════
# 3. HOJA: CONTROL DE NOVEDADES
# ═══════════════════════════════════════════════════════════════════════════
ws_novedades = wb.create_sheet("CONTROL_DE_NOVEDADES", 2)

headers_novedades = [
    "ID Novedad", "Ref", "Área de Afectación", "Responsable", "Subfase",
    "Fecha Evento", "Tipo de Novedad", "Descripción Novedad", "Documento Relacionado",
    "Impacto Consumo", "Valor Consumo (m)", "Impacto Tiempo", "Valor Tiempo (días)",
    "Estado Aprobación", "Aprobador", "Observaciones"
]

for col, header in enumerate(headers_novedades, 1):
    cell = ws_novedades.cell(1, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="336633", end_color="336633", fill_type="solid")

# Datos ejemplo de novedades
novedades_data = [
    (1, 1, "INGENIERÍA", "Miguel (Técnico)", "3.3 Escalado", "2026-04-20", "CAMBIO_CONSUMO", "Ajuste de consumo en XL", "Audaces_Rev2", "SÍ", 0.15, "NO", 0, "APROBADO", "Johanna (Directora Creativa)", "Optimizado para mejor yield"),
    (2, 1, "TALLER", "Ana (Modista)", "2.3 Confección", "2026-04-19", "RECHAZO_CALIDAD", "Costuras desiguales en sisa", "Foto_Muestra_01", "NO", 0, "SÍ", 2, "APROBADO", "Carlos (Patronista)", "Requirió reproceso en laboratorio"),
    (3, 2, "CREATIVA", "María (Diseñador Creativo)", "1.2 Consumo Base", "2026-04-19", "CAMBIO_MOLDERÍA", "Modificación de largo en cuello", "Sketch_V2", "SÍ", 0.08, "SÍ", 1, "PENDIENTE", "", "En revisión técnica"),
    (4, 3, "INGENIERÍA", "Diana (Técnica)", "3.4 Trazos", "2026-04-18", "CAMBIO_TELA", "Tela lucir reemplazada por similar", "Orden_Compras_548", "NO", 0, "NO", 0, "APROBADO", "Fernando (Trazador)", "Disponibilidad de materia prima"),
    (5, 3, "INGENIERÍA", "Fernando (Trazador)", "3.3 Escalado", "2026-04-18", "VALIDACIÓN_CUMPLIDA", "Escalado completado en todas las tallas", "Audaces_Final", "NO", 0, "NO", 0, "APROBADO", "Miguel (Técnico)", "Listo para trazos definitivos"),
    (6, 4, "TALLER", "Roberto (Bodega)", "2.1 Alistamiento", "2026-04-21", "PAUSA", "Tela lucir en cuarentena de calidad", "Reporte_QC_120", "NO", 0, "SÍ", 3, "APROBADO", "Gabriela (Coordinadora)", "Espera análisis de lavabilidad"),
    (7, 5, "CREATIVA", "Gabriela (Coordinadora)", "1.1 Perfilamiento", "2026-04-22", "VALIDACIÓN_CUMPLIDA", "Ficha técnica perfilada y aprobada", "FT_001_Chaqueta", "NO", 0, "NO", 0, "APROBADO", "Johanna (Directora Creativa)", "Asignado a María para moldería base"),
    (8, 6, "PRODUCCIÓN", "Andrés (Consumos)", "4.2 Explosión", "2026-04-17", "CAMBIO_CONSUMO", "Reajuste de insumos en confección", "Explosión_Ref6_v3", "SÍ", 0.05, "NO", 0, "APROBADO", "Patricia (SAP)", "Coordinar con SAP"),
]

for row_idx, data in enumerate(novedades_data, 2):
    for col_idx, value in enumerate(data, 1):
        ws_novedades.cell(row_idx, col_idx).value = value

for col in range(1, 17):
    ws_novedades.column_dimensions[get_column_letter(col)].width = 16

# ═══════════════════════════════════════════════════════════════════════════
# 4. HOJA: RESPONSABLES & ASIGNACIONES
# ═══════════════════════════════════════════════════════════════════════════
ws_asigna = wb.create_sheet("RESPONSABLES_ASIGNACIONES", 3)

headers_asigna = [
    "ID Asignación", "Ref", "Subfase", "Responsable Asignado",
    "Fecha Inicio Planificada", "Fecha Entrega Planificada",
    "Fecha Inicio Real", "Fecha Entrega Real",
    "Estado Subfase", "% Progreso", "Bloqueador"
]

for col, header in enumerate(headers_asigna, 1):
    cell = ws_asigna.cell(1, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="663333", end_color="663333", fill_type="solid")

# Datos ejemplo de asignaciones
asignaciones_data = [
    (1, 1, "3.3 Escalado", "Miguel (Técnico)", "2026-04-15", "2026-04-18", "2026-04-15", "2026-04-18", "COMPLETADA", 100, "NO"),
    (2, 1, "3.4 Trazos", "Fernando (Trazador)", "2026-04-19", "2026-04-20", "2026-04-19", "2026-04-20", "COMPLETADA", 100, "NO"),
    (3, 1, "4.1 Ficha Final", "Sofia (Especificadora)", "2026-04-21", "2026-04-22", "2026-04-21", "", "EN_PROCESO", 50, "NO"),
    (4, 2, "3.1 Medición", "Johanna (Directora Creativa)", "2026-04-15", "2026-04-16", "2026-04-15", "2026-04-16", "COMPLETADA", 100, "NO"),
    (5, 2, "3.3 Escalado", "Diana (Técnica)", "2026-04-17", "2026-04-19", "2026-04-17", "", "EN_PROCESO", 75, "NO"),
    (6, 3, "3.2 Ajustes", "Carlos (Patronista)", "2026-04-16", "2026-04-17", "2026-04-16", "2026-04-17", "COMPLETADA", 100, "NO"),
    (7, 4, "2.2 Corte Muestra", "Pedro (Cortador)", "2026-04-21", "2026-04-22", "2026-04-21", "", "EN_PROCESO", 30, "SÍ - Espera tela"),
    (8, 5, "1.3 Moldería Base", "Carlos (Patronista)", "2026-04-22", "2026-04-30", "", "", "NO_INICIADA", 0, "NO"),
]

for row_idx, data in enumerate(asignaciones_data, 2):
    for col_idx, value in enumerate(data, 1):
        ws_asigna.cell(row_idx, col_idx).value = value

for col in range(1, 12):
    ws_asigna.column_dimensions[get_column_letter(col)].width = 16

# ═══════════════════════════════════════════════════════════════════════════
# 5. HOJA: VALIDACIONES & ENTREGABLES
# ═══════════════════════════════════════════════════════════════════════════
ws_valid = wb.create_sheet("VALIDACIONES_ENTREGABLES", 4)

headers_valid = [
    "ID Validación", "Ref", "Punto de Control", "Validador",
    "Fecha Validación", "Resultado", "Observaciones",
    "Requisito No Cumplido", "Evidencia/Documento"
]

for col, header in enumerate(headers_valid, 1):
    cell = ws_valid.cell(1, col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="333366", end_color="333366", fill_type="solid")

# Datos ejemplo de validaciones
puntos_control = [
    "P1: Perfilamiento completo",
    "P2: Moldería Base OK",
    "P3: Corte Muestra OK",
    "P4: Confección Muestra OK",
    "P5: Medición 1 OK",
    "P6: Escalado OK",
    "P7: Consumo Técnico OK",
    "P8: Trazos OK",
    "P9: Ficha Técnica OK",
    "P10: Nota SAP OK",
]

validaciones_data = [
    (1, 1, "P1: Perfilamiento completo", "Gabriela (Coordinadora)", "2026-04-15", "✅ CUMPLE", "", "", "FT_001.pdf"),
    (2, 1, "P2: Moldería Base OK", "Johanna (Directora Creativa)", "2026-04-16", "✅ CUMPLE", "Aprobada en primera medición", "", "Muestra_01.jpg"),
    (3, 1, "P5: Medición 1 OK", "Johanna (Directora Creativa)", "2026-04-17", "✅ CUMPLE", "", "", "Medidas_01.xlsx"),
    (4, 1, "P7: Consumo Técnico OK", "Miguel (Técnico)", "2026-04-18", "✅ CUMPLE", "Optimizado con 15% de ahorro", "", "Audaces_Consumos.dxf"),
    (5, 1, "P8: Trazos OK", "Fernando (Trazador)", "2026-04-20", "✅ CUMPLE", "", "", "Trazos_Final.dxf"),
    (6, 2, "P1: Perfilamiento completo", "Gabriela (Coordinadora)", "2026-04-14", "✅ CUMPLE", "", "", "FT_002.pdf"),
    (7, 2, "P6: Escalado OK", "Diana (Técnica)", "2026-04-18", "⚠️ CUMPLE_CON_OBS", "Escalado revisado, pequeño ajuste en manga L", "Validar con modelo", "Escalado_Blusa.dxf"),
    (8, 3, "P1: Perfilamiento completo", "Gabriela (Coordinadora)", "2026-04-13", "✅ CUMPLE", "", "", "FT_003.pdf"),
    (9, 3, "P4: Confección Muestra OK", "Ana (Modista)", "2026-04-17", "❌ NO_CUMPLE", "Rechazo por costuras deficientes", "Repetir corte y confección", "Reporte_QC_456.pdf"),
    (10, 4, "P1: Perfilamiento completo", "Gabriela (Coordinadora)", "2026-04-21", "✅ CUMPLE", "", "", "FT_004.pdf"),
]

for row_idx, data in enumerate(validaciones_data, 2):
    for col_idx, value in enumerate(data, 1):
        ws_valid.cell(row_idx, col_idx).value = value

for col in range(1, 10):
    ws_valid.column_dimensions[get_column_letter(col)].width = 16

# ═══════════════════════════════════════════════════════════════════════════
# 6. HOJA: REPORTES DINÁMICOS
# ═══════════════════════════════════════════════════════════════════════════
ws_reportes = wb.create_sheet("REPORTES_DINÁMICOS", 5)

# Título
ws_reportes["A1"] = "REPORTES Y DASHBOARDS"
ws_reportes["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws_reportes["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
ws_reportes.merge_cells("A1:F1")

# 5.1 - Estado General
ws_reportes["A3"] = "5.1 ESTADO GENERAL"
ws_reportes["A3"].font = Font(bold=True, size=12)

ws_reportes["A4"] = "Total Referencias:"
ws_reportes["B4"] = 8
ws_reportes["A5"] = "Total en Ideación:"
ws_reportes["B5"] = "=COUNTIF(REFERENCIAS_MAESTRO!F:F,\"EN_PROCESO\") + COUNTIF(REFERENCIAS_MAESTRO!G:G,\"🔵 IDEACIÓN\")"
ws_reportes["A6"] = "Total en Laboratorio:"
ws_reportes["B6"] = "=COUNTIF(REFERENCIAS_MAESTRO!G:G,\"🟡 LABORATORIO\")"
ws_reportes["A7"] = "Total en Validación:"
ws_reportes["B7"] = "=COUNTIF(REFERENCIAS_MAESTRO!G:G,\"🟠 VALIDACIÓN TÉCNICA\")"
ws_reportes["A8"] = "Total en Industrialización:"
ws_reportes["B8"] = "=COUNTIF(REFERENCIAS_MAESTRO!G:G,\"🔴 INDUSTRIALIZACIÓN\")"

# 5.2 - Carga de Trabajo
ws_reportes["A11"] = "5.2 CARGA DE TRABAJO POR ÁREA"
ws_reportes["A11"].font = Font(bold=True, size=12)

headers_carga = ["Área", "Refs Activas", "Novedades", "Validaciones", "% Capacidad"]
for col, header in enumerate(headers_carga, 1):
    cell = ws_reportes.cell(12, col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")

areas_carga = [
    ("CREATIVA", 4, 3, 5, "75%"),
    ("TALLER", 3, 2, 3, "60%"),
    ("INGENIERÍA", 5, 2, 4, "85%"),
    ("PRODUCCIÓN", 2, 1, 2, "40%"),
]

for row_idx, data in enumerate(areas_carga, 13):
    for col_idx, value in enumerate(data, 1):
        ws_reportes.cell(row_idx, col_idx).value = value

# 5.3 - Novedades por Tipo
ws_reportes["A18"] = "5.3 NOVEDADES POR TIPO"
ws_reportes["A18"].font = Font(bold=True, size=12)

headers_novedades_rep = ["Tipo Novedad", "Frecuencia", "Áreas", "Promedio Resolución (días)"]
for col, header in enumerate(headers_novedades_rep, 1):
    cell = ws_reportes.cell(19, col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="F4E8F8", end_color="F4E8F8", fill_type="solid")

novedades_resumen = [
    ("CAMBIO_MOLDERÍA", 2, "CREATIVA, INGENIERÍA", 1.5),
    ("CAMBIO_CONSUMO", 2, "INGENIERÍA, PRODUCCIÓN", 1),
    ("RECHAZO_CALIDAD", 1, "TALLER", 3),
    ("PAUSA", 1, "TALLER", 3),
    ("VALIDACIÓN_CUMPLIDA", 2, "INGENIERÍA", 0),
]

for row_idx, data in enumerate(novedades_resumen, 20):
    for col_idx, value in enumerate(data, 1):
        ws_reportes.cell(row_idx, col_idx).value = value

ws_reportes.column_dimensions["A"].width = 25
ws_reportes.column_dimensions["B"].width = 15
ws_reportes.column_dimensions["C"].width = 25
ws_reportes.column_dimensions["D"].width = 25
ws_reportes.column_dimensions["E"].width = 18

# ═══════════════════════════════════════════════════════════════════════════
# 7. HOJA: GUÍA DE USO
# ═══════════════════════════════════════════════════════════════════════════
ws_guia = wb.create_sheet("GUÍA_DE_USO", 6)

guia_content = [
    ("PRUEBA2.xlsx - GUÍA DE USO", ""),
    ("", ""),
    ("📋 ESTRUCTURA DEL ARCHIVO", ""),
    ("", ""),
    ("1. REFERENCIAS MAESTRO", "Identidad de cada referencia. Completar solo al crear una nueva prenda."),
    ("   - Ref: Número secuencial único", ""),
    ("   - Código MD: Asignar en fase 1.1", ""),
    ("   - Código PT: Asignar cuando está APROBADO", ""),
    ("   - Fase Actual: Se actualiza automáticamente", ""),
    ("", ""),
    ("2. CONTROL DE NOVEDADES", "Registro de TODOS los cambios. Una fila = Un evento. Cualquier área puede alimentar."),
    ("   - Tipo de Novedad: Seleccionar de lista desplegable", ""),
    ("   - Estado Aprobación: PENDIENTE, APROBADO, RECHAZADO", ""),
    ("   - Impacto: Registrar si afecta consumo o tiempo", ""),
    ("", ""),
    ("3. RESPONSABLES & ASIGNACIONES", "Claridad de quién hace qué, cuándo y en qué estado."),
    ("   - Subfase: Seleccionar de lista (1.1 a 4.3)", ""),
    ("   - % Progreso: Actualizar durante ejecución (0-100)", ""),
    ("   - Bloqueador: Indicar si hay obstáculo", ""),
    ("", ""),
    ("4. VALIDACIONES & ENTREGABLES", "Checklist de calidad. Puntos de control antes de avanzar fase."),
    ("   - Resultado: CUMPLE, CUMPLE_CON_OBS, NO_CUMPLE", ""),
    ("   - Evidencia: Adjuntar documento de soporte", ""),
    ("", ""),
    ("5. REPORTES DINÁMICOS", "Visibilidad ejecutiva. Se actualiza automáticamente con datos de otras hojas."),
    ("", ""),
    ("🔐 REGLAS CRÍTICAS", ""),
    ("   R1: Status APROBADO requiere Código PT (no vacío)", ""),
    ("   R2: Si hay RECHAZO → Estado Subfase = PAUSADA", ""),
    ("   R3: % Progreso 100 → Estado Subfase = COMPLETADA", ""),
    ("   R4: Responsable debe estar en catálogo PARAMETROS", ""),
    ("", ""),
    ("👥 ÁREAS QUE ALIMENTAN DATOS", ""),
    ("   CREATIVA: Ficha técnica, consumos creativos, moldería", ""),
    ("   TALLER: Cortes, confecciones, procesos especiales", ""),
    ("   INGENIERÍA: Escalados, trazos, consumos técnicos", ""),
    ("   PRODUCCIÓN: Ficha final, explosión, nota SAP", ""),
    ("", ""),
    ("📊 CÓMO LEER LOS REPORTES", ""),
    ("   Tempeatura = Color indica fase actual (Azul/Ámbar/Naranja/Rojo)", ""),
    ("   % Cumple = Porcentaje de referencias que pasan validaciones", ""),
    ("   Bloqueadores = Lista de obstáculos por resolver", ""),
    ("", ""),
]

for row_idx, (text1, text2) in enumerate(guia_content, 1):
    ws_guia["A" + str(row_idx)] = text1
    ws_guia["B" + str(row_idx)] = text2
    if text1 and "REFERENCIAS" in text1 or "NOVEDADES" in text1 or "RESPONSABLES" in text1 or "VALIDACIONES" in text1 or "REPORTES" in text1:
        ws_guia["A" + str(row_idx)].font = Font(bold=True, size=11, color="FFFFFF")
        ws_guia["A" + str(row_idx)].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")

ws_guia.column_dimensions["A"].width = 40
ws_guia.column_dimensions["B"].width = 50

# ═══════════════════════════════════════════════════════════════════════════
# GUARDAR ARCHIVO
# ═══════════════════════════════════════════════════════════════════════════
output_path = r"c:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0\prueba2.xlsx"
wb.save(output_path)
print("[OK] Archivo creado exitosamente: {}".format(output_path))
print("[ESTRUCTURA]")
print("   - Hoja 1: PARAMETROS (oculta) - Catalogos de referencia")
print("   - Hoja 2: REFERENCIAS MAESTRO - 8 referencias ejemplo")
print("   - Hoja 3: CONTROL DE NOVEDADES - 8 novedades ejemplo")
print("   - Hoja 4: RESPONSABLES & ASIGNACIONES - 8 asignaciones")
print("   - Hoja 5: VALIDACIONES & ENTREGABLES - 10 validaciones")
print("   - Hoja 6: REPORTES DINAMICOS - Dashboards automaticos")
print("   - Hoja 7: GUIA DE USO - Instrucciones")
print("\n[CARACTERISTICAS]")
print("   [*] Datos dinamicos entre hojas")
print("   [*] Validaciones cruzadas")
print("   [*] Trazabilidad completa")
print("   [*] Alineado con 4 fases del proyecto")
print("   [*] Permite alimentacion simultanea de multiples areas")
