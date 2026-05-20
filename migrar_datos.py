"""
Script de migracion de datos desde archivos en Documentos/ hacia prueba.xlsx
+ creacion de hoja PARAMETROS con listas maestras para dropdowns.

Arquitectura:
  - VALIDACION DE TELAS WS27: 1 fila por tela/insumo (1NF) -> SECUENCIA REFERENCIA
  - CONTRAMUESTRAS WS27 MATRIZ: 1 fila por referencia -> complementa datos de FASE 3 y 4
  - FO APPAREL WS27: 1 fila por referencia -> complementa FASE 1, STATUS, TIME COLLECTION
  - FO APPAREL FW26: 1 fila por referencia -> otra coleccion aparte
  - PARAMETROS POR TELA: -> hoja PARAMETROS
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import os
import re
from collections import defaultdict

DOCS = r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0\Documentos"
PRUEBA = r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0\prueba.xlsx"

# ── Cargar prueba.xlsx ──
wb = openpyxl.load_workbook(PRUEBA)
ws = wb["SECUENCIA REFERENCIA"]

# ── Construir indice de columnas (nombre_col -> numero_col) ──
col_index = {}
for col in range(1, ws.max_column + 1):
    name = ws.cell(row=3, column=col).value
    if name:
        col_index[str(name).strip().upper()] = col

print(f"Columnas mapeadas: {len(col_index)}")

def find_col(keyword):
    """Busca columna por palabra clave (case-insensitive)."""
    for name, col in col_index.items():
        if keyword.upper() in name:
            return col
    return None

# ── Estilos para datos ──
thin_border = Border(
    left=Side(style='thin', color='808080'),
    right=Side(style='thin', color='808080'),
    top=Side(style='thin', color='808080'),
    bottom=Side(style='thin', color='808080')
)
data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def safe(val):
    """Convierte a string limpio, reemplaza #N/A, N/A."""
    if val is None or str(val).strip() in ('', '#N/A', 'N/A', '#REF!'):
        return ''
    s = str(val).strip()
    if s.upper() == 'N/A':
        return ''
    return s

def clean_consumo(val):
    """Limpia valores de consumo (ej. '2,09' -> 2.09, '2.16 MTS' -> 2.16)."""
    if not val:
        return ''
    s = str(val).strip()
    s = s.upper().replace('MTS', '').replace('MT', '').replace('CMS', '').strip()
    s = s.replace(',', '.')
    try:
        return str(float(s))
    except:
        return s

def set_cell(ws, row, col_keyword, value, default_align=True):
    """Escribe en celda buscando columna por keyword. Retorna True si escribio."""
    if not value or str(value).strip() == '':
        return False
    col = find_col(col_keyword)
    if col:
        cell = ws.cell(row=row, column=col)
        if cell.value is None or str(cell.value).strip() == '':
            cell.value = str(value).strip()[:32767]
            if default_align:
                cell.alignment = data_align
            cell.border = thin_border
    return col is not None

def set_cell_direct(ws, row, col_num, value):
    """Escribe directamente por numero de columna."""
    if not value or str(value).strip() == '':
        return
    cell = ws.cell(row=row, column=col_num)
    if cell.value is None or str(cell.value).strip() == '':
        cell.value = str(value).strip()[:32767]
        cell.alignment = data_align
        cell.border = thin_border

# ═══════════════════════════════════════════════════════════
# 1. PROCESAR VALIDACION DE TELAS WS27 (estructura 1NF)
# ═══════════════════════════════════════════════════════════
print("\n=== [1/5] VALIDACION DE TELAS WS27 ===")
csv_path = os.path.join(DOCS, "Copia de COLECCION WS27 - 1.VALIDACION DE TELAS.csv")

rows_data = []
with open(csv_path, 'r', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    all_rows = list(reader)

# Encontrar la fila de headers real (row 4, index 3 en 0-based)
# El header row tiene "REF,FOTO,REFERENCIA,STATUS,MOD.ARTE,..."
header_idx = None
for i, row in enumerate(all_rows):
    if row and row[0].strip() == 'REF' and 'FOTO' in str(row):
        header_idx = i
        break

if header_idx is None:
    header_idx = 3  # fallback row 4

headers = [str(h).strip() for h in all_rows[header_idx]]
print(f"Header row index: {header_idx}, columns: {len(headers)}")
print(f"Headers: {headers[:15]}...")

# Mapeo manual de columnas del CSV por posicion
# Estructura del CSV (indice 0-based):
CSV_COL = {
    'REF': 0, 'FOTO': 1, 'REFERENCIA': 2, 'STATUS': 3,
    'MOD_ARTE': 4, 'UBI_TRAZO': 5, 'VARIACION_COLOR': 6,
    'LARGO': 8, 'USO_PRENDA': 9, 'CODIGO_TELA': 10,
    'DESCRIPCION_TELA': 11, 'ANCHO': 12, 'TELA_FOTO': 13,
    'CONSUMO_BASE': 14,
    # EQUIPO CREATIVO (cols 15-20)
    'DISENADOR_CREATIVO': 15, 'CAMBIO_MOLDERIA': 16,
    'CONSUMO1': 17, 'CONSUMO2': 18, 'CONSUMO3': 19, 'OBS_CREATIVO': 20,
    # DISENADOR TECNICO (cols 21-)
    'DISENADOR_TECNICO': 21,
    # SOLIDO
    'TALLA_SOLIDO': 22, 'UND_SOLIDO': 23,
    'CONSUMO1_SOLIDO': 24, 'CONSUMO2_SOLIDO': 25, 'CONSUMO3_SOLIDO': 26,
    # MOD ARTE
    'TALLA_MOD': 27, 'UND_MOD': 28,
    'CONSUMO1_MOD': 29, 'CONSUMO2_MOD': 30, 'CONSUMO3_MOD': 31,
    # UBI TRAZO
    'TALLA_UBI': 32, 'UND_UBI': 33,
    'CONSUMO1_UBI': 34, 'CONSUMO2_UBI': 35, 'CONSUMO3_UBI': 36,
    # OBSERVACIONES/COMENTARIOS tecnico
    'OBS_TECNICO': 37,
    # A RIESGO / VALIDADO
    'A_RIESGO': 38,
    # TRAZO SOLIDO
    'TRAZO_TALLA_SOLIDO': 39, 'TRAZO_UND_SOLIDO': 40,
    'TRAZO_CONS1': 41, 'TRAZO_CONS2': 42, 'TRAZO_CONS3': 43, 'TRAZO_CONS4': 44,
    # TRAZO MOD ARTE
    'TRAZO_TALLA_MOD': 45, 'TRAZO_UND_MOD': 46,
    'TRAZO_CONS_MOD1': 47, 'TRAZO_CONS_MOD2': 48, 'TRAZO_CONS_MOD3': 49,
    # TRAZO UBI (cols 50-56)
    'TRAZO_TALLA_UBI': 50, 'TRAZO_UND_UBI': 51,
    'TRAZO_CONS_UBI1': 52, 'TRAZO_CONS_UBI2': 53,
    # CONSUMOS COSTEO (cols 54-66)
    'COSTEO_TALLA_SOLIDO': 54, 'COSTEO_UND_SOLIDO': 55,
    'COSTEO_CONS_SOLIDO': 56,
    'COSTEO_TALLA_MOD': 57, 'COSTEO_UND_MOD': 58,
    'COSTEO_CONS_MOD': 59,
    'COSTEO_TALLA_UBI': 60, 'COSTEO_UND_UBI': 61,
    'COSTEO_CONS_UBI': 62,
    'OBS_COSTEO': 63,
    'TEC': 64, 'TRAZADOR': 65,
    'TERMINADO_TALLER': 66,
}

current_ref = None
current_foto = None
current_nombre = None
current_status = None
current_mod_arte = None
current_ubi_trazo = None
current_var_color = None
current_largo = None
current_disenador_c = None
current_cambio_mold = None
current_consumo1 = None
current_consumo2 = None
current_consumo3 = None
current_obs_creativo = None

data_start = header_idx + 1
row_count = 0
skipped_lines = 0

# Find the last data row (before summary stats)
last_data_row = len(all_rows)

for i, csv_row in enumerate(all_rows[data_start:], start=data_start):
    if i >= len(all_rows):
        break
    
    ref_val = safe(csv_row[CSV_COL['REF']]) if len(csv_row) > CSV_COL['REF'] else ''
    
    # Skip non-data rows (empty REF, summary rows, etc.)
    if not ref_val:
        # Check if this row has tela data (continuation of previous ref)
        uso = safe(csv_row[CSV_COL['USO_PRENDA']]) if len(csv_row) > CSV_COL['USO_PRENDA'] else ''
        cod_tela = safe(csv_row[CSV_COL['CODIGO_TELA']]) if len(csv_row) > CSV_COL['CODIGO_TELA'] else ''
        if current_ref and (uso or cod_tela):
            ref_val = current_ref  # Keep same ref
        else:
            # Check for summary row
            last_cols = [safe(csv_row[c]) if c < len(csv_row) else '' for c in range(64, 68)]
            summary_text = ' '.join(last_cols)
            if any(kw in summary_text.upper() for kw in ['CONSUMOS', 'REF TERMINADAS', 'REF PDT', 'CONSUMOS EQUIPO']):
                skipped_lines += 1
                continue
            skipped_lines += 1
            continue
    
    # Skip non-numeric REF
    if not ref_val.isdigit():
        skipped_lines += 1
        continue
    
    # Update reference-level data
    if csv_row[CSV_COL['REF']] and safe(csv_row[CSV_COL['REF']]):
        current_ref = ref_val
        current_foto = safe(csv_row[CSV_COL['FOTO']]) if len(csv_row) > CSV_COL['FOTO'] else ''
        current_nombre = safe(csv_row[CSV_COL['REFERENCIA']]) if len(csv_row) > CSV_COL['REFERENCIA'] else ''
        current_status = safe(csv_row[CSV_COL['STATUS']]) if len(csv_row) > CSV_COL['STATUS'] else ''
        current_mod_arte = safe(csv_row[CSV_COL['MOD_ARTE']]) if len(csv_row) > CSV_COL['MOD_ARTE'] else ''
        current_ubi_trazo = safe(csv_row[CSV_COL['UBI_TRAZO']]) if len(csv_row) > CSV_COL['UBI_TRAZO'] else ''
        current_var_color = safe(csv_row[CSV_COL['VARIACION_COLOR']]) if len(csv_row) > CSV_COL['VARIACION_COLOR'] else ''
        current_largo = safe(csv_row[CSV_COL['LARGO']]) if len(csv_row) > CSV_COL['LARGO'] else ''
        current_disenador_c = safe(csv_row[CSV_COL['DISENADOR_CREATIVO']]) if len(csv_row) > CSV_COL['DISENADOR_CREATIVO'] else ''
        current_cambio_mold = safe(csv_row[CSV_COL['CAMBIO_MOLDERIA']]) if len(csv_row) > CSV_COL['CAMBIO_MOLDERIA'] else ''
        current_consumo1 = clean_consumo(csv_row[CSV_COL['CONSUMO1']]) if len(csv_row) > CSV_COL['CONSUMO1'] else ''
        current_consumo2 = clean_consumo(csv_row[CSV_COL['CONSUMO2']]) if len(csv_row) > CSV_COL['CONSUMO2'] else ''
        current_consumo3 = clean_consumo(csv_row[CSV_COL['CONSUMO3']]) if len(csv_row) > CSV_COL['CONSUMO3'] else ''
        current_obs_creativo = safe(csv_row[CSV_COL['OBS_CREATIVO']]) if len(csv_row) > CSV_COL['OBS_CREATIVO'] else ''
    
    # Build data row in SECUENCIA REFERENCIA (start from row 5, after example)
    target_row = 5 + row_count
    
    def getv(key):
        idx = CSV_COL.get(key, -1)
        if idx >= 0 and idx < len(csv_row):
            return safe(csv_row[idx])
        return ''
    
    # ── FASE 1: PERFILAMIENTO ──
    set_cell(ws, target_row, "REF", current_ref)
    set_cell(ws, target_row, "FOTO", current_foto)
    set_cell(ws, target_row, "NOMBRE REFERENCIA", current_nombre)
    set_cell(ws, target_row, "STATUS", current_status)
    set_cell(ws, target_row, "COLECCION", "WINTER SUN (WS27)")
    set_cell(ws, target_row, "LARGO", current_largo)
    set_cell(ws, target_row, "MOD ARTE SI/NO", current_mod_arte)
    set_cell(ws, target_row, "UBI TRAZO SI/NO", current_ubi_trazo)
    set_cell(ws, target_row, "VARIACION COLOR SI/NO", current_var_color)
    set_cell(ws, target_row, "DISENADOR CREATIVO", current_disenador_c)
    set_cell(ws, target_row, "CAMBIO MOLDERIA SI/NO", current_cambio_mold)
    set_cell(ws, target_row, "CONSUMO BASE", clean_consumo(csv_row[CSV_COL['CONSUMO_BASE']]) if len(csv_row) > CSV_COL['CONSUMO_BASE'] else '')
    set_cell(ws, target_row, "CONSUMO CREATIVO 1", clean_consumo(getv('CONSUMO1')))
    set_cell(ws, target_row, "CONSUMO CREATIVO 2", clean_consumo(getv('CONSUMO2')))
    set_cell(ws, target_row, "CONSUMO CREATIVO 3", clean_consumo(getv('CONSUMO3')))
    set_cell(ws, target_row, "OBSERVACIONES CREATIVO", current_obs_creativo)
    
    # ── FASE 2: TELAS E INSUMOS ──
    uso = getv('USO_PRENDA')
    cod_tela = getv('CODIGO_TELA')
    desc_tela = getv('DESCRIPCION_TELA')
    ancho = getv('ANCHO')
    foto_tela = getv('TELA_FOTO')
    
    set_cell(ws, target_row, "USO EN PRENDA", uso)
    set_cell(ws, target_row, "CODIGO TELA/INSUMO", cod_tela)
    set_cell(ws, target_row, "DESCRIPCION TELA", desc_tela)
    
    # Clean ancho
    ancho_clean = safe(ancho)
    if ancho_clean:
        ancho_clean = ancho_clean.replace(',', '.')
        set_cell(ws, target_row, "ANCHO TELA", ancho_clean)
    
    set_cell(ws, target_row, "FOTO TELA", foto_tela)
    
    # ── FASE 3: DISENADOR TECNICO, CONSUMOS TECNICOS ──
    set_cell(ws, target_row, "DISENADOR TECNICO ASIGNADO", getv('DISENADOR_TECNICO'))
    
    # Consumos solidos
    set_cell(ws, target_row, "CONSUMO SOLIDO", clean_consumo(getv('CONSUMO1_SOLIDO')))
    set_cell(ws, target_row, "TALLA CALCULO CONSUMO", getv('TALLA_SOLIDO'))
    set_cell(ws, target_row, "OBSERVACIONES TECNICO", getv('OBS_TECNICO'))
    
    # ── FASE 3: TRAZADOR ──
    set_cell(ws, target_row, "TRAZADOR ASIGNADO", getv('TRAZADOR'))
    set_cell(ws, target_row, "CONSUMO TRAZADOR SOLIDO", clean_consumo(getv('TRAZO_CONS1')))
    
    # ── CONSUMOS COSTEO ──
    set_cell(ws, target_row, "CONSUMO COSTEO SOLIDO", clean_consumo(getv('COSTEO_CONS_SOLIDO')))
    set_cell(ws, target_row, "A RIESGO / VALIDADO POR LINA PENA", getv('A_RIESGO'))
    set_cell(ws, target_row, "TERMINADO TALLER PARA COSTEO TELAS", getv('TERMINADO_TALLER'))
    
    row_count += 1

print(f"VALIDACION DE TELAS: {row_count} filas migradas, {skipped_lines} lineas saltadas")

# ═══════════════════════════════════════════════════════════
# 2. PARSEO CONTRAMUESTRAS WS27 - MATRIZ
# ═══════════════════════════════════════════════════════════
print("\n=== [2/5] CONTRAMUESTRAS WS27 - MATRIZ ===")
csv_cm = os.path.join(DOCS, "Copia de CONTRAMUESTRAS WS27 - MATRIZ.csv")

with open(csv_cm, 'r', encoding='utf-8-sig') as f:
    cm_rows = list(csv.reader(f))

# Row 9 (index 8) has primary headers, row 10 (index 9) has sub-headers nested
# Actually looking at the file, data rows start at 17 (index 16) based on earlier reading
# Let me get the header row which should be around row 9-10
# Data rows start around row 17-18 (after headers)
# Each data row has REF in first column or is blank for derived rows

cm_refs = {}  # ref -> dict of all column values (first row of that ref)
cm_data_start = None
for i, row in enumerate(cm_rows):
    if row and row[0].strip() and row[0].strip().isdigit():
        if cm_data_start is None:
            cm_data_start = i
        ref = row[0].strip()
        if ref not in cm_refs:
            cm_refs[ref] = row

print(f"CONTRAMUESTRAS: {len(cm_refs)} referencias encontradas (data start row {cm_data_start})")

# Map CONTRAMUESTRAS columns to SECUENCIA REFERENCIA
# Based on the descripcion proyecto.md, the CONTRAMUESTRAS file has a similar structure
# but the column layout differs from FO files. Let me map the key columns:

# From FO row 3 header analysis:
# Col 0: REF, 1: IMAGEN, 2: CODIGO MD, 3: CODIGO PT, 4: NOMBRE REF, 5: COLOR, 6: COD COLOR
# Col 7: FOTO BASADO EN, 8: PT, 9: BASADO EN, 10: STATUS, 11: DISEÑADOR
# Col 12: STATUS TALLER, 13: MODISTA, 14: FOTOS INTERNAS
# Col 15: LINEA, 16: SUBLINEA, 17: TIPO REF, 18: TALLAJE, 19: LARGO, 20: CLOSURE
# Col 21: LINNED, 22: INCLUDES, 23: INCLUDES PAQ COMPLETO
# ... (continues)

# However, the CONTRAMUESTRAS CSV has different column order.
# Let me build a flexible approach: find matches by header name in the CONTRAMUESTRAS header rows.

# For CONTRAMUESTRAS, the structure starts with same columns but diverges.
# Looking at row 9-10 of CONTRAMUESTRAS:
# Row 9: #REF!, IMAGEN, #REF!, #REF!, #REF!, #REF!, #REF!, FOTO BASADO EN, #REF!, #REF!, #REF!, #REF!, STATUS TALLER, MODISTA, FOTOS INTERNAS...
# Row 10 (continuation): #REF!, IMAGEN, #REF!, #REF!, #REF!, #REF!, #REF!, FOTO BASADO EN, #REF!, #REF!, #REF!, #REF!, ...
# So the first columns map similarly to FO:
# Col 0: REF (A)
# Col 1: IMAGEN (B) 
# Col 2: CODIGO MD (C)
# Col 3: CODIGO PT (D)
# Col 4: NOMBRE REFERENCIA (E)
# Col 5: COLOR (F)
# Col 6: COD COLOR (G)
# Col 7: FOTO BASADO EN (H)
# Col 8: PT referente (I)
# Col 9: BASADO EN (J)
# Col 10: STATUS (K)
# Col 11: DISEÑADOR (L)
# Col 12: STATUS TALLER (M)
# Col 13: MODISTA (N)
# Col 14: FOTOS INTERNAS (O)

# The CONTRAMUESTRAS then has more columns specific to its purpose.

# Let me map the CONTRAMUESTRAS to SECUENCIA REFERENCIA by finding matching REF in existing rows
# and adding data from CONTRAMUESTRAS columns.

cm_mapped = 0
for ref_num, cm_row in cm_refs.items():
    # Find matching row in SECUENCIA REFERENCIA
    matched = False
    for r in range(5, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value  # REF column
        if cell_val and str(cell_val).strip() == ref_num:
            matched = True
            target = r
            
            def cmget(idx):
                return safe(cm_row[idx]) if idx < len(cm_row) else ''
            
            # FASE 1 data
            set_cell(ws, target, "CODIGO MD", cmget(2))
            set_cell(ws, target, "CODIGO PT", cmget(3))
            set_cell(ws, target, "COLOR", cmget(5))
            set_cell(ws, target, "COD COLOR", cmget(6))
            set_cell(ws, target, "FOTO BASADO EN", cmget(7))
            set_cell(ws, target, "PT REFERENTE", cmget(8))
            set_cell(ws, target, "NOMBRE REFERENTE", cmget(9))
            set_cell(ws, target, "LINEA", cmget(15))
            set_cell(ws, target, "SUBLINEA", cmget(16))
            set_cell(ws, target, "TIPO REF", cmget(17))
            set_cell(ws, target, "TALLAJE", cmget(18))
            set_cell(ws, target, "LARGO", cmget(19))
            set_cell(ws, target, "CLOSURE", cmget(20))
            set_cell(ws, target, "FOTOS INTERNAS", cmget(14))
            
            # STATUS
            set_cell(ws, target, "STATUS TALLER", cmget(12))
            set_cell(ws, target, "MODISTA ASIGNADA", cmget(13))
            
            # Following columns in CONTRAMUESTRAS continue with various data
            # Let's map more based on known positions
            # Row 10 headers give clues about remaining columns
            
            cm_mapped += 1
            break
    
    if not matched and ref_num:
        # Create new row for unmatched refs
        target = ws.max_row + 1
        set_cell_direct(ws, target, 1, ref_num)  # REF
        
        def cmget(idx):
            return safe(cm_row[idx]) if idx < len(cm_row) else ''
        
        set_cell(ws, target, "CODIGO MD", cmget(2))
        set_cell(ws, target, "CODIGO PT", cmget(3))
        set_cell(ws, target, "COLOR", cmget(5))
        set_cell(ws, target, "STATUS", cmget(10))
        set_cell(ws, target, "DISENADOR CREATIVO", cmget(11))
        set_cell(ws, target, "STATUS TALLER", cmget(12))
        set_cell(ws, target, "MODISTA ASIGNADA", cmget(13))
        set_cell(ws, target, "LINEA", cmget(15))
        set_cell(ws, target, "SUBLINEA", cmget(16))
        set_cell(ws, target, "TIPO REF", cmget(17))
        set_cell(ws, target, "TALLAJE", cmget(18))
        cm_mapped += 1

print(f"CONTRAMUESTRAS: {cm_mapped} referencias migradas")

# ═══════════════════════════════════════════════════════════
# 3. PARSEO FO APPAREL 2026 - WS27
# ═══════════════════════════════════════════════════════════
print("\n=== [3/5] FO APPAREL WS27 ===")
csv_fo_ws27 = os.path.join(DOCS, "Copia de FO... APPAREL 2026 - WS27.csv")

with open(csv_fo_ws27, 'r', encoding='utf-8-sig') as f:
    fo_rows = list(csv.reader(f))

# Find header row (has REF in col 0)
fo_header_idx = None
for i, row in enumerate(fo_rows):
    if row and row[0].strip() == 'REF':
        fo_header_idx = i
        break

if fo_header_idx is None:
    fo_header_idx = 2  # row 3

fo_headers = [str(h).strip() for h in fo_rows[fo_header_idx]]
print(f"FO WS27 Header row: {fo_header_idx}, cols: {len(fo_headers)}")
print(f"Sample headers: {fo_headers[:20]}...")

fo_data_start = fo_header_idx + 1
fo_mapped = 0

# Map FO columns by header name
fo_col_map = {}
for i, h in enumerate(fo_headers):
    fo_col_map[h.upper().strip()] = i

for i in range(fo_data_start, len(fo_rows)):
    row = fo_rows[i]
    if not row or not row[0].strip():
        continue
    ref = safe(row[0])
    if not ref.isdigit():
        continue
    
    def foget(name):
        """Get value from FO row by column header name."""
        idx = None
        for h, j in fo_col_map.items():
            if name.upper().replace(' ', '') in h.replace(' ', '').replace('#REF!', '').replace('REF!', ''):
                idx = j
                break
        if idx is not None and idx < len(row):
            return safe(row[idx])
        return ''
    
    # Find target row
    target = None
    for r in range(5, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip() == ref:
            target = r
            break
    
    if target is None:
        target = ws.max_row + 1
        set_cell_direct(ws, target, 1, ref)
    
    # Map known headers from FO
    # REF, IMAGEN, CODIGO MD, CODIGO PT, NOMBRE REFERENCIA, COLOR, COD COLOR,
    # FOTO BASADO EN, PT, BASADO EN, STATUS, DISEÑADOR,
    # STATUS TALLER, MODISTA, FOTOS INTERNAS,
    # LINEA, SUBLINEA, TIPO REF, TALLAJE, LARGO, CLOSURE, LINNED,
    # INCLUDES, INCLUDES DE PAQUETE COMPLETO EN PLANTA
    # CODIGO TELA DE LUCIR PRINCIPAL EN MUESTRA, FOTO TELA, DESCRIPCION TELA, ANCHO TELA, BASE TEXTIL
    # CODIGO TELA DE FORRO, FOTO TELA, DESCRIPCION TELA, ANCHO TELA
    # MOD. DE ARTE, UBICACION EN TRAZO, ALL OVER, SUGERENCIA MOD/UBC
    # VARIACION DE COLOR, REF, TIPO DE EMPAQUE
    # APLICA O NO (bordado), APLICA O NO APLICA, DESCRIPCION (bordado)
    # APLICA O NO APLICA (semielaborado), DESCRIPCION
    # PROVEEDOR, PROCESO EXTERNO, COSTO
    
    # Direct index-based mapping (since headers may have subtle differences)
    def fo(idx):
        return safe(row[idx]) if idx < len(row) else ''
    
    set_cell(ws, target, "CODIGO MD", fo(2))
    set_cell(ws, target, "CODIGO PT", fo(3))
    set_cell(ws, target, "NOMBRE REFERENCIA", fo(4))
    set_cell(ws, target, "COLOR", fo(5))
    set_cell(ws, target, "COD COLOR", fo(6))
    set_cell(ws, target, "FOTO BASADO EN", fo(7))
    set_cell(ws, target, "PT REFERENTE", fo(8))
    set_cell(ws, target, "NOMBRE REFERENTE", fo(9))
    set_cell(ws, target, "STATUS", fo(10))
    set_cell(ws, target, "DISENADOR CREATIVO", fo(11))
    set_cell(ws, target, "STATUS TALLER", fo(12))
    set_cell(ws, target, "MODISTA ASIGNADA", fo(13))
    
    # Check if NOT cancelled/just-for-show
    cancel_status = str(fo(10)).upper()
    if cancel_status not in ('CANCELADO CORTADO', 'CANCELADO SIN CORTAR', 'JUST FOR SHOW', 'CANCELADO POR COMERCIAL'):
        set_cell(ws, target, "FOTOS INTERNAS", fo(14))
        set_cell(ws, target, "LINEA", fo(15))
        set_cell(ws, target, "SUBLINEA", fo(16))
        set_cell(ws, target, "TIPO REF", fo(17))
        set_cell(ws, target, "TALLAJE", fo(18))
        set_cell(ws, target, "LARGO", fo(19))
        set_cell(ws, target, "CLOSURE", fo(20))
        set_cell(ws, target, "LINNED", fo(21))
        set_cell(ws, target, "INCLUDES", fo(22))
        set_cell(ws, target, "INCLUDES PAQ. COMPLETO PLANTA", fo(23))
        
        # Tela de lucir
        set_cell(ws, target, "CODIGO TELA/INSUMO", fo(24))
        set_cell(ws, target, "FOTO TELA", fo(25))
        set_cell(ws, target, "DESCRIPCION TELA", fo(26))
        set_cell(ws, target, "ANCHO TELA", fo(27))
        set_cell(ws, target, "BASE TEXTIL", fo(28))
        
        # Forro
        set_cell(ws, target, "CODIGO TELA FORRO", fo(29))
        set_cell(ws, target, "DESCRIPCION FORRO", fo(31))
        set_cell(ws, target, "ANCHO FORRO", fo(32))
        
        # Catalogacion
        set_cell(ws, target, "MOD ARTE SI/NO", fo(33))
        set_cell(ws, target, "UBI TRAZO SI/NO", fo(34))
        set_cell(ws, target, "ALL OVER SI/NO", fo(35))
        set_cell(ws, target, "SUGERENCIA MOD/UBC", fo(36))
        set_cell(ws, target, "VARIACION COLOR SI/NO", fo(37))
        set_cell(ws, target, "REF VARIACION COLOR", fo(38))
        set_cell(ws, target, "TIPO DE EMPAQUE", fo(39))
        
        # Bordado / Semielaborado / Proceso externo
        set_cell(ws, target, "TIENE BORDADO?", fo(40))
        set_cell(ws, target, "DESCRIPCION BORDADO", fo(42))
        set_cell(ws, target, "TIENE SEMIELABORADO?", fo(43))
        set_cell(ws, target, "DESCRIPCION SEMIELABORADO", fo(44))
        set_cell(ws, target, "PROVEEDOR EXT.", fo(45))
        set_cell(ws, target, "PROCESO EXTERNO", fo(46))
        set_cell(ws, target, "COSTO PROCESO EXT.", fo(47))
        
        # COLECCION
        set_cell(ws, target, "COLECCION", "WINTER SUN (WS27)")

    fo_mapped += 1

print(f"FO WS27: {fo_mapped} referencias migradas")

# ═══════════════════════════════════════════════════════════
# 4. PARSEO FO APPAREL 2026 - FW26
# ═══════════════════════════════════════════════════════════
print("\n=== [4/5] FO APPAREL FW26 ===")
csv_fo_fw26 = os.path.join(DOCS, "Copia de FO... APPAREL 2026 - FW26.csv")

with open(csv_fo_fw26, 'r', encoding='utf-8-sig') as f:
    fo_fw26_rows = list(csv.reader(f))

fw_header_idx = None
for i, row in enumerate(fo_fw26_rows):
    if row and row[0].strip() == 'REF':
        fw_header_idx = i
        break
if fw_header_idx is None:
    fw_header_idx = 3

fw_data_start = fw_header_idx + 1
fw_mapped = 0

for i in range(fw_data_start, len(fo_fw26_rows)):
    row = fo_fw26_rows[i]
    if not row or not row[0].strip():
        continue
    ref = safe(row[0])
    if not ref.isdigit():
        continue
    
    def fw(idx):
        return safe(row[idx]) if idx < len(row) else ''
    
    # FW26 gets different collection label
    target = None
    for r in range(5, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip() == ref:
            target = r
            break
    
    if target is None:
        target = ws.max_row + 1
        set_cell_direct(ws, target, 1, ref)
    
    set_cell(ws, target, "CODIGO MD", fw(2))
    set_cell(ws, target, "CODIGO PT", fw(3))
    set_cell(ws, target, "NOMBRE REFERENCIA", fw(4))
    set_cell(ws, target, "COLOR", fw(5))
    set_cell(ws, target, "COD COLOR", fw(6))
    set_cell(ws, target, "FOTO BASADO EN", fw(7))
    set_cell(ws, target, "PT REFERENTE", fw(8))
    set_cell(ws, target, "NOMBRE REFERENTE", fw(9))
    set_cell(ws, target, "STATUS", fw(10))
    set_cell(ws, target, "DISENADOR CREATIVO", fw(11))
    
    # FW26 has STATUS TALLER, MODISTA, FOTOS INTERNAS at cols 12-14 but with #REF! headers
    # Use default approach
    set_cell(ws, target, "LINEA", fw(15))
    set_cell(ws, target, "SUBLINEA", fw(16))
    set_cell(ws, target, "TIPO REF", fw(17))
    set_cell(ws, target, "TALLAJE", fw(18))
    set_cell(ws, target, "LARGO", fw(19))
    set_cell(ws, target, "CLOSURE", fw(20))
    set_cell(ws, target, "LINNED", fw(21))
    set_cell(ws, target, "INCLUDES", fw(22))
    
    # Tela data
    set_cell(ws, target, "CODIGO TELA/INSUMO", fw(24))
    set_cell(ws, target, "DESCRIPCION TELA", fw(26))
    set_cell(ws, target, "ANCHO TELA", fw(27))
    set_cell(ws, target, "BASE TEXTIL", fw(28))
    set_cell(ws, target, "CODIGO TELA FORRO", fw(29))
    set_cell(ws, target, "DESCRIPCION FORRO", fw(31))
    set_cell(ws, target, "ANCHO FORRO", fw(32))
    
    set_cell(ws, target, "MOD ARTE SI/NO", fw(33))
    set_cell(ws, target, "UBI TRAZO SI/NO", fw(34))
    set_cell(ws, target, "ALL OVER SI/NO", fw(35))
    set_cell(ws, target, "VARIACION COLOR SI/NO", fw(37))
    set_cell(ws, target, "REF VARIACION COLOR", fw(38))
    set_cell(ws, target, "TIPO DE EMPAQUE", fw(39))
    
    set_cell(ws, target, "TIENE BORDADO?", fw(40))
    set_cell(ws, target, "DESCRIPCION BORDADO", fw(42))
    set_cell(ws, target, "TIENE SEMIELABORADO?", fw(43))
    set_cell(ws, target, "DESCRIPCION SEMIELABORADO", fw(44))
    set_cell(ws, target, "PROVEEDOR EXT.", fw(45))
    set_cell(ws, target, "PROCESO EXTERNO", fw(46))
    set_cell(ws, target, "COSTO PROCESO EXT.", fw(47))
    
    set_cell(ws, target, "COLECCION", "FALL WINTER (FW26)")
    
    # Unidades de produccion: cols after COMEX section (approx cols 137-150 for FW26)
    # Based on FW26 header row 4 (index 3), after "COMPOSICIONES..." there are 
    # 0,2,4,6,8,10,12,XS,S,M,L,XL,TOTAL at specific columns
    # Let me extract from the header row
    fw_headers = [str(h).strip() for h in fo_fw26_rows[fw_header_idx]]
    for t_idx, t_name in enumerate(['0', '2', '4', '6', '8', '10', '12', 'XS', 'S', 'M', 'L', 'XL']):
        t_col = None
        for j, h in enumerate(fw_headers):
            if h.strip() == t_name:
                t_col = j
                break
        if t_col is not None and t_col < len(row):
            val = safe(row[t_col])
            if val:
                # Map to target column by talla name
                talla_map = {'0': 'TALLA 0', '2': 'TALLA 2', '4': 'TALLA 4', '6': 'TALLA 6',
                             '8': 'TALLA 8', '10': 'TALLA 10', '12': 'TALLA 12',
                             'XS': 'TALLA XS', 'S': 'TALLA S', 'M': 'TALLA M',
                             'L': 'TALLA L', 'XL': 'TALLA XL'}
                set_cell(ws, target, talla_map.get(t_name, t_name), val)
    
    # TOTAL
    for j, h in enumerate(fw_headers):
        if h.strip() == 'TOTAL':
            val = safe(row[j])
            if val:
                set_cell(ws, target, "TOTAL UNIDADES", val)
            break
    
    fw_mapped += 1

print(f"FO FW26: {fw_mapped} referencias migradas")

# ═══════════════════════════════════════════════════════════
# 5. PARAMETROS POR TELA -> hoja PARAMETROS
# ═══════════════════════════════════════════════════════════
print("\n=== [5/5] PARAMETROS POR TELA -> PARAMETROS sheet ===")
csv_param = os.path.join(DOCS, "PARAMETROS POR TELA - BIK. BOTTOM.csv")

with open(csv_param, 'r', encoding='utf-8-sig') as f:
    param_rows = list(csv.reader(f))

print(f"PARAMETROS POR TELA: {len(param_rows)} lineas")

# ═══════════════════════════════════════════════════════════
# CREAR HOJA PARAMETROS
# ═══════════════════════════════════════════════════════════
print("\n=== CREANDO HOJA PARAMETROS ===")

# Delete existing PARAMETROS sheet if exists
if "PARAMETROS" in wb.sheetnames:
    del wb["PARAMETROS"]

ws_param = wb.create_sheet("PARAMETROS")

# Extract unique values from migrated data for dropdown lists
statuses = set()
disenadores_creativos = set()
disenadores_tecnicos = set()
trazadores = set()
modistas = set()
lineas = set()
sublineas = set()
tipos_ref = set()
tallajes = set()
usos_prenda = set()
bases_textiles = set()
largos = set()

for r in range(5, ws.max_row + 1):
    statuses.add(safe(ws.cell(row=r, column=find_col("STATUS") or 1).value))
    disenadores_creativos.add(safe(ws.cell(row=r, column=find_col("DISENADOR CREATIVO") or 1).value))
    disenadores_tecnicos.add(safe(ws.cell(row=r, column=find_col("DISENADOR TECNICO ASIGNADO") or 1).value))
    trazadores.add(safe(ws.cell(row=r, column=find_col("TRAZADOR ASIGNADO") or 1).value))
    modistas.add(safe(ws.cell(row=r, column=find_col("MODISTA ASIGNADA") or 1).value))
    lineas.add(safe(ws.cell(row=r, column=find_col("LINEA") or 1).value))
    sublineas.add(safe(ws.cell(row=r, column=find_col("SUBLINEA") or 1).value))
    tipos_ref.add(safe(ws.cell(row=r, column=find_col("TIPO REF") or 1).value))
    usos_prenda.add(safe(ws.cell(row=r, column=find_col("USO EN PRENDA") or 1).value))
    bases_textiles.add(safe(ws.cell(row=r, column=find_col("BASE TEXTIL") or 1).value))
    largos.add(safe(ws.cell(row=r, column=find_col("LARGO") or 1).value))

# Remove empty strings
for s in [statuses, disenadores_creativos, disenadores_tecnicos, trazadores,
          modistas, lineas, sublineas, tipos_ref, tallajes, usos_prenda,
          bases_textiles, largos]:
    s.discard('')
    s.discard('#N/A')
    s.discard('#REF!')

# Also add standard values
statuses.update(['EN PROCESO', 'APROBADO', 'CANCELADO', 'CANCELADO CORTADO',
                 'CANCELADO SIN CORTAR', 'JUST FOR SHOW', 'CANCELADO POR COMERCIAL',
                 'TERMINADO', 'PAUSADO', 'CANCELADO ES PERSONAL DE JO'])
tallajes.update(['0-2-4-6-8-10-12', 'XS-S-M-L-XL', '2-4-6-8-10-12', '1-2-3-4-5-6'])
usos_prenda.update(['TELA LUCIR', 'TELA FORRO', 'FUSIONABLE', 'SESGOS LUCIR',
                     'SESGOS FORRO', 'SESGOS FUSIONABLE', 'ENTRETELA', 'CINTA',
                     'RESORTE', 'HILO RESORTE', 'CIERRE', 'ARGOLIA', 'BOTON',
                     'CORDON', 'BORLA', 'MARQUILLA', 'ELASTICO'])
bases_textiles.update(['LINEN', 'COTTON', 'SILK', 'WOOL', 'LYCRA', 'LYCRA VITA',
                        'LYCRA BAHIA', 'LYCRA CRINKLE', 'CRINKLE LYCRA', 'SILK CDC',
                        'COTTON JACQUARD', 'PRINTED JACQUARD', 'LEATHER', 'CUERO',
                        'LIGHT LINEN', 'VISCOSE', 'POLYAMIDE', 'POLYESTER',
                        'TAFETA SEDA', 'DUCHESS SATIN', 'CREPE DE CHINE', 'CREPE DE SEDA'])

# Write PARAMETROS sheet
param_header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
param_header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
param_fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")

categorias = [
    ("STATUS", sorted(statuses, key=lambda x: x.lower())),
    ("DISEÑADORES CREATIVOS", sorted(disenadores_creativos, key=lambda x: x.lower())),
    ("DISEÑADORES TECNICOS", sorted(disenadores_tecnicos, key=lambda x: x.lower())),
    ("TRAZADORES", sorted(trazadores, key=lambda x: x.lower())),
    ("MODISTAS", sorted(modistas, key=lambda x: x.lower())),
    ("LINEAS", sorted(lineas, key=lambda x: x.lower())),
    ("SUBLINEAS", sorted(sublineas, key=lambda x: x.lower())),
    ("TIPOS REF", sorted(tipos_ref, key=lambda x: x.lower())),
    ("TALLAJES", sorted(tallajes, key=lambda x: x.lower())),
    ("USOS EN PRENDA", sorted(usos_prenda, key=lambda x: x.lower())),
    ("BASES TEXTILES", sorted(bases_textiles, key=lambda x: x.lower())),
    ("LARGOS", sorted(largos, key=lambda x: x.lower())),
    ("SI/NO", ["SI", "NO"]),
    ("VEREDICTOS", ["APROBADA DIRECTA", "APROBADA CON COMENTARIOS", "RECHAZADA"]),
    ("TIPO BORDADO", ["SOBRE PRENDA ARMADA", "SEMIELABORADO", "AMBOS", "EN PRENDA", "EMBROIDERED STRAP EDGE",
                       "EMBROIDERED AND EMBROIDERED STRAPS"]),
    ("TIPO PROCESO EXTERNO", ["LAVANDERIA", "BORDADO", "DRAPEADO", "DESCCOLADO", "TINTORERIA"]),
    ("CATALOGACION", ["SOLIDO", "MODIFICACION DE ARTE", "UBICACION EN TRAZO", "ALL OVER",
                       "ALL OVER CON SENTIDO", "ALL OVER CON ORIENTACION"]),
    ("COLECCIONES", ["WINTER SUN (WS27)", "FALL WINTER (FW26)", "RESORT RTW (RS26)", "SPRING SUMMER (SS26)",
                      "SUMMER VACATION (SV26)", "PREFALL RTW (PF26)"]),
    ("DROPS", ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]),
    ("PRIORIDADES", ["1","2","3","4","5","6","7","8","9","10"]),
    ("DIFICULTAD", ["BAJA", "INTERMEDIA", "ALTA", "MUY ALTA"]),
    ("TIPO TEJIDO", ["TEJIDO PLANO", "TEJIDO DE PUNTO", "CUERO", "DENIM"]),
    ("WOVEN/KNITTED", ["WOVEN", "KNITTED"]),
    ("TIPO CORTE", ["PRENDA COMPLETA", "LABORATORIO", "PIEZA", "REPOSICION"]),
]

col = 1
for cat_name, values in categorias:
    # Header
    cell = ws_param.cell(row=1, column=col, value=cat_name)
    cell.fill = param_header_fill
    cell.font = param_header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    
    # Values
    for j, val in enumerate(values):
        cell = ws_param.cell(row=2 + j, column=col, value=val)
        if j % 2 == 0:
            cell.fill = param_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=10)
    
    ws_param.column_dimensions[get_column_letter(col)].width = 28
    col += 1

# Agregar PARAMETROS POR TELA como otra seccion de la hoja
col += 1  # spacing
if param_rows:
    for i, row in enumerate(param_rows):
        for j, value in enumerate(row):
            cell = ws_param.cell(row=2 + i, column=col + j, value=safe(value) if value else '')
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=8)
            if i < 3:  # header rows
                cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                cell.font = Font(name="Calibri", size=8, bold=True)
    for j in range(len(param_rows[0]) if param_rows else 0):
        ws_param.column_dimensions[get_column_letter(col + j)].width = 12

ws_param.freeze_panes = "B2"
print(f"PARAMETROS: {len(categorias)} listas maestras creadas + datos de telas")

# ═══════════════════════════════════════════════════════════
# GUARDAR
# ═══════════════════════════════════════════════════════════
print("\n=== GUARDANDO prueba.xlsx ===")
wb.save(PRUEBA)
print(f"Archivo guardado: {PRUEBA}")
print(f"Total filas SECUENCIA REFERENCIA: {ws.max_row - 4}")
print("=== MIGRACION COMPLETADA ===")
