import gspread
import openpyxl

# ========= CONFIG =========
SHEET_ID = "1aRhZHtFwVmMg6y7MfC-x_3l68UpZClQubHaDqAqxOkY"
EXCEL_FILE = "prueba3.xlsx"
SHEET_NAME = "CONSUMO_MATERIALES"
CREDS_FILE = "google_credenciales.json"

# Mapeo de columnas en Excel (1-indexed: A=1, B=2, ..., AE=31)
COL = {
    "ID_Ref": 1,              # A
    "Cod_Material": 8,        # H (NO la B que es FOTO_Ref)
    "Tipo_Material": 6,       # F
    "Consumo_Creativo": 16,   # P
    "Ubicacion_Trazo": 12,    # L
    "Modificacion_Arte": 13,  # M
    "Consumo_Tecnico": 17,    # Q
    "Consumo_Verificacion": 19,  # S
    "A_Riesgo": 26,           # Z
    "Validado_Por": 27,       # AA
    "Consumo_Trazador": 18,   # R
    "Trazador": 28,           # AB
    "Observaciones": 31,      # AE
}

# ========= CONECTAR =========
gc = gspread.service_account(filename=CREDS_FILE)
sheet = gc.open_by_key(SHEET_ID)

# ========= LEER GOOGLE SHEETS =========
creativos = sheet.worksheet("CONSUMO_CREATIVOS").get_all_values()
tecnicos = sheet.worksheet("CONSUMO_TECNICOS").get_all_values()
trazadores = sheet.worksheet("CONSUMO_TRAZADORES").get_all_values()

# ========= ABRIR EXCEL =========
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]

# ========= FUNCION AUXILIAR: buscar o crear fila por ID_Ref + Cod_Material =========
def buscar_o_crear_fila(ws, id_ref, cod_mat):
    """Busca fila existente en Excel que coincida con ID_Ref + Cod_Material.
       Si no existe, crea una nueva al final y la retorna."""
    for row in range(2, ws.max_row + 1):  # fila 1 = headers
        if (str(ws.cell(row, COL["ID_Ref"]).value) == str(id_ref) and
            str(ws.cell(row, COL["Cod_Material"]).value) == str(cod_mat)):
            return row
    # No encontrada → crear nueva fila
    nueva = ws.max_row + 1
    ws.cell(nueva, COL["ID_Ref"], id_ref)
    ws.cell(nueva, COL["Cod_Material"], cod_mat)
    return nueva

# ========= ESCRIBIR: CREATIVOS (col 1=ID_Ref, 2=Cod_Material, 3=Tipo, 4=Consumo, 5=ModArte, 6=UbiTrazo, 7=Obs) =========
print("Sincronizando CREATIVOS...")
for fila in creativos[1:]:
    if not fila[0]:
        continue
    id_ref = fila[0]
    cod_mat = fila[1]
    row = buscar_o_crear_fila(ws, id_ref, cod_mat)
    ws.cell(row, COL["Tipo_Material"], fila[2])
    ws.cell(row, COL["Consumo_Creativo"], float(fila[3].replace(",", ".")) if fila[3] else None)
    ws.cell(row, COL["Modificacion_Arte"], fila[4])
    ws.cell(row, COL["Ubicacion_Trazo"], fila[5])
    if len(fila) > 6:
        ws.cell(row, COL["Observaciones"], fila[6])

# ========= ESCRIBIR: TECNICOS (col 1=ID_Ref, 2=Cod_Material, 3=Tipo, 4=Consumo, 5=A_Riesgo, 6=Validado_Por, 7=Obs) =========
print("Sincronizando TECNICOS...")
for fila in tecnicos[1:]:
    if not fila[0]:
        continue
    id_ref = fila[0]
    cod_mat = fila[1]
    row = buscar_o_crear_fila(ws, id_ref, cod_mat)
    ws.cell(row, COL["Consumo_Tecnico"], float(fila[3].replace(",", ".")) if fila[3] else None)
    ws.cell(row, COL["A_Riesgo"], fila[4])
    ws.cell(row, COL["Validado_Por"], fila[5])
    # Consumo_Verificacion (col 7 en google sheets si existe)
    if len(fila) > 7 and fila[7]:
        ws.cell(row, COL["Consumo_Verificacion"], float(fila[7].replace(",", ".")) if fila[7] else None)
    # Observaciones
    obs_idx = 8 if len(fila) > 8 else 7
    if len(fila) > obs_idx and fila[obs_idx]:
        obs_actual = ws.cell(row, COL["Observaciones"]).value or ""
        ws.cell(row, COL["Observaciones"], f"{obs_actual} | TECNICO: {fila[obs_idx]}" if obs_actual else fila[obs_idx])

# ========= ESCRIBIR: TRAZADORES (col 1=ID_Ref, 2=Cod_Material, 3=Tipo, 4=Consumo, 5=TallaBase, 6=Und, 7=Obs) =========
print("Sincronizando TRAZADORES...")
for fila in trazadores[1:]:
    if not fila[0]:
        continue
    id_ref = fila[0]
    cod_mat = fila[1]
    row = buscar_o_crear_fila(ws, id_ref, cod_mat)
    ws.cell(row, COL["Consumo_Trazador"], float(fila[3].replace(",", ".")) if fila[3] else None)
    ws.cell(row, COL["Trazador"], fila[6] if len(fila) > 6 else None)

# ========= GUARDAR =========
wb.save(EXCEL_FILE)
print("Sincronización completada")