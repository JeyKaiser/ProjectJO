"""
Genera VALIDACION_DE_TELAS_v3.0.xlsx
Arquitectura hibrida: hojas por rol + coordinacion completa
5 hojas: _MAESTRA(oculta), COORDINACION, CREATIVO, TECNICO, TRAZO_Y_CORTE
"""

import openpyxl
from openpyxl.styles import (Font, PatternFill, Border, Side, Alignment,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL   = PatternFill(start_color='46bdc6', end_color='46bdc6', fill_type='solid')
HEADER_FONT   = Font(name='Calibri', size=10, bold=True, color='ffffff')
DATA_FONT     = Font(name='Calibri', size=10, color='434343')
LOCKED_FILL   = PatternFill(start_color='f3f3f3', end_color='f3f3f3', fill_type='solid')
EDITABLE_FILL = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')

CREATIVO_FILL = PatternFill(start_color='b6d7a8', end_color='b6d7a8', fill_type='solid')
TECNICO_FILL  = PatternFill(start_color='fff2cc', end_color='fff2cc', fill_type='solid')
TRAZO_FILL    = PatternFill(start_color='fce5cd', end_color='fce5cd', fill_type='solid')
CONTRA_FILL   = PatternFill(start_color='d9d2e9', end_color='d9d2e9', fill_type='solid')
MATERIAL_FILL = PatternFill(start_color='c9daf8', end_color='c9daf8', fill_type='solid')
BASICO_FILL   = PatternFill(start_color='e6e6e6', end_color='e6e6e6', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='bbbbbb'),
    right=Side(style='thin', color='bbbbbb'),
    top=Side(style='thin', color='bbbbbb'),
    bottom=Side(style='thin', color='bbbbbb')
)
BOLD_BORDER = Border(
    left=Side(style='medium', color='666666'),
    right=Side(style='medium', color='666666'),
    top=Side(style='thin', color='bbbbbb'),
    bottom=Side(style='thin', color='bbbbbb')
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_headers(ws, row=1, max_col=10, fill=None):
    if fill is None:
        fill = HEADER_FILL
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = fill
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

def add_dropdown(ws, col_letter, options, min_row=2, max_row=200):
    dv = DataValidation(type='list', formula1='"' + ','.join(options) + '"', allow_blank=True)
    dv.error = 'Valor no valido'
    ws.add_data_validation(dv)
    dv.add(f'{col_letter}{min_row}:{col_letter}{max_row}')

def auto_col_width(ws, max_col=30, min_w=7, max_w=18):
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = min_w
    for col in range(1, max_col + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            for cell_val in row:
                if cell_val:
                    max_len = max(max_len, min(len(str(cell_val)), max_w))
        ws.column_dimensions[get_column_letter(col)].width = max(min_w, max_len + 2)

def apply_readonly_style(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = LOCKED_FILL
        c.border = THIN_BORDER
        c.font = Font(name='Calibri', size=10, color='888888')

def apply_editable_style(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = EDITABLE_FILL
        c.border = THIN_BORDER
        c.font = Font(name='Calibri', size=10, color='434343')

def create_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ═══════════════════════════════════════════════════════════════
    # Hoja 0: _MAESTRA (oculta)
    # ═══════════════════════════════════════════════════════════════
    ws_m = wb.create_sheet('_MAESTRA', 0)

    # Row 1: Fase headers
    fase_headers = {}
    fase_headers[0] = 'FASE 1 - INFORMACION BASICA'
    fase_headers[1] = 'FASE 2 - MATERIALES'
    fase_headers[2] = 'FASE 3 - DISENO CREATIVO'
    fase_headers[3] = 'FASE 3 - DISENO TECNICO'
    fase_headers[4] = 'FASE 3 - TRAZO Y CORTE'
    fase_headers[5] = 'FASE 4 - CONTRAMUESTRA'

    col_headers = [
        'REF','FOTO','REFERENCIA','STATUS','MOD.ARTE','UBI.TRAZO',
        'VAR.COLOR 1','VAR.COLOR 2','LARGO',
        'USO PRENDA','CODIGO TELA','DESC TELA','ANCHO','TELA FOTO','CONS BASE',
        'DISE CREATIVO','CAMBIO MOLDERIA','CONS 1','CONS 2','CONS 3','OBSERVACIONES',
        'DISE TECNICO',
        'TALLA SOL','UNDS SOL','CONS1 SOL','CONS2 SOL','CONS3 SOL',
        'TALLA MA','UNDS MA','CONS1 MA','CONS2 MA','CONS3 MA',
        'TALLA UT','UNDS UT','CONS1 UT','CONS2 UT','CONS3 UT','OBSERVACIONES TEC',
        'TALLA SOL TR','UNDS SOL TR','CONS1 SOL TR','CONS2 SOL TR','CONS3 SOL TR','CONS4 SOL TR',
        'TALLA MA TR','UNDS MA TR','CONS1 MA TR','CONS2 MA TR','CONS3 MA TR','CONS4 MA TR',
        'TALLA UT TR','UNDS UT TR','CONS1 UT TR','CONS2 UT TR','OBSERVACIONES TR',
        'TALLA SOL CM','UNDS SOL CM','CONS1 SOL CM','CONS2 SOL CM','CONS3 SOL CM','CONS4 SOL CM',
        'TALLA MA CM','UNDS MA CM','CONS1 MA CM','CONS2 MA CM','CONS3 MA CM','CONS4 MA CM',
        'TALLA UT CM','UNDS UT CM','CONS1 UT CM','CONS2 UT CM','OBSERVACIONES CM',
    ]

    total_cols = len(col_headers)  # 73

    # Row 1: Phase headers
    ws_m.cell(row=1, column=1, value='FASE 1 - INFORMACION BASICA')
    ws_m.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws_m.cell(row=1, column=10, value='FASE 2 - MATERIALES')
    ws_m.merge_cells(start_row=1, start_column=10, end_row=1, end_column=15)
    ws_m.cell(row=1, column=16, value='FASE 3 - DISEÑO CREATIVO')
    ws_m.merge_cells(start_row=1, start_column=16, end_row=1, end_column=21)
    ws_m.cell(row=1, column=22, value='FASE 3 - DISEÑO TECNICO')
    ws_m.merge_cells(start_row=1, start_column=22, end_row=1, end_column=38)
    ws_m.cell(row=1, column=39, value='FASE 3 - TRAZO Y CORTE')
    ws_m.merge_cells(start_row=1, start_column=39, end_row=1, end_column=55)
    ws_m.cell(row=1, column=56, value='FASE 4 - CONTRAMUESTRA')
    ws_m.merge_cells(start_row=1, start_column=56, end_row=1, end_column=72)

    # Row 2: Column headers
    for ci, h in enumerate(col_headers):
        c = ws_m.cell(row=2, column=ci + 1, value=h)
        c.font = HEADER_FONT
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER

    # Color phase headers
    for ci in range(1, 10):
        ws_m.cell(row=1, column=ci).fill = BASICO_FILL
        ws_m.cell(row=2, column=ci).fill = BASICO_FILL
    for ci in range(10, 16):
        ws_m.cell(row=1, column=ci).fill = MATERIAL_FILL
        ws_m.cell(row=2, column=ci).fill = MATERIAL_FILL
    for ci in range(16, 22):
        ws_m.cell(row=1, column=ci).fill = CREATIVO_FILL
        ws_m.cell(row=2, column=ci).fill = CREATIVO_FILL
    for ci in range(22, 39):
        ws_m.cell(row=1, column=ci).fill = TECNICO_FILL
        ws_m.cell(row=2, column=ci).fill = TECNICO_FILL
    for ci in range(39, 56):
        ws_m.cell(row=1, column=ci).fill = TRAZO_FILL
        ws_m.cell(row=2, column=ci).fill = TRAZO_FILL
    for ci in range(56, 73):
        ws_m.cell(row=1, column=ci).fill = CONTRA_FILL
        ws_m.cell(row=2, column=ci).fill = CONTRA_FILL

    ws_m.freeze_panes = 'A3'
    ws_m.sheet_state = 'hidden'
    auto_col_width(ws_m, total_cols, 6, 14)

    # ═══════════════════════════════════════════════════════════════
    # Hoja 1: COORDINACION
    # ═══════════════════════════════════════════════════════════════
    ws_c = wb.create_sheet('COORDINACION', 1)

    # Row 1: Phase headers (same as master)
    ws_c.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws_c.cell(row=1, column=1, value='FASE 1 - INFORMACION BASICA').fill = BASICO_FILL
    ws_c.merge_cells(start_row=1, start_column=10, end_row=1, end_column=15)
    ws_c.cell(row=1, column=10, value='FASE 2 - MATERIALES').fill = MATERIAL_FILL
    ws_c.merge_cells(start_row=1, start_column=16, end_row=1, end_column=21)
    ws_c.cell(row=1, column=16, value='CREATIVO').fill = CREATIVO_FILL
    ws_c.merge_cells(start_row=1, start_column=22, end_row=1, end_column=38)
    ws_c.cell(row=1, column=22, value='DISEÑO TECNICO').fill = TECNICO_FILL
    ws_c.merge_cells(start_row=1, start_column=39, end_row=1, end_column=55)
    ws_c.cell(row=1, column=39, value='TRAZO Y CORTE').fill = TRAZO_FILL
    ws_c.merge_cells(start_row=1, start_column=56, end_row=1, end_column=72)
    ws_c.cell(row=1, column=56, value='CONTRAMUESTRA').fill = CONTRA_FILL

    # Row 2: Column headers
    for ci, h in enumerate(col_headers):
        c = ws_c.cell(row=2, column=ci + 1, value=h)
        c.font = HEADER_FONT
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        if ci < 9: c.fill = BASICO_FILL
        elif ci < 15: c.fill = MATERIAL_FILL
        elif ci < 21: c.fill = CREATIVO_FILL
        elif ci < 38: c.fill = TECNICO_FILL
        elif ci < 55: c.fill = TRAZO_FILL
        else: c.fill = CONTRA_FILL

    ws_c.freeze_panes = 'D3'
    auto_col_width(ws_c, total_cols, 6, 14)

    # ═══════════════════════════════════════════════════════════════
    # Hoja 2: CREATIVO
    # ═══════════════════════════════════════════════════════════════
    ws_cr = wb.create_sheet('CREATIVO', 2)
    cr_headers = [
        ('REF', '🔒'), ('FOTO', '🔒'), ('REFERENCIA', '🔒'),
        ('STATUS', '🔒'), ('MOD.ARTE', '🔒'), ('UBI.TRAZO', '🔒'),
        ('USO PRENDA', '🔒'), ('CODIGO TELA', '🔒'), ('DESC TELA', '🔒'),
        ('ANCHO', '🔒'), ('TELA FOTO', '🔒'), ('CONS BASE', '🔒'),
        ('DISEÑADOR CREATIVO', '✏️'), ('CAMBIO MOLDERIA', '✏️'),
        ('CONS 1', '✏️'), ('CONS 2', '✏️'), ('CONS 3', '✏️'),
        ('OBSERVACIONES', '✏️'),
    ]

    # Row 1: role bar
    ws_cr.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws_cr.cell(row=1, column=1, value='CONTEXTO (solo lectura)').fill = BASICO_FILL
    ws_cr.cell(row=1, column=1).font = Font(name='Calibri', size=9, bold=True, color='555555')
    ws_cr.merge_cells(start_row=1, start_column=7, end_row=1, end_column=12)
    ws_cr.cell(row=1, column=7, value='MATERIALES (solo lectura)').fill = MATERIAL_FILL
    ws_cr.cell(row=1, column=7).font = Font(name='Calibri', size=9, bold=True, color='555555')
    ws_cr.merge_cells(start_row=1, start_column=13, end_row=1, end_column=18)
    ws_cr.cell(row=1, column=13, value='MIS CONSUMOS (editable)').fill = CREATIVO_FILL
    ws_cr.cell(row=1, column=13).font = Font(name='Calibri', size=9, bold=True, color='555555')

    # Row 2: column headers
    for ci, (h, _) in enumerate(cr_headers):
        c = ws_cr.cell(row=2, column=ci + 1, value=h)
        c.font = HEADER_FONT
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        if ci < 6: c.fill = BASICO_FILL
        elif ci < 12: c.fill = MATERIAL_FILL
        else: c.fill = CREATIVO_FILL

    ws_cr.freeze_panes = 'A3'
    auto_col_width(ws_cr, len(cr_headers), 7, 18)

    # ═══════════════════════════════════════════════════════════════
    # Hoja 3: TECNICO
    # ═══════════════════════════════════════════════════════════════
    ws_t = wb.create_sheet('TECNICO', 3)
    tech_headers = [
        ('REF','🔒'),('FOTO','🔒'),('REFERENCIA','🔒'),
        ('STATUS','🔒'),('MOD.ARTE','🔒'),('UBI.TRAZO','🔒'),
        ('USO PRENDA','🔒'),('CODIGO TELA','🔒'),('DESC TELA','🔒'),
        ('ANCHO','🔒'),('TELA FOTO','🔒'),('CONS BASE','🔒'),
        ('DISEÑADOR TECNICO','✏️'),
        ('TALLA SOL','✏️'),('UNDS SOL','✏️'),('CONS1 SOL','✏️'),('CONS2 SOL','✏️'),('CONS3 SOL','✏️'),
        ('TALLA MA','✏️'),('UNDS MA','✏️'),('CONS1 MA','✏️'),('CONS2 MA','✏️'),('CONS3 MA','✏️'),
        ('TALLA UT','✏️'),('UNDS UT','✏️'),('CONS1 UT','✏️'),('CONS2 UT','✏️'),('CONS3 UT','✏️'),
        ('OBSERVACIONES','✏️'),
    ]

    ws_t.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws_t.cell(row=1, column=1, value='CONTEXTO 🔒').fill = BASICO_FILL
    ws_t.merge_cells(start_row=1, start_column=7, end_row=1, end_column=12)
    ws_t.cell(row=1, column=7, value='MATERIALES 🔒').fill = MATERIAL_FILL
    ws_t.merge_cells(start_row=1, start_column=13, end_row=1, end_column=13)
    ws_t.cell(row=1, column=13, value='DIS. TEC ✏️').fill = TECNICO_FILL
    ws_t.merge_cells(start_row=1, start_column=14, end_row=1, end_column=18)
    ws_t.cell(row=1, column=14, value='SOLIDO ✏️').fill = TECNICO_FILL
    ws_t.merge_cells(start_row=1, start_column=19, end_row=1, end_column=23)
    ws_t.cell(row=1, column=19, value='MOD ARTE ✏️').fill = TECNICO_FILL
    ws_t.merge_cells(start_row=1, start_column=24, end_row=1, end_column=28)
    ws_t.cell(row=1, column=24, value='UBI TRAZO ✏️').fill = TECNICO_FILL
    ws_t.cell(row=1, column=29, value='OBS ✏️').fill = TECNICO_FILL

    for ci, (h, _) in enumerate(tech_headers):
        c = ws_t.cell(row=2, column=ci + 1, value=h)
        c.font = HEADER_FONT
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        if ci < 6: c.fill = BASICO_FILL
        elif ci < 12: c.fill = MATERIAL_FILL
        else: c.fill = TECNICO_FILL

    ws_t.freeze_panes = 'A3'
    auto_col_width(ws_t, len(tech_headers), 7, 14)

    # ═══════════════════════════════════════════════════════════════
    # Hoja 4: TRAZO_Y_CORTE
    # ═══════════════════════════════════════════════════════════════
    ws_tr = wb.create_sheet('TRAZO_Y_CORTE', 4)
    trazo_headers = [
        ('REF','🔒'),('FOTO','🔒'),('REFERENCIA','🔒'),
        ('STATUS','🔒'),('MOD.ARTE','🔒'),('UBI.TRAZO','🔒'),
        ('USO PRENDA','🔒'),('CODIGO TELA','🔒'),('DESC TELA','🔒'),
        ('ANCHO','🔒'),('TELA FOTO','🔒'),('CONS BASE','🔒'),
        ('TALLA SOL','✏️'),('UNDS SOL','✏️'),('CONS1 SOL','✏️'),('CONS2 SOL','✏️'),('CONS3 SOL','✏️'),('CONS4 SOL','✏️'),
        ('TALLA MA','✏️'),('UNDS MA','✏️'),('CONS1 MA','✏️'),('CONS2 MA','✏️'),('CONS3 MA','✏️'),('CONS4 MA','✏️'),
        ('TALLA UT','✏️'),('UNDS UT','✏️'),('CONS1 UT','✏️'),('CONS2 UT','✏️'),
        ('OBSERVACIONES','✏️'),
    ]

    ws_tr.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws_tr.cell(row=1, column=1, value='CONTEXTO 🔒').fill = BASICO_FILL
    ws_tr.merge_cells(start_row=1, start_column=7, end_row=1, end_column=12)
    ws_tr.cell(row=1, column=7, value='MATERIALES 🔒').fill = MATERIAL_FILL
    ws_tr.merge_cells(start_row=1, start_column=13, end_row=1, end_column=18)
    ws_tr.cell(row=1, column=13, value='SOLIDO ✏️').fill = TRAZO_FILL
    ws_tr.merge_cells(start_row=1, start_column=19, end_row=1, end_column=24)
    ws_tr.cell(row=1, column=19, value='MOD ARTE ✏️').fill = TRAZO_FILL
    ws_tr.merge_cells(start_row=1, start_column=25, end_row=1, end_column=28)
    ws_tr.cell(row=1, column=25, value='UBI TRAZO ✏️').fill = TRAZO_FILL
    ws_tr.cell(row=1, column=29, value='OBS ✏️').fill = TRAZO_FILL

    for ci, (h, _) in enumerate(trazo_headers):
        c = ws_tr.cell(row=2, column=ci + 1, value=h)
        c.font = HEADER_FONT
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        if ci < 6: c.fill = BASICO_FILL
        elif ci < 12: c.fill = MATERIAL_FILL
        else: c.fill = TRAZO_FILL

    ws_tr.freeze_panes = 'A3'
    auto_col_width(ws_tr, len(trazo_headers), 7, 14)

    # ── Save ──
    path = 'Documentos/NUEVA_PROPUESTA/VALIDACION_DE_TELAS_v3.0.xlsx'
    wb.save(path)
    print(f'Archivo generado: {path}')
    print(f'Hojas: {wb.sheetnames}')
    for name in wb.sheetnames:
        ws = wb[name]
        print(f'  {name}: {ws.max_row}r x {ws.max_column}c')

if __name__ == '__main__':
    create_workbook()
