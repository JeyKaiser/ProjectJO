"""Completa prueba.xlsx: crea hoja PARAMETROS con listas maestras."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv, os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

PRUEBA = r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0\prueba.xlsx"
DOCS = r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0\Documentos"

wb = openpyxl.load_workbook(PRUEBA)
ws = wb["SECUENCIA REFERENCIA"]

thin_border = Border(
    left=Side(style='thin', color='808080'),
    right=Side(style='thin', color='808080'),
    top=Side(style='thin', color='808080'),
    bottom=Side(style='thin', color='808080')
)

def safe(val):
    if val is None or str(val).strip() in ('', '#N/A', 'N/A', '#REF!'):
        return ''
    return str(val).strip()

def find_col(keyword):
    for col in range(1, ws.max_column + 1):
        name = ws.cell(row=3, column=col).value
        if name and keyword.upper() in str(name).upper():
            return col
    return None

# ── Extraer valores unicos de los datos migrados ──
print("Extrayendo valores unicos de SECUENCIA REFERENCIA...")

def extract_set(col_keyword):
    col = find_col(col_keyword)
    result = set()
    if col:
        for r in range(5, ws.max_row + 1):
            val = safe(ws.cell(row=r, column=col).value)
            if val:
                result.add(val)
    return result

statuses = extract_set("STATUS")
disenadores_c = extract_set("DISENADOR CREATIVO")
disenadores_t = extract_set("DISENADOR TECNICO ASIGNADO")
trazadores = extract_set("TRAZADOR ASIGNADO")
modistas = extract_set("MODISTA ASIGNADA")
lineas = extract_set("LINEA")
sublineas = extract_set("SUBLINEA")
tipos_ref = extract_set("TIPO REF")
tallajes = extract_set("TALLAJE")
largos = extract_set("LARGO")
usos_prenda = extract_set("USO EN PRENDA")
bases_textiles = extract_set("BASE TEXTIL")

print(f"  Status: {len(statuses)} | Creativos: {len(disenadores_c)} | Tecnicos: {len(disenadores_t)}")
print(f"  Trazadores: {len(trazadores)} | Modistas: {len(modistas)} | Lineas: {len(lineas)}")
print(f"  Usos: {len(usos_prenda)} | Bases: {len(bases_textiles)} | Largos: {len(largos)}")

# ── Suplementar con valores estandar ──
statuses.update(['EN PROCESO', 'APROBADO', 'CANCELADO', 'CANCELADO CORTADO',
                 'CANCELADO SIN CORTAR', 'JUST FOR SHOW', 'CANCELADO POR COMERCIAL',
                 'TERMINADO', 'PAUSADO', 'CANCELADO ES PERSONAL DE JO'])
tallajes.update(['0-2-4-6-8-10-12', 'XS-S-M-L-XL', '2-4-6-8-10-12'])
usos_prenda.update(['TELA LUCIR', 'TELA FORRO', 'FUSIONABLE', 'SESGOS LUCIR',
                     'SESGOS FORRO', 'SESGOS FUSIONABLE', 'ENTRETELA', 'CINTA',
                     'RESORTE', 'HILO RESORTE', 'CIERRE', 'ARGOLIA', 'BOTON'])
bases_textiles.update(['LINEN', 'COTTON', 'SILK', 'WOOL', 'LYCRA', 'LYCRA VITA',
                        'LYCRA BAHIA', 'LYCRA CRINKLE', 'CRINKLE LYCRA', 'SILK CDC',
                        'COTTON JACQUARD', 'PRINTED JACQUARD', 'LEATHER', 'CUERO',
                        'LIGHT LINEN', 'VISCOSE', 'POLYAMIDE', 'POLYESTER',
                        'TAFETA SEDA', 'DUCHESS SATIN', 'CREPE DE CHINE', 'CREPE DE SEDA'])

# ── Crear hoja PARAMETROS ──
if "PARAMETROS" in wb.sheetnames:
    del wb["PARAMETROS"]
ws_param = wb.create_sheet("PARAMETROS")

header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
alt_fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
data_align = Alignment(horizontal="center", vertical="center")

def sorted_set(s):
    return sorted(s, key=lambda x: x.lower().strip())

categorias = [
    ("STATUS", sorted_set(statuses)),
    ("DISE\u00d1ADORES CREATIVOS", sorted_set(disenadores_c)),
    ("DISE\u00d1ADORES T\u00c9CNICOS", sorted_set(disenadores_t)),
    ("TRAZADORES", sorted_set(trazadores)),
    ("MODISTAS", sorted_set(modistas)),
    ("L\u00cdNEAS", sorted_set(lineas)),
    ("SUBL\u00cdNEAS", sorted_set(sublineas)),
    ("TIPOS REF", sorted_set(tipos_ref)),
    ("TALLAJES", sorted_set(tallajes)),
    ("LARGOS", sorted_set(largos)),
    ("USOS EN PRENDA", sorted_set(usos_prenda)),
    ("BASES TEXTILES", sorted_set(bases_textiles)),
    ("SI/NO", ["SI", "NO"]),
    ("VEREDICTOS DIR. CREATIVA", ["APROBADA DIRECTA", "APROBADA CON COMENTARIOS", "RECHAZADA"]),
    ("TIPO BORDADO", ["SOBRE PRENDA ARMADA", "SEMIELABORADO", "AMBOS", "EN PRENDA",
                       "EMBROIDERED STRAP EDGE", "EMBROIDERED AND EMBROIDERED STRAPS"]),
    ("TIPO PROCESO EXTERNO", ["LAVANDERIA", "BORDADO", "DRAPEADO", "DESCOLADO", "TINTORERIA", "MONTAJE MANIQUI"]),
    ("CATALOGACION", ["SOLIDO", "MODIFICACION DE ARTE", "UBICACION EN TRAZO", "ALL OVER",
                       "ALL OVER CON SENTIDO", "ALL OVER CON ORIENTACION",
                       "REPROGRAMACION MOD ARTE", "APOYO ESPECIAL"]),
    ("COLECCIONES", ["WINTER SUN (WS27)", "FALL WINTER (FW26)", "RESORT RTW (RS26)",
                      "SPRING SUMMER (SS26)", "SUMMER VACATION (SV26)", "PREFALL RTW (PF26)"]),
    ("DROPS", [chr(c) for c in range(65, 91)]),  # A-Z
    ("PRIORIDADES", [str(i) for i in range(1, 11)]),
    ("DIFICULTAD", ["BAJA", "INTERMEDIA", "ALTA", "MUY ALTA"]),
    ("TIPO TEJIDO", ["TEJIDO PLANO", "TEJIDO DE PUNTO", "CUERO", "DENIM", "GASA"]),
    ("WOVEN/KNITTED", ["WOVEN", "KNITTED"]),
    ("TIPO CORTE", ["PRENDA COMPLETA", "LABORATORIO", "PIEZA", "REPOSICION", "REPOSICION CONTRAMUESTRA"]),
    ("STATUS PROCESO EXT.", ["PENDIENTE", "EN PROCESO", "TERMINADO", "RECHAZADO"]),
    ("CONFECCION STATUS", ["EN PROCESO", "TERMINADO", "PAUSADO", "CANCELADO", "RECHAZADO"]),
    ("TIPO RECHAZO CALIDAD", ["MEDIDAS", "CALIDAD DE CONFECCION", "COHERENCIA CON FICHA TECNICA",
                                "MATERIA PRIMA", "INSUMOS", "OTRO"]),
    ("CLASIFICACION HALLAZGO", ["INCONSISTENCIAS", "FALTA DE INFORMACION", "FALTANTES",
                                  "CAMBIOS", "FALTA DE ANALISIS", "OTROS"]),
    ("TIPO EMPAQUE", ["CAJA INDIVIDUAL", "BOLSA COLECTIVA", "CAJA COLECTIVA", "BOLSA INDIVIDUAL",
                       "MARQUILLA SATINADA FORRADA", "SIN EMPAQUE ESPECIAL"]),
    ("CLOSURE TYPES", ["BACK ZIPPER", "BACK BUTTONS", "BACK ZIPPER AND BUTTOM", "SNAPS",
                        "WITH SASH", "N/A", "SIN CIERRE"]),
    ("LINNED", ["SI", "NO", "ONLY THE TOP", "ONLY THE BOTTOM", "N/A"]),
    ("SENTIDO SESGO", ["AL HILO", "A TRAVEZ", "AL SESGO"]),
    ("PERMITE GIRAR MOLDE", ["90", "180", "PEINE"]),
]

print(f"\nCreando {len(categorias)} listas maestras...")

col = 1
for cat_name, values in categorias:
    # Header
    cell = ws_param.cell(row=1, column=col, value=cat_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    
    # Values
    for j, val in enumerate(values):
        cell = ws_param.cell(row=2 + j, column=col, value=val)
        if j % 2 == 0:
            cell.fill = alt_fill
        cell.alignment = data_align
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=9)
    
    ws_param.column_dimensions[get_column_letter(col)].width = 30
    col += 1

# ── Cargar PARAMETROS POR TELA data ──
print("Agregando datos de PARAMETROS POR TELA...")
csv_param = os.path.join(DOCS, "PARAMETROS POR TELA - BIK. BOTTOM.csv")
with open(csv_param, 'r', encoding='utf-8-sig') as f:
    param_rows = list(csv.reader(f))

if param_rows:
    start_col = col + 1
    telas_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    telas_font = Font(color="000000", bold=True, size=8, name="Calibri")
    label = ws_param.cell(row=1, column=start_col, value="PARAMETROS POR TELA - BIKINI BOTTOM")
    label.fill = telas_fill
    label.font = Font(color="000000", bold=True, size=10, name="Calibri")
    for c in range(start_col, start_col + len(param_rows[0])):
        ws_param.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + len(param_rows[0]) - 1)
        break
    
    for i, row in enumerate(param_rows):
        for j, value in enumerate(row):
            cell = ws_param.cell(row=2 + i, column=start_col + j, value=safe(value))
            cell.alignment = data_align
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=8)
            if i == 0:  # first row styling
                cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
                cell.font = Font(name="Calibri", size=8, bold=True)
    for j in range(len(param_rows[0])):
        ws_param.column_dimensions[get_column_letter(start_col + j)].width = 14
    
    print(f"  Datos: {len(param_rows)} filas x {len(param_rows[0])} columnas")

ws_param.freeze_panes = "B2"
wb.save(PRUEBA)

# ── Verificar ──
print(f"\nArchivo guardado: {PRUEBA}")
print(f"Hoja SECUENCIA REFERENCIA: {ws.max_row - 4} filas de datos")
print(f"Hoja PARAMETROS: {ws_param.max_row - 1} filas x {ws_param.max_column} columnas")
print("=== COMPLETADO ===")
