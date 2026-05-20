"""
Script para crear el archivo 'prueba.xlsx' con la secuencia consolidada
del ciclo de vida de una referencia de coleccion JO.
Fases: Ideacion (Azul), Laboratorio (Ambar), Validacion (Naranja), Industrializacion (Rojo)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SECUENCIA REFERENCIA"

# ── Colores por Fase ──
class Colors:
    FASE1 = "4A90D9"    # Azul Claro - Ideacion y Diseno
    FASE1_BG = "A3C9E8" # Fondo azul claro
    FASE2 = "E8A838"    # Ambar - Laboratorio y Prototipado
    FASE2_BG = "F4D389" # Fondo amarillo
    FASE3 = "E07B3E"    # Naranja - Validacion Tecnica
    FASE3_BG = "F0B27A" # Fondo naranja claro
    FASE4 = "D94A4A"    # Rojo - Industrializacion
    FASE4_BG = "F1948A" # Fondo rojo claro
    WHITE = "FFFFFF"
    HEADER_FONT = "000000"
    BORDER = "808080"

thin_border = Border(
    left=Side(style='thin', color=Colors.BORDER),
    right=Side(style='thin', color=Colors.BORDER),
    top=Side(style='thin', color=Colors.BORDER),
    bottom=Side(style='thin', color=Colors.BORDER)
)

def style_header(ws, row, col_start, col_end, fill_color, font_color="FFFFFF", text="", bold=True, size=11):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.font = Font(color=font_color, bold=bold, size=size, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    if text:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        ws.cell(row=row, column=col_start).value = text

def set_col_widths(ws, widths_dict):
    for col_num, width in widths_dict.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

# ═══════════════════════════════════════════════════════════════
# DEFINICION DE LA SECUENCIA (ROW 3: columnas individuales)
# ═══════════════════════════════════════════════════════════════

# Cada bloque: (texto_fase, texto_subarea, [columnas], color)
# Fila 1 = Fase macro (merged across all columns of that phase)
# Fila 2 = Sub-area header (merged within sub-area)
# Fila 3 = Columnas individuales

COL = 1  # contador de columna

def add_phase(phase_name, color, subareas):
    """
    subareas: list of (subarea_name, [column_names], sub_color or None)
    Each subarea gets row2 merged header, row3 individual cols.
    """
    global COL
    start = COL
    for sub_name, columns, sub_color in subareas:
        sub_start = COL
        for col_name in columns:
            ws.cell(row=3, column=COL).value = col_name
            cell = ws.cell(row=3, column=COL)
            cell.fill = PatternFill(start_color=sub_color or color, end_color=sub_color or color, fill_type="solid")
            cell.font = Font(bold=True, size=9, name="Calibri", color="000000" if sub_color else "FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            COL += 1
        end = COL - 1
        if sub_name:
            style_header(ws, 2, sub_start, end, color, "000000", sub_name, bold=True, size=10)
    end = COL - 1
    style_header(ws, 1, start, end, color, "FFFFFF", phase_name, bold=True, size=12)

# ── FASE 1: IDEACION Y DISENO (Azul) ──
add_phase(
    "🔵 FASE 1: IDEACION Y DISENO",
    Colors.FASE1,
    [
        ("1.1 PERFILAMIENTO - Coord. Diseno", [
            "REF", "CODIGO MD", "CODIGO PT", "NOMBRE REFERENCIA", "COLOR",
            "COD COLOR", "LINEA", "SUBLINEA", "TIPO REF", "COLECCION",
            "TEMPORADA", "CAPSULA", "TIPO PRENDA", "ES NUEVO?"
        ], Colors.FASE1_BG),
        ("1.2 REFERENTES", [
            "FOTO BASADO EN", "PT REFERENTE", "NOMBRE REFERENTE",
            "COLECCION REFERENTE"
        ], Colors.FASE1_BG),
        ("1.3 CARACTERISTICAS PRENDA", [
            "TALLAJE", "LARGO", "CLOSURE", "LINNED",
            "INCLUDES", "INCLUDES PAQ. COMPLETO PLANTA"
        ], Colors.FASE1_BG),
        ("1.4 DISENADOR CREATIVO", [
            "DISENADOR CREATIVO", "FECHA ASIGNACION CREATIVO"
        ], Colors.FASE1_BG),
        ("1.5 CONSUMO BASE (Estimacion Visual)", [
            "CONSUMO BASE", "CAMBIO MOLDERIA SI/NO",
            "CONSUMO CREATIVO 1", "CONSUMO CREATIVO 2",
            "CONSUMO CREATIVO 3", "OBSERVACIONES CREATIVO"
        ], Colors.FASE1_BG),
        ("1.6 MOLDERIA BASE (Patron Inicial)", [
            "FECHA INICIO MOLDERIA DISENO", "FECHA FIN MOLDERIA DISENO",
            "COMENTARIOS MOLDERIA DISENO", "REQUIERE MUESTRA SI/NO",
            "FOTOS INTERNAS"
        ], Colors.FASE1_BG),
    ]
)

# ── FASE 2: LABORATORIO Y PROTOTIPADO (Ambar) ──
add_phase(
    "🟡 FASE 2: LABORATORIO Y PROTOTIPADO",
    Colors.FASE2,
    [
        ("2.1 ALISTAMIENTO (Telas e Insumos)", [
            "USO EN PRENDA", "CODIGO TELA/INSUMO", "DESCRIPCION TELA",
            "ANCHO TELA", "BASE TEXTIL", "FOTO TELA", "COLOR TELA",
            "NOVEDAD CALIDAD SI/NO", "TIPO NOVEDAD", "ESPERANDO COMPRAS?"
        ], Colors.FASE2_BG),
        ("2.2 TELA FORRO / FUSIONABLE", [
            "CODIGO TELA FORRO", "DESCRIPCION FORRO", "ANCHO FORRO",
            "CODIGO FUSIONABLE", "DESCRIPCION FUSIONABLE", "ANCHO FUSIONABLE"
        ], Colors.FASE2_BG),
        ("2.3 CORTE DE MUESTRA", [
            "CORTADOR ASIGNADO", "FECHA CORTE MUESTRA",
            "CONSUMO TELA MUESTRA (mts)", "OBSERVACIONES CORTE"
        ], Colors.FASE2_BG),
        ("2.4 CONFECCION DE MUESTRA", [
            "MODISTA ASIGNADA", "FECHA INICIO CONFECCION", "FECHA ENTREGA CONFECCION",
            "STATUS CONFECCION", "OBSERVACIONES MODISTA",
            "ESPECIFICACION CONFECCION #1"
        ], Colors.FASE2_BG),
        ("2.5 PROCESOS ESPECIALES", [
            "TIENE BORDADO?", "TIPO BORDADO", "DESCRIPCION BORDADO",
            "TIENE SEMIELABORADO?", "DESCRIPCION SEMIELABORADO",
            "PROVEEDOR EXT.", "PROCESO EXTERNO", "COSTO PROCESO EXT.",
            "STATUS PROCESO EXT."
        ], Colors.FASE2_BG),
    ]
)

# ── FASE 3: VALIDACION TECNICA (Naranja) ──
add_phase(
    "🟠 FASE 3: VALIDACION TECNICA",
    Colors.FASE3,
    [
        ("3.1 MEDICION Y TALLAJE", [
            "MEDICION FASE 1 - FECHA", "COMENTARIOS MEDICION 1",
            "VEREDICTO DIR. CREATIVA", "MEDICION FASE 2 - FECHA",
            "COMENTARIOS MEDICION 2", "MEDICION FASE 3 - FECHA",
            "COMENTARIOS MEDICION 3", "MEDICION FASE 4 - FECHA",
            "COMENTARIOS MEDICION 4", "MEDICION FASE 5 - FECHA",
            "COMENTARIOS MEDICION 5", "FOTO CONTRAMUESTRA"
        ], Colors.FASE3_BG),
        ("3.2 AJUSTES DE MOLDERIA", [
            "REQUIERE AJUSTE SI/NO", "TIPO AJUSTE",
            "COMENTARIOS AJUSTE", "FECHA INICIO MOLDERIA ING",
            "FECHA FINALIZACION MOLDERIA ING", "COMENTARIOS HALLAZGOS MOLDERIA"
        ], Colors.FASE3_BG),
        ("3.3 DISENADOR TECNICO (Escalado)", [
            "DISENADOR TECNICO ASIGNADO", "DISENADOR CONTRAMUESTRA",
            "TIPO CONSUMO (SOLIDO/MOD/UBI)", "COMENTARIOS INGENIERIA"
        ], Colors.FASE3_BG),
        ("3.4 CALCULO CONSUMO TECNICO", [
            "CONSUMO SOLIDO", "CONSUMO MOD ARTE", "CONSUMO UBI TRAZO",
            "TALLA CALCULO CONSUMO", "UNIDAD CALCULO",
            "OBSERVACIONES TECNICO"
        ], Colors.FASE3_BG),
        ("3.5 TRAZOS DE PRODUCCION (Trazador)", [
            "TRAZADOR ASIGNADO", "CONSUMO TRAZADOR SOLIDO",
            "CONSUMO TRAZADOR MOD ARTE", "CONSUMO TRAZADOR UBI TRAZO",
            "AHORRO TRAZADOR (mts)", "% AHORRO TRAZADOR",
            "OBSERVACIONES TRAZO", "COMENTARIOS TRAZO"
        ], Colors.FASE3_BG),
        ("3.6 CATALOGACION ESPECIAL", [
            "MOD ARTE SI/NO", "UBI TRAZO SI/NO", "ALL OVER SI/NO",
            "VARIACION COLOR SI/NO", "REF VARIACION COLOR",
            "SUGERENCIA MOD/UBC"
        ], Colors.FASE3_BG),
    ]
)

# ── FASE 4: INDUSTRIALIZACION (Rojo) ──
add_phase(
    "🔴 FASE 4: INDUSTRIALIZACION",
    Colors.FASE4,
    [
        ("4.1 CIERRE FICHA FINAL", [
            "ESPECIFICADORA", "FECHA INICIO ESPECIFICACION",
            "FECHA FINAL ESPECIFICACION", "REVISION MATERIALES SAP",
            "FECHA ENTREGA FICHA BORDADO", "FECHA ENTREGA A FICHA"
        ], Colors.FASE4_BG),
        ("4.2 MARQUILLAS Y COMPOSICIONES (Muestra)", [
            "DESC USA/UK (Muestra)", "FIBER COMPOSITION (Muestra)",
            "WOVEN/KNITTED (Muestra)", "INSIDE (Muestra)", "INCLUDE (Muestra)",
            "OBSERVACIONES COMPOSICION (Muestra)", "INFO SAP COMPOSICION"
        ], Colors.FASE4_BG),
        ("4.3 MARQUILLAS Y COMPOSICIONES (Produccion)", [
            "DESC USA/UK (Prod)", "FIBER COMPOSITION (Prod)",
            "WOVEN/KNITTED (Prod)", "INSIDE (Prod)", "INCLUDE (Prod)",
            "OBSERVACIONES COMPOSICION (Prod)", "COMEX"
        ], Colors.FASE4_BG),
        ("4.4 CUIDADOS DE LA PRENDA", [
            "PROCESO ESPECIAL CUIDADOS", "LAVADO", "DESMANCHE",
            "SECADO", "PLANCHADO", "CUIDADOS INCLUDES PRENDA",
            "LOGO LAVADO", "LOGO DESMANCHE", "LOGO SECADO", "LOGO PLANCHA"
        ], Colors.FASE4_BG),
        ("4.5 EXPLOSION CONTRAMUESTRA", [
            "NOMBRE CONTRAMUESTRA (OT)", "CODIGO OT", "TALLA CONTRAMUESTRA",
            "DESCRIPCION COLOR CONTRAMUESTRA", "UNIDADES CORTADAS",
            "FOTO CONTRAMUESTRA", "REPROGRAMACION CONTRAMUESTRA SI/NO",
            "PRIORIDAD CONTRAMUESTRA", "FECHA META ENTREGA",
            "DROPS", "STATUS CONTRAMUESTRA"
        ], Colors.FASE4_BG),
        ("4.6 NOTA FABRICACION SAP", [
            "NOTA DE FABRICACION", "GESTION DE LA NOTA (FECHA)",
            "FECHA TRASLADO SAP", "FECHA DESPACHO ZF (PRODUCCION)",
            "CODIGO DE BARRAS", "UBICACION"
        ], Colors.FASE4_BG),
        ("4.7 UNIDADES DE PRODUCCION", [
            "TALLA 0", "TALLA 2", "TALLA 4", "TALLA 6", "TALLA 8",
            "TALLA 10", "TALLA 12", "TALLA XS", "TALLA S", "TALLA M",
            "TALLA L", "TALLA XL", "TOTAL UNIDADES"
        ], Colors.FASE4_BG),
        ("4.8 MAQUILA", [
            "TIPO DE TEJIDO", "COMPLEJIDAD CORTE", "ENVIO CORTE A MAQUILAS",
            "COMPLEJIDAD CONFECCION", "ENVIO CONFECCION A MAQUILAS"
        ], Colors.FASE4_BG),
        ("4.9 MONTAJE EN MANIQUI", [
            "TIPO DE MONTAJE MANIQUI", "PROYECTO MONTAJE MANIQUI",
        ], Colors.FASE4_BG),
        ("4.10 VALIDACION MP (Novedades Prod.)", [
            "FECHA VALIDACION MP", "AREA DE AFECTACION",
            "CLASIFICACION HALLAZGO", "MP INVOLUCRADA",
            "CLASIFICACION MP", "TIPO DE EJECUCION"
        ], Colors.FASE4_BG),
        ("4.11 FEEDBACK PRODUCCION (Escalado)", [
            "LIBERACION FICHA FISICA", "HALLAZGO REPROGRAMACION",
            "CLASIFICACION SITUACIONES PROD", "TIPO SITUACION PLANTA",
            "OBSERVACIONES ESCALADO"
        ], Colors.FASE4_BG),
        ("4.12 FEEDBACK: Ficha Especificacion", [
            "CLASIFICACION SITUACIONES PROD (Ficha)", "VARIABLES (Ficha)",
            "OBSERVACIONES (Ficha)"
        ], Colors.FASE4_BG),
        ("4.13 FEEDBACK: Bordado", [
            "CLASIFICACION SITUACIONES PROD (Bordado)", "VARIABLES (Bordado)",
            "OBSERVACIONES (Bordado)"
        ], Colors.FASE4_BG),
        ("4.14 FEEDBACK: Consumos", [
            "CLASIFICACION SITUACIONES PROD (Consumos)", "VARIABLES (Consumos)",
            "OBSERVACIONES (Consumos)"
        ], Colors.FASE4_BG),
        ("4.15 FEEDBACK: Compras/Calidad/Maquila/Costeo", [
            "CLASIFICACION SITUACIONES PROD (CCMC)", "VARIABLES (CCMC)",
            "AREA ENCARGADA", "OBSERVACIONES (CCMC)"
        ], Colors.FASE4_BG),
        ("4.16 AUDACES (Cierre)", [
            "CORRECCION REALIZADA", "SITUACIONES EN CORTE"
        ], Colors.FASE4_BG),
    ]
)

# ── COLUMNAS TRANSVERSALES (STATUS y ENTREGABLES) ──
add_phase(
    "📊 STATUS Y ENTREGABLES",
    "6C7A89",  # Gris azulado
    [
        ("STATUS GENERAL", [
            "STATUS", "STATUS TALLER", "STATUS BORDADO",
            "TERMINADO TALLER PARA COSTEO TELAS", "ESTADO PRENDA PLANTA PROD",
            "TIPO RECHAZO CALIDAD PLANTA", "FEEDBACK CALIDAD PLANTA"
        ], "B0BEC5"),
        ("ENTREGABLES POR AREA", [
            "ENTREGABLE CREATIVO OK?", "ENTREGABLE TECNICO (En Proceso)",
            "ENTREGABLE TECNICO (Completo) OK?", "ENTREGABLE TRAZADOR OK?",
            "ENVIO DE MOD ARTE OK?", "FECHA LIBERACION DISENO A INGENIERIA",
            "FECHA LIBERACION INGENIERIA A PRODUCCION"
        ], "B0BEC5"),
    ]
)

# ── TIME COLLECTION ──
add_phase(
    "⏱ TIME COLLECTION (Planificacion Temporal)",
    "1ABC9C",  # Verde turquesa
    [
        ("PRIORIDADES Y DIFICULTAD", [
            "PRIORIDADES FIRST BUY", "PRIORIDADES BORDADO",
            "PRIORIDADES TEXTIL STOCK", "DIFICULTAD PRENDA",
            "DIFICULTAD BORDADO", "TIRAS CONTINUAS SI/NO",
            "BORDADO (SEMIELABORADO/EN PRENDA)", "GRUPO/ESTILO"
        ], "8ED8C8"),
        ("COMENTARIOS EQUIPO", [
            "COMENTARIOS INGENIERIA (TIME)", "COMENTARIOS TRAZO (TIME)",
            "COMENTARIOS COSTEO", "SUGERENCIA MOD/UBC (TIME)",
            "REQUIERE MUESTRA SI/NO (TIME)", "ESCALADO MOLDERIA #1",
            "COMENTARIOS TELAS UBICACION/MOD"
        ], "8ED8C8"),
    ]
)

# ── PROCESOS EXTERNOS Y CORTE (Transversales) ──
add_phase(
    "🔧 DETALLE PROCESOS EXTERNOS Y CORTE",
    "8E44AD",  # Purpura
    [
        ("PROCESO EXTERNO DETALLE", [
            "TIPO PROCESO EXTERNO (Detalle)", "FECHA RECEPCION PIEZA",
            "FECHA ENTREGA PIEZA", "STATUS PROCESO EXTERNO"
        ], "D2B4DE"),
        ("CORTE CONTRAMUESTRAS (Historial)", [
            "FECHA CORTE #1", "TIPO CORTE #1", "FECHA CORTE #2",
            "TIPO CORTE #2", "FECHA CORTE #3", "TIPO CORTE #3",
            "FECHA CORTE #4", "TIPO CORTE #4", "OBSERVACIONES CORTE CM",
            "TOTAL CORTES PIEZAS", "TOTAL CORTES MUESTRAS"
        ], "D2B4DE"),
        ("CONFECCION CONTRAMUESTRA", [
            "MODISTA CM", "FECHA INICIO CONFECCION CM",
            "FECHA ENTREGA CONFECCION CM", "STATUS CONFECCION CM",
            "OBSERVACIONES CONFECCION MODISTA CM",
            "SITUACIONES CONFECCION DIS. TECNICO 1",
            "SITUACIONES CONFECCION DIS. TECNICO 2",
            "TIEMPO CONFECCION INGENIERIA (min)"
        ], "D2B4DE"),
    ]
)

# ── CALIDAD CONFECCION ──
add_phase(
    "✅ CALIDAD DE CONFECCION",
    "2ECC71",  # Verde esmeralda
    [
        ("CALIDAD PLANTA PRODUCCION", [
            "ESTADO PRENDA CALIDAD PLANTA", "TIPO RECHAZO CALIDAD",
            "FEEDBACK COMENTARIOS CALIDAD", "PROGRAMACION ZA"
        ], "A9DFBF"),
    ]
)

# ── EMPAQUES E INSUMOS ──
add_phase(
    "📦 EMPAQUES E INSUMOS ADICIONALES",
    "E67E22",  # Naranja oscuro
    [
        ("EMPAQUES", [
            "TIPO DE EMPAQUE", "MARQUILLA SATINADA FORRADA"
        ], "F0B27A"),
    ]
)

# ── CONSUMOS COSTEO ──
add_phase(
    "💰 CONSUMOS COSTEO",
    "C0392B",  # Rojo oscuro
    [
        ("VERIFICACION COSTEO", [
            "CONSUMO COSTEO SOLIDO", "CONSUMO COSTEO MOD ARTE",
            "CONSUMO COSTEO UBI TRAZO", "OBSERVACIONES COSTEO VERIF",
            "A RIESGO / VALIDADO POR LINA PENA"
        ], "F5B7B1"),
        ("REFERENTE CONSUMO ESTANDAR", [
            "REFERENTE CONSUMO ESTANDAR", "CONSUMO REFERENTE",
            "CONSUMO TECNICO (Est)", "CONSUMO TRAZADOR (Est)",
            "REFERENTE CONSUMO APROXIMADO", "VALIDADO", "BORDADO STATUS FISICO"
        ], "F5B7B1"),
    ]
)

# ── INSUMOS CON RECORRIDOS ──
add_phase(
    "🧵 INSUMOS, SESGOS Y RECORRIDOS",
    "5D6D7E",  # Gris pizarra
    [
        ("SESGOS Y TIRAS", [
            "SESGO LUCIR CONSUMO LINEAL", "SESGO FORRO CONSUMO LINEAL",
            "SESGO FUSIONABLE CONSUMO LINEAL", "ANCHO SESGO (cms)",
            "SENTIDO SESGO (AL HILO/TRAVES/SESGO)", "TIRAS CONTINUAS (DETALLE)"
        ], "ABB2B9"),
        ("INSUMOS NO TEXTILES", [
            "CODIGO INSUMO", "DESCRIPCION INSUMO", "UNIDAD INSUMO",
            "CANTIDAD USADA POR PRENDA"
        ], "ABB2B9"),
    ]
)

# ── INFO LOGISTICA ──
add_phase(
    "🚚 INFORMACION LOGISTICA",
    "2980B9",  # Azul marino
    [
        ("DATOS LOGISTICOS", [
            "ACTUALIZACION DE MODELO", "NOTA DE FABRICACION (Log)",
            "GESTION DE LA NOTA (Log)", "CODIGO DE BARRAS (Log)",
            "UBICACION (Log)", "REVISION CREATIVOS"
        ], "85C1E9"),
        ("ENTREGA PARCIAL", [
            "ENTREGA PARCIAL (Fecha Estimada)",
        ], "85C1E9"),
    ]
)

# ── REF PDT POR TERMINAR ──
add_phase(
    "⚠ PENDIENTES POR TERMINAR",
    "F39C12",  # Amarillo oscuro
    [
        ("REFERENCIAS PENDIENTES TALLER", [
            "REF TERMINADAS TALLER SIN CONSUMOS",
            "REF PDT POR TERMINAR TALLER SIN CONSUMOS",
            "CONSUMOS EQUIPO TECNICO INGRESADOS SAP",
            "CONSUMOS EQUIPO TECNICO SIN INGRESAR SAP (SARA)"
        ], "F9E79F"),
    ]
)

# ── METADATA ──
add_phase(
    "📋 METADATA DEL REGISTRO",
    "34495E",  # Azul grisaceo oscuro
    [
        ("CONTROL DE CAMBIOS", [
            "FECHA CREACION REGISTRO", "FECHA ULTIMA MODIFICACION",
            "USUARIO ULTIMA MODIFICACION", "VERSION ARCHIVO",
            "NOTAS INTERNAS"
        ], "7FB3D8"),
    ]
)

# ── Ajustar anchos de columna ──
for col in range(1, COL):
    ws.column_dimensions[get_column_letter(col)].width = 16

# ── Congelar paneles (filas 1-3 + columna A) ──
ws.freeze_panes = "B4"

# ── Agregar filtros en fila 3 ──
ws.auto_filter.ref = f"A3:{get_column_letter(COL-1)}3"

# ── Agregar fila de ejemplo (CASO_001 del JSON) ──
ejemplo = [
    # FASE 1: Perfilamiento
    1, "MD-001", "PT-00001", "Vestido Nuevo con Mod Arte", "ECRU/SAND",
    "1591", "DRESSES", "MAXI", "DS", "WS27",
    "WINTER SUN", "CAPSULA A", "Vestido", "SI",
    # Referentes
    "", "", "", "",
    # Caracteristicas
    "0-2-4-6-8-10-12", "MAXI 115CM", "BACK ZIPPER", "SI",
    "NO", "",
    # Disenador Creativo
    "Claudia", "2026-04-24",
    # Consumo Base
    "2.41 MTS", "NO", "2.16 MTS", "", "",
    "Estimacion inicial segun referente PT03402",
    # Moldería
    "2026-04-24", "2026-04-26", "Molderia base creada en talla 2", "SI", "NO",
    # FASE 2: Alistamiento
    "TELA LUCIR", "TE00008887", "SOLID LINEN JULIA 43003", "1.5 mts",
    "LINEN", "", "VANILA 1591", "NO", "", "NO",
    # Forro
    "", "", "", "", "", "",
    # Corte
    "Andres Cortador", "2026-04-28", "3.5 mts", "Corte sin novedades",
    # Confección
    "Isabel Valencia", "2026-04-29", "2026-04-30", "TERMINADO",
    "Confección OK", "Revisar margen fusionable en pretinas",
    # Procesos especiales
    "NO", "", "", "NO", "", "", "", "", "",
    # FASE 3: Medición
    "2026-05-02", "OK, revisar largo", "APROBADA CON COMENTARIOS",
    "2026-05-05", "Largo ajustado OK", "", "", "", "", "", "", "NO",
    # Ajustes Molderia
    "NO", "", "", "", "", "",
    # Disenador Técnico
    "Daniela Garcia", "Andrea Jacome", "SOLIDO", "Prenda sin novedades tecnicas",
    # Cálculo Consumo
    "2.09", "", "", "8", "UNIDAD", "Se trabajo ancho 1.39 por hallazgos de planta",
    # Trazos
    "Kelly Mitrovich", "2.03", "", "", "0.06", "2.87%",
    "Ajuste por encogimiento", "Trazo validado",
    # Catalogación
    "SI", "NO", "NO", "NO", "", "",
]

# Rellenar el resto de columnas con vacío para el ejemplo
ejemplo.extend([""] * (COL - 1 - len(ejemplo)))

for c in range(1, COL):
    ws.cell(row=4, column=c).value = ejemplo[c-1] if c-1 < len(ejemplo) else ""
    ws.cell(row=4, column=c).border = thin_border
    ws.cell(row=4, column=c).alignment = Alignment(horizontal="center", vertical="center")

# ── Altura de filas ──
ws.row_dimensions[1].height = 28
ws.row_dimensions[2].height = 22
ws.row_dimensions[3].height = 35
ws.row_dimensions[4].height = 20

# ── Guardar ──
output_path = r"C:\Users\jchacon\Documents\JEFERSON STUDY\Analisis de secuencia de coleccion\version 2.0\prueba.xlsx"
wb.save(output_path)
print(f"Archivo creado exitosamente: {output_path}")
print(f"Total columnas: {COL-1}")
